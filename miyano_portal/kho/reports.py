"""Báo cáo kho khách hàng: Nhập-Xuất-Tồn, thẻ kho, cảnh báo hạn dùng.

Tính trực tiếp từ `Customer Stock Ledger Entry` — nguồn sự thật duy nhất.
KHÔNG dùng `Customer Stock Lot Balance` cho hai báo cáo N-X-T và thẻ kho: đó
chỉ là cache TỒN HIỆN TẠI, không trả lời được câu hỏi lịch sử ("tồn đầu kỳ
ngày X là bao nhiêu"). Cache chỉ đúng vai cho cảnh báo hạn dùng — một câu hỏi
VỀ HIỆN TẠI ("lô nào đang tồn và sắp hết hạn"), nơi nó nhất quán với kho_ton().

Một bộ cột (COLUMNS) cho mỗi loại báo cáo, dùng CHUNG cho JSON trả về portal
và cho xuất Excel — theo round-tripping-spreadsheets: màn hình và file xuất
không được lệch cột. Nhãn cột ở đây PHẢI khớp từng chữ với mảng tương ứng
trong `frontend/src/kho-bao-cao-columns.js` — test_kho_reports.py khẳng định
điều đó bằng cách đọc lại chính file .js.

Dòng sổ có `da_dao=1` KHÔNG bị lọc bỏ khỏi bất kỳ tổng nào ở đây: đó là dòng
gốc đã bị đảo, còn dòng đảo (âm, chứng từ riêng, ngày = ngày huỷ) là một dòng
sổ khác, hoàn toàn hợp lệ và phải được cộng vào đúng kỳ của NÓ. Lọc theo
`da_dao` sẽ làm dòng gốc "biến mất" khỏi kỳ nó được ghi trong khi dòng đảo vẫn
còn, phá vỡ hằng đẳng thức tồn đầu + nhập − xuất = tồn cuối ngay lập tức.
"""

import io

import frappe
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from miyano_portal.kho import ledger, voucher

EPS = ledger.EPS

NXT_COLUMNS = [
	("Mã vật tư", "ma_vat_tu"),
	("Tên vật tư", "ten_vat_tu"),
	("ĐVT", "dvt"),
	("Tồn đầu - SL", "ton_dau_sl"),
	("Tồn đầu - Thành tiền", "ton_dau_tt"),
	("Nhập - SL", "nhap_sl"),
	("Nhập - Thành tiền", "nhap_tt"),
	("Xuất - SL", "xuat_sl"),
	("Xuất - Thành tiền", "xuat_tt"),
	("Tồn cuối - SL", "ton_cuoi_sl"),
	("Tồn cuối - Thành tiền", "ton_cuoi_tt"),
]

# Cùng tám cột số của NXT_COLUMNS, thêm Số lô / Hạn sử dụng, bớt Mã/Tên/ĐVT
# (đã cố định ở mức vật tư khi bung xuống lô — xem BaoCaoNXT.vue).
NXT_LOT_COLUMNS = [
	("Số lô", "so_lo"),
	("Hạn sử dụng", "han_su_dung"),
	("Tồn đầu - SL", "ton_dau_sl"),
	("Tồn đầu - Thành tiền", "ton_dau_tt"),
	("Nhập - SL", "nhap_sl"),
	("Nhập - Thành tiền", "nhap_tt"),
	("Xuất - SL", "xuat_sl"),
	("Xuất - Thành tiền", "xuat_tt"),
	("Tồn cuối - SL", "ton_cuoi_sl"),
	("Tồn cuối - Thành tiền", "ton_cuoi_tt"),
]

THE_KHO_COLUMNS = [
	("Ngày", "ngay"),
	("Số chứng từ", "chung_tu"),
	("Loại chứng từ", "loai_chung_tu"),
	("Đối tác / Nơi nhận", "doi_tac"),
	("Số lô", "so_lo"),
	("SL nhập", "sl_nhap"),
	("SL xuất", "sl_xuat"),
	("Tồn luỹ kế", "ton_luy_ke"),
]

CANH_BAO_COLUMNS = [
	("Mã vật tư", "ma_vat_tu"),
	("Tên vật tư", "ten_vat_tu"),
	("ĐVT", "dvt"),
	("Số lô", "so_lo"),
	("Hạn sử dụng", "han_su_dung"),
	("Số ngày còn lại", "so_ngay_con_lai"),
	("Số lượng tồn", "so_luong"),
	("Trạng thái", "trang_thai"),
]

# Gap 2 (review E4 phần B): hai bộ cột còn thiếu cho xuất Excel của Nhật ký
# vật tư và NXT theo đợt hàng — thêm vào đây để dùng CHUNG giữa
# nhat_ky_rows_export()/bao_cao_dot_rows() và kho_bao_cao_excel(), đúng
# nguyên tắc round-tripping-spreadsheets như ba bộ cột trên.
NHAT_KY_COLUMNS = [
	("Ngày", "ngay"),
	("Số phiếu", "phieu"),
	("Loại", "loai"),
	("Nguồn / NCC", "nguon"),
	("Đợt", "dot"),
	("Lô", "lo"),
	("Hạn dùng", "han"),
	("SL nhập", "sl_nhap"),
	("SL xuất", "sl_xuat"),
	("Đơn giá", "don_gia"),
	("Tồn sau giao dịch", "ton_sau"),
	("Người ghi sổ", "nguoi_ghi_so"),
]

DOT_COLUMNS = [
	("Đợt (phiếu nhập)", "dot"),
	("Ngày nhận", "ngay_nhan"),
	("Nguồn / NCC", "nguon"),
	("Chứng từ", "chung_tu"),
	("Vật tư", "vat_tu"),
	("Lô", "lo"),
	("Hạn dùng", "han_su_dung"),
	("SL nhập", "sl_nhap"),
	("Giá trị nhập", "gia_tri_nhap"),
	("Đã xuất", "da_xuat"),
	("Còn lại", "con_lai"),
	("Tuổi tồn (ngày)", "tuoi_ton_ngay"),
	("% tiêu thụ", "pct_tieu_thu"),
	("Chậm luân chuyển", "cham_luan_chuyen"),
]

_CHUNG_TU_LABEL = {
	"Customer Stock Receipt": "Phiếu nhập",
	"Customer Stock Issue": "Phiếu xuất",
}


def _r(v) -> float:
	"""Làm tròn rác dấu phẩy động về 0 (cùng ngưỡng EPS với ledger.py)."""
	v = float(v or 0)
	return 0.0 if abs(v) < EPS else round(v, 6)


def _new_bucket() -> dict:
	return {
		"ton_dau_sl": 0.0, "ton_dau_tt": 0.0,
		"nhap_sl": 0.0, "nhap_tt": 0.0,
		"xuat_sl": 0.0, "xuat_tt": 0.0,
	}


def _close(bucket: dict) -> dict:
	"""Tồn cuối LUÔN được TÍNH, không bao giờ đọc từ nơi khác (kể cả cache):
	đó là điều duy nhất khiến tồn đầu + nhập − xuất = tồn cuối đúng theo cấu
	trúc, không phải nhờ hai phép tính độc lập tình cờ khớp nhau."""
	out = dict(bucket)
	out["ton_cuoi_sl"] = out["ton_dau_sl"] + out["nhap_sl"] - out["xuat_sl"]
	out["ton_cuoi_tt"] = out["ton_dau_tt"] + out["nhap_tt"] - out["xuat_tt"]
	return {k: _r(v) for k, v in out.items()}


def _vat_tu_info(kho: str) -> dict:
	"""Tra cứu tên/ĐVT cho MỌI vật tư từng có trong kho, kể cả đã `active=0`:
	một báo cáo lịch sử phải hiển thị được vật tư đã ngừng dùng nếu nó có
	phát sinh trong kỳ — khác với kho_vat_tu_list() (chỉ vật tư đang dùng,
	dùng để CHỌN khi lập phiếu mới)."""
	rows = frappe.get_all(
		"Customer Warehouse Item",
		filters={"kho": kho},
		fields=["name", "ma_vat_tu", "ten_vat_tu", "dvt"],
	)
	return {r["name"]: r for r in rows}


def ton_hien_tai_rows(kho: str, tim: str | None = None) -> list[dict]:
	"""Tồn hiện tại của MỘT kho, gộp các lô về một dòng cho mỗi vật tư.

	Chuyển từ `api/kho.py::kho_ton()` (Phase 2) sang đây nguyên trạng — cùng
	một phép gộp, giờ dùng chung bởi cả `kho_ton()` (portal, một kho suy từ
	phiên đăng nhập) và `kho.desk_reports.ton_kho_khach_hang_rows()` (desk,
	Phase 6, lặp qua NHIỀU kho rồi gọi lại đúng hàm này cho từng kho). Không
	được viết lại phép cộng này lần thứ hai ở bất cứ đâu khác.

	Đọc từ `Customer Stock Lot Balance` — đây là câu hỏi VỀ HIỆN TẠI, giống
	`canh_bao_han_rows()`, không phải lịch sử."""
	lots = frappe.get_all(
		"Customer Stock Lot Balance",
		filters={"kho": kho, "so_luong": [">", EPS]},
		fields=["vat_tu", "so_lo", "han_su_dung", "so_luong", "gia_tri"],
	)

	gop = {}
	for lot in lots:
		g = gop.setdefault(lot["vat_tu"], {
			"vat_tu": lot["vat_tu"], "so_luong": 0.0, "gia_tri": 0.0,
			"so_lo_count": 0, "han_gan_nhat": None,
		})
		g["so_luong"] += float(lot["so_luong"])
		g["gia_tri"] += float(lot["gia_tri"] or 0)
		g["so_lo_count"] += 1
		han = lot["han_su_dung"]
		if han and (g["han_gan_nhat"] is None or han < g["han_gan_nhat"]):
			g["han_gan_nhat"] = han

	out = []
	for vat_tu, g in gop.items():
		vt = frappe.db.get_value(
			"Customer Warehouse Item", vat_tu,
			["ma_vat_tu", "ten_vat_tu", "dvt", "item_code"], as_dict=True,
		)
		if not vt:
			continue
		if tim:
			hay = f"{vt.ma_vat_tu} {vt.ten_vat_tu}".lower()
			if tim.lower() not in hay:
				continue
		out.append({**g, **{
			"ma_vat_tu": vt.ma_vat_tu, "ten_vat_tu": vt.ten_vat_tu,
			"dvt": vt.dvt, "item_code": vt.item_code or "",
		}})
	return sorted(out, key=lambda r: r["ten_vat_tu"])


def nxt_data(kho: str, tu_ngay, den_ngay) -> dict:
	"""Gộp sổ kho theo từng vật tư VÀ theo từng lô trong một lượt quét.

	Trả dict thô {"items": {...}, "lots": {vat_tu: {so_lo: {...}}}, "lot_han":
	{vat_tu: {so_lo: han}}} — nxt_item_rows()/nxt_lot_rows() tự chọn mức và tự
	lọc/enrich. Một lượt quét duy nhất cho cả hai mức để bung xuống lô không
	phải quét lại sổ lần thứ hai.
	"""
	tu = frappe.utils.getdate(tu_ngay)
	den = frappe.utils.getdate(den_ngay)
	if tu > den:
		frappe.throw("Từ ngày phải trước hoặc bằng Đến ngày.", frappe.ValidationError)

	entries = frappe.get_all(
		"Customer Stock Ledger Entry",
		filters={"kho": kho, "ngay": ["<=", den]},
		fields=["vat_tu", "so_lo", "han_su_dung", "ngay", "so_luong", "gia_tri"],
	)

	per_item: dict[str, dict] = {}
	per_lot: dict[str, dict] = {}
	lot_han: dict[str, dict] = {}

	for e in entries:
		vt = e["vat_tu"]
		lo = e["so_lo"]
		ngay = frappe.utils.getdate(e["ngay"])
		sl = float(e["so_luong"])
		gt = float(e["gia_tri"] or 0)

		item_b = per_item.setdefault(vt, _new_bucket())
		lot_b = per_lot.setdefault(vt, {}).setdefault(lo, _new_bucket())
		if e.get("han_su_dung"):
			lot_han.setdefault(vt, {}).setdefault(lo, e["han_su_dung"])

		# Phân vùng TOÀN PHẦN theo (trước kỳ) / (nhập, so_luong>0) / (xuất,
		# còn lại): một dòng so_luong==0 (không xảy ra trong thực tế nhưng
		# không được để lọt qua cả hai nhánh) rơi vào "xuất" với biên độ 0 —
		# không mất, không đếm hai lần.
		for b in (item_b, lot_b):
			if ngay < tu:
				b["ton_dau_sl"] += sl
				b["ton_dau_tt"] += gt
			elif sl > 0:
				b["nhap_sl"] += sl
				b["nhap_tt"] += gt
			else:
				b["xuat_sl"] += -sl
				b["xuat_tt"] += -gt

	return {
		"items": {vt: _close(b) for vt, b in per_item.items()},
		"lots": {
			vt: {lo: _close(b) for lo, b in lots.items()}
			for vt, lots in per_lot.items()
		},
		"lot_han": lot_han,
	}


def nxt_item_rows(kho: str, tu_ngay, den_ngay, tim: str | None = None) -> list[dict]:
	"""Một dòng cho mỗi vật tư CÓ PHÁT SINH TRONG SỔ tính tới den_ngay — tức
	là có dòng sổ trước tu_ngay (tồn đầu khác 0 hoặc từng khác 0) HOẶC có dòng
	sổ trong kỳ. Vật tư tồn cuối bằng 0 do vừa nhập vừa xuất hết trong kỳ VẪN
	xuất hiện — không lọc theo tồn cuối ở bất cứ đâu trong hàm này."""
	data = nxt_data(kho, tu_ngay, den_ngay)
	info = _vat_tu_info(kho)
	out = []
	for vt, closed in data["items"].items():
		meta = info.get(vt)
		if not meta:
			continue
		if tim:
			hay = f"{meta['ma_vat_tu']} {meta['ten_vat_tu']}".lower()
			if tim.lower() not in hay:
				continue
		out.append({
			"vat_tu": vt, "ma_vat_tu": meta["ma_vat_tu"],
			"ten_vat_tu": meta["ten_vat_tu"], "dvt": meta["dvt"],
			**closed,
		})
	return sorted(out, key=lambda r: r["ten_vat_tu"])


def nxt_lot_rows(kho: str, vat_tu: str, tu_ngay, den_ngay) -> list[dict]:
	"""Bung một vật tư xuống mức lô cho CÙNG khoảng ngày. Tổng của các dòng
	trả về ở đây, trên cả tám cột số, PHẢI bằng đúng dòng vật tư tương ứng của
	nxt_item_rows() — cả hai đọc từ cùng một lượt quét nxt_data()."""
	data = nxt_data(kho, tu_ngay, den_ngay)
	lots = data["lots"].get(vat_tu, {})
	han = data["lot_han"].get(vat_tu, {})
	out = []
	for lo, closed in lots.items():
		out.append({"so_lo": lo, "han_su_dung": han.get(lo), **closed})
	return sorted(
		out, key=lambda r: (r["han_su_dung"] is None, r["han_su_dung"] or "", r["so_lo"])
	)


def the_kho_rows(kho: str, vat_tu: str, tu_ngay, den_ngay) -> list[dict]:
	"""Thẻ kho của một vật tư: mọi dòng sổ trong [tu_ngay, den_ngay], theo thứ
	tự thời gian, với cột tồn luỹ kế bắt đầu từ đúng tồn đầu kỳ (tổng sổ trước
	tu_ngay) — không bắt đầu từ 0, nếu không con số sẽ là hư cấu với mọi kỳ
	không mở đúng từ đầu lịch sử.

	Sắp theo `ngay asc, creation asc, name asc`: `ngay` một mình không đủ duy
	nhất (nhiều chứng từ cùng ngày), và trường `sort_field` mặc định của
	doctype (`creation`) một mình lại sai thứ tự khi có phiếu ghi lùi ngày —
	đúng lý do rebuild_lot_balance() đã ghi trong docstring của nó.
	"""
	tu = frappe.utils.getdate(tu_ngay)
	den = frappe.utils.getdate(den_ngay)
	if tu > den:
		frappe.throw("Từ ngày phải trước hoặc bằng Đến ngày.", frappe.ValidationError)

	# SUM() ở SQL thay vì kéo cả lịch sử-trước-kỳ về Python rồi sum() (M-5,
	# review E4 phần B) — một kho lâu năm có thể có hàng nghìn dòng trước
	# tu_ngay, không cần rời DB để cộng một cột số.
	balance = frappe.utils.flt(frappe.db.get_value(
		"Customer Stock Ledger Entry",
		{"kho": kho, "vat_tu": vat_tu, "ngay": ["<", tu]},
		"sum(so_luong)",
	))

	entries = frappe.get_all(
		"Customer Stock Ledger Entry",
		filters={"kho": kho, "vat_tu": vat_tu, "ngay": ["between", [tu, den]]},
		fields=["name", "ngay", "so_lo", "so_luong", "chung_tu_type", "chung_tu", "creation"],
		order_by="ngay asc, creation asc, name asc",
	)

	receipt_names = {e["chung_tu"] for e in entries if e["chung_tu_type"] == "Customer Stock Receipt"}
	issue_names = {e["chung_tu"] for e in entries if e["chung_tu_type"] == "Customer Stock Issue"}
	nguoi_giao = {}
	if receipt_names:
		nguoi_giao = dict(frappe.get_all(
			"Customer Stock Receipt", filters={"name": ["in", list(receipt_names)]},
			fields=["name", "nguoi_giao"], as_list=True,
		))
	noi_nhan = {}
	if issue_names:
		noi_nhan = {
			r["name"]: r for r in frappe.get_all(
				"Customer Stock Issue", filters={"name": ["in", list(issue_names)]},
				fields=["name", "noi_nhan", "nguoi_nhan"],
			)
		}

	out = []
	for e in entries:
		sl = float(e["so_luong"])
		balance += sl
		if e["chung_tu_type"] == "Customer Stock Receipt":
			doi_tac = nguoi_giao.get(e["chung_tu"]) or ""
		else:
			r = noi_nhan.get(e["chung_tu"]) or {}
			doi_tac = " / ".join(x for x in [r.get("noi_nhan"), r.get("nguoi_nhan")] if x)
		out.append({
			"ngay": e["ngay"],
			"chung_tu": e["chung_tu"],
			"loai_chung_tu": _CHUNG_TU_LABEL.get(e["chung_tu_type"], e["chung_tu_type"]),
			"doi_tac": doi_tac,
			"so_lo": e["so_lo"],
			"sl_nhap": _r(sl) if sl > 0 else 0.0,
			"sl_xuat": _r(-sl) if sl < 0 else 0.0,
			"ton_luy_ke": _r(balance),
		})
	return out


_NHAT_KY_TRANG = 50


def _nhat_ky_filtered_rows(
	kho: str, vat_tu: str, tu_ngay, den_ngay,
	so_lo: str | None = None, loai: str | None = None, nguon: str | None = None,
	khoa_phong: str | None = None,
) -> list[dict]:
	"""Toàn bộ dòng nhật ký khớp bộ lọc, theo thứ tự thời gian, CHƯA phân
	trang — dùng chung cho cả `nhat_ky_rows()` (màn hình, 50 dòng/trang) và
	`nhat_ky_rows_export()` (Excel, NL-8.3: bắt buộc chọn kỳ nhưng KHÔNG giới
	hạn 50 dòng như màn hình — một khách xuất báo cáo quý phải lấy đủ cả quý,
	không phải trang đầu). Một phép tính, hai cách tiêu thụ, đúng khuôn
	NXT_COLUMNS/NXT_LOT_COLUMNS dùng chung nxt_data() ở trên.

	Cột "Tồn sau giao dịch" là tồn luỹ kế CHẠY QUA TOÀN BỘ dòng sổ trong kỳ,
	theo ĐÚNG thứ tự thời gian mà the_kho_rows() dùng (ngay asc, creation asc,
	name asc) — cùng một phép cộng dồn, không viết lại lần hai. Luỹ kế được
	tính TRƯỚC khi áp các bộ lọc hiển thị (so_lo/loai/nguon): các bộ lọc đó chỉ
	chọn DÒNG NÀO được hiển thị, không đổi Ý NGHĨA của cột tồn — một khách lọc
	theo một lô vẫn phải thấy đúng tồn CỦA CẢ VẬT TƯ tại thời điểm đó, giống
	một sao kê ngân hàng vẫn hiện số dư thật dù đang lọc theo loại giao dịch.

	Bất biến (test bắt được): khi kỳ phủ hết lịch sử và không lọc gì, tồn sau
	giao dịch của dòng CUỐI CÙNG phải bằng đúng tồn hiện tại của vật tư theo
	`ton_hien_tai_rows()`/`kho_ton` — hai đường tính độc lập (sổ luỹ kế vs.
	cache `Customer Stock Lot Balance`) phải khớp nhau.

	Dòng `da_dao=1` (phiếu gốc đã bị đảo) VẪN xuất hiện, mờ đi phía client —
	sổ là chứng từ, giấu dòng đã đảo là làm sai lệch bản ghi (BR-D2).

	"Đợt" (BR-D1: đợt = một phiếu nhập) chỉ có nghĩa cho dòng NHẬP — dòng nhập
	mang chính tên phiếu của nó làm đợt. Dòng XUẤT không gắn với một đợt duy
	nhất (xem `bao_cao_dot_rows()`: một lần xuất có thể ăn vào nhiều đợt khác
	nhau qua phân bổ FIFO phân tích) nên cột "đợt"/"nguồn" để trống ở dòng xuất
	— gán một đợt xem như của dòng xuất sẽ là suy diễn.
	"""
	tu = frappe.utils.getdate(tu_ngay)
	den = frappe.utils.getdate(den_ngay)
	if tu > den:
		frappe.throw("Từ ngày phải trước hoặc bằng Đến ngày.", frappe.ValidationError)

	# SUM() ở SQL thay vì kéo cả lịch sử-trước-kỳ về Python rồi sum() (M-5,
	# review E4 phần B) — một kho lâu năm có thể có hàng nghìn dòng trước
	# tu_ngay, không cần rời DB để cộng một cột số.
	balance = frappe.utils.flt(frappe.db.get_value(
		"Customer Stock Ledger Entry",
		{"kho": kho, "vat_tu": vat_tu, "ngay": ["<", tu]},
		"sum(so_luong)",
	))

	entries = frappe.get_all(
		"Customer Stock Ledger Entry",
		filters={"kho": kho, "vat_tu": vat_tu, "ngay": ["between", [tu, den]]},
		fields=["name", "ngay", "so_lo", "han_su_dung", "so_luong", "don_gia",
		        "chung_tu_type", "chung_tu", "da_dao", "owner", "creation"],
		order_by="ngay asc, creation asc, name asc",
	)

	receipt_names = {e["chung_tu"] for e in entries if e["chung_tu_type"] == "Customer Stock Receipt"}
	receipts = {}
	if receipt_names:
		receipts = {
			r["name"]: r for r in frappe.get_all(
				"Customer Stock Receipt", filters={"name": ["in", list(receipt_names)]},
				fields=["name", "loai_nhap", "ncc", "phieu_goc"],
			)
		}
	ncc_names = {r["ncc"] for r in receipts.values() if r.get("ncc")}
	ncc_ten = {}
	if ncc_names:
		ncc_ten = dict(frappe.get_all(
			"Customer Supplier", filters={"name": ["in", list(ncc_names)]},
			fields=["name", "ten_ncc"], as_list=True,
		))

	# Đầu mối để biết một dòng XUẤT có phải bút toán bù trừ (I-1, review E4
	# phần B) hay không — chỉ cần loai_xuat, không cần phieu_goc/ncc như bên
	# nhập vì nguồn/đợt của dòng xuất vốn đã luôn để trống.
	issue_names = {e["chung_tu"] for e in entries if e["chung_tu_type"] == "Customer Stock Issue"}
	issues = {}
	if issue_names:
		issues = {
			r["name"]: r for r in frappe.get_all(
				"Customer Stock Issue", filters={"name": ["in", list(issue_names)]},
				fields=["name", "loai_xuat", "khoa_phong"],
			)
		}

	all_rows = []
	for e in entries:
		sl = float(e["so_luong"])
		balance += sl
		is_receipt = e["chung_tu_type"] == "Customer Stock Receipt"
		nguon_dong = ""
		dot = ""
		la_dao = False
		if is_receipt:
			rc = receipts.get(e["chung_tu"]) or {}
			la_dao = rc.get("loai_nhap") == voucher.LOAI_DAO
			if la_dao:
				# I-1 (review E4 phần B): `_tao_phieu_dao()` không copy `ncc`
				# sang phiếu đảo, nên tra receipts[...] cho CHÍNH dòng đảo sẽ
				# luôn rơi vào nhánh else bên dưới và gán nhầm "Miyano" cho
				# mọi lần huỷ phiếu Mua ngoài — quy hàng của NCC cho Miyano.
				# Dòng đảo không phải một đợt hàng thật (xem bao_cao_dot_rows,
				# nơi chính module này loại nó khỏi danh sách đợt) nên không
				# gán "nguồn" cho nó; `dot` trỏ NGƯỢC về đợt gốc mà nó bù trừ
				# (`phieu_goc`) để người đọc lần được, thay vì mang chính tên
				# phiếu đảo — cái tên đó không phải một mã đợt.
				dot = rc.get("phieu_goc") or ""
			else:
				dot = e["chung_tu"]
				if rc.get("loai_nhap") == "Mua ngoài (NCC khác)" and rc.get("ncc"):
					nguon_dong = ncc_ten.get(rc["ncc"], rc["ncc"])
				else:
					nguon_dong = "Miyano"
		else:
			iss = issues.get(e["chung_tu"]) or {}
			la_dao = iss.get("loai_xuat") == voucher.LOAI_DAO
		all_rows.append({
			"ngay": e["ngay"],
			"phieu": e["chung_tu"],
			"loai": "Nhập" if is_receipt else "Xuất",
			"nguon": nguon_dong,
			"dot": dot,
			"lo": e["so_lo"],
			"han": e["han_su_dung"],
			"sl_nhap": _r(sl) if sl > 0 else 0.0,
			"sl_xuat": _r(-sl) if sl < 0 else 0.0,
			"don_gia": float(e["don_gia"] or 0),
			"ton_sau": _r(balance),
			"nguoi_ghi_so": e["owner"],
			"da_dao": bool(e["da_dao"]),
			# `da_dao` đánh dấu dòng GỐC đã bị đảo; `la_dao` đánh dấu CHÍNH
			# dòng này LÀ bút toán bù trừ — hai cờ khác nhau và cố tình không
			# gộp làm một: một cặp huỷ luôn có đúng một dòng mang mỗi cờ.
			# Trước bản sửa này dòng đảo không mang dấu hiệu nào, hiện ra như
			# một giao dịch bình thường — client cần cả hai để tô đúng UI.
			"la_dao": la_dao,
			# E8/US-E8.4: khoa phòng nhận, để nhật ký lọc được theo khoa
			# (chỉ có nghĩa cho dòng XUẤT — dòng nhập luôn rỗng, không suy
			# diễn gán khoa cho một lượt nhập hàng).
			"khoa_phong": (issues.get(e["chung_tu"]) or {}).get("khoa_phong") if not is_receipt else None,
		})

	filtered = all_rows
	if so_lo:
		filtered = [r for r in filtered if r["lo"] == so_lo]
	if loai:
		filtered = [r for r in filtered if r["loai"] == loai]
	if nguon:
		filtered = [r for r in filtered if r["nguon"] == nguon]
	if khoa_phong:
		filtered = [r for r in filtered if r["khoa_phong"] == khoa_phong]
	return filtered


def nhat_ky_rows(
	kho: str, vat_tu: str, tu_ngay, den_ngay,
	so_lo: str | None = None, loai: str | None = None,
	nguon: str | None = None, trang: int = 1, khoa_phong: str | None = None,
	so_dong_moi_trang: int = _NHAT_KY_TRANG,
) -> dict:
	"""Nhật ký vật tư (US-E4.6, UC-43, BR-D2) — bản MÀN HÌNH, phân trang server
	(mặc định 50 dòng, brief 2026-08-15 cho phép khách chọn 10/20/50 qua
	PhanTrang.vue — `so_dong_moi_trang` thay cho hằng số cứng cũ). Phép
	tính thật nằm ở `_nhat_ky_filtered_rows()`; xem docstring ở đó cho các
	bất biến (đối chiếu kho_ton, dòng da_dao không bị giấu, luỹ kế tính
	trước khi lọc). `khoa_phong` (E8/US-E8.4) là lọc HIỂN THỊ như
	so_lo/loai/nguon — không đổi ý nghĩa cột tồn luỹ kế."""
	filtered = _nhat_ky_filtered_rows(
		kho, vat_tu, tu_ngay, den_ngay, so_lo, loai, nguon, khoa_phong
	)
	trang = max(1, int(trang or 1))
	so_dong_moi_trang = max(1, int(so_dong_moi_trang or _NHAT_KY_TRANG))
	start = (trang - 1) * so_dong_moi_trang
	return {
		"tong_dong": len(filtered),
		"trang": trang,
		"so_dong_moi_trang": so_dong_moi_trang,
		"dong": filtered[start:start + so_dong_moi_trang],
	}


def nhat_ky_rows_export(
	kho: str, vat_tu: str, tu_ngay, den_ngay,
	so_lo: str | None = None, loai: str | None = None, nguon: str | None = None,
	khoa_phong: str | None = None,
) -> list[dict]:
	"""Bản XUẤT EXCEL của nhật ký vật tư (Gap 2, review E4 phần B) — CÙNG bộ
	lọc và CÙNG phép tính với màn hình (`_nhat_ky_filtered_rows()`), nhưng
	KHÔNG cắt theo trang: NL-8.3 chỉ bắt buộc chọn kỳ khi xuất, không giới
	hạn số dòng — 50 dòng/trang là giới hạn HIỂN THỊ, không phải giới hạn dữ
	liệu."""
	return _nhat_ky_filtered_rows(
		kho, vat_tu, tu_ngay, den_ngay, so_lo, loai, nguon, khoa_phong
	)


def canh_bao_han_rows(kho: str, so_ngay: int = 90) -> list[dict]:
	"""Lô đã hết hạn (còn tồn) VÀ lô hết hạn trong `so_ngay` ngày tới, gần nhất
	trước — một sắp xếp ASCENDING duy nhất theo `han_su_dung` cho cả hai nhóm
	đã đủ: nhóm đã hết hạn (ngày trong quá khứ) luôn đứng trước nhóm sắp hết
	hạn (ngày trong tương lai), và trong mỗi nhóm ngày gần `hôm nay` nhất lên
	đầu — hết hạn lâu nhất trước (cấp bách nhất), sắp hết hạn sớm nhất trước.

	Lô KHÔNG có `han_su_dung` (US-E4.8, VĐ-2) rơi vào một nhóm riêng "Không có
	hạn dùng", xếp sau mọi lô có hạn, KHÔNG được tính vào "Đã hết hạn"/"Sắp
	hết hạn". Trước bản sửa này, lọc thẳng bằng
	`frappe.get_all(..., filters={"han_su_dung": ["<=", han_toi]})` khiến
	Frappe tự bọc điều kiện đó bằng `coalesce(han_su_dung, '0001-01-01') <=
	han_toi` (xem `frappe.model.db_query.DatabaseQuery.prepare_filter_condition`)
	— NULL luôn thoả `<= han_toi`, nên MỌI lô không khai hạn bị kéo vào kết
	quả, rồi `frappe.utils.getdate(None)` lại âm thầm trả về NGÀY HÔM NAY,
	khiến các lô đó hiện ra là "Sắp hết hạn" với 0 ngày còn lại. Đọc TOÀN BỘ
	lô còn tồn của kho (không đưa `han_su_dung` vào filter) rồi tự phân loại
	ở Python là cách duy nhất tránh được lớp ifnull/coalesce đó.

	Đọc từ `Customer Stock Lot Balance` — cố ý, xem docstring đầu file: đây là
	câu hỏi VỀ HIỆN TẠI, không phải lịch sử.
	"""
	today = frappe.utils.getdate(frappe.utils.today())
	han_toi = frappe.utils.add_days(today, int(so_ngay))
	lots = frappe.get_all(
		"Customer Stock Lot Balance",
		filters={"kho": kho, "so_luong": [">", EPS]},
		fields=["vat_tu", "so_lo", "han_su_dung", "so_luong"],
	)
	info = _vat_tu_info(kho)
	out = []
	for lot in lots:
		meta = info.get(lot["vat_tu"]) or {}
		base = {
			"vat_tu": lot["vat_tu"],
			"ma_vat_tu": meta.get("ma_vat_tu", ""),
			"ten_vat_tu": meta.get("ten_vat_tu", ""),
			"dvt": meta.get("dvt", ""),
			"so_lo": lot["so_lo"],
			"so_luong": _r(lot["so_luong"]),
		}
		if not lot["han_su_dung"]:
			out.append({
				**base,
				"han_su_dung": None,
				"so_ngay_con_lai": None,
				"trang_thai": "Không có hạn dùng",
			})
			continue
		han = frappe.utils.getdate(lot["han_su_dung"])
		if han > han_toi:
			continue
		out.append({
			**base,
			"han_su_dung": han,
			"so_ngay_con_lai": (han - today).days,
			"trang_thai": "Đã hết hạn" if han < today else "Sắp hết hạn",
		})
	return sorted(
		out, key=lambda r: (r["han_su_dung"] is None, r["han_su_dung"] or today, r["so_lo"])
	)


_NGUONG_CHAM_MAC_DINH = 90  # khớp default trong 20_DataDict.md §1.3 / meta DocType


def _nguong_cham_luan_chuyen() -> int:
	"""Ngưỡng chậm luân chuyển (ngày). Field CHƯA TỪNG được cấu hình = mặc
	định `_NGUONG_CHAM_MAC_DINH` (90); field được cấu hình `<= 0` = "không áp
	ngưỡng" (một lựa chọn nghiệp vụ hợp lệ, khác "chưa cấu hình").

	KHÔNG dùng `get_single_value` ở đây (C-1, review E4 phần B — sửa LẠI một
	lần nữa sau khi bản vá đầu tiên của C-1 vẫn sai): `get_single_value` LUÔN
	ép giá trị qua `cast_fieldtype(df.fieldtype, value)` trước khi trả về —
	với field `Int`, `cast_fieldtype` gọi `cint(value)`, mà `cint(None) == 0`.
	Nghĩa là "chưa từng có dòng nào trong tabSingles" (chưa cấu hình) và "có
	dòng, giá trị 0" (cố ý tắt ngưỡng) đi qua CÙNG MỘT ĐƯỜNG và ra CÙNG MỘT
	SỐ 0 — không có cách nào phân biệt hai trạng thái đó bằng
	`get_single_value`, dù kiểm `gia_tri in (None, "")` sau đó trông có vẻ
	đúng (đo trực nghiệm: kiểm tra đó không bao giờ đúng, vì `gia_tri` đã là
	`0` — một int — TRƯỚC KHI tới được phép so sánh). Đây chính là lý do
	`get_singles_dict(..., cast=False)` (mặc định) là lựa chọn ĐÚNG: nó trả
	dict THÔ, không ép kiểu, chỉ chứa field NÀO THẬT SỰ có dòng trong
	`tabSingles` — field chưa từng lưu đơn giản KHÔNG CÓ MẶT trong dict, phân
	biệt được với field có dòng nhưng giá trị chuỗi `"0"`.

	Không mất tính "phòng thủ hai lớp": patch `v1_6.seed_portal_settings_defaults`
	seed 90 xuống `tabSingles` một lần khi migrate; hàm này vẫn tự rơi về
	`_NGUONG_CHAM_MAC_DINH` cho site/DB test chưa chạy patch đó.
	"""
	raw = frappe.db.get_singles_dict("Miyano Portal Settings").get("nguong_cham_luan_chuyen_ngay")
	if raw in (None, ""):
		return _NGUONG_CHAM_MAC_DINH
	return frappe.utils.cint(raw)


def bao_cao_dot_rows(
	kho: str, tu_ngay, den_ngay, vat_tu: str | None = None, nguon: str | None = None,
) -> list[dict]:
	"""NXT theo đợt hàng, phân bổ FIFO (US-E4.7, UC-44, BR-D1/D3).

	"Đợt" = một phiếu nhập ĐÃ GHI SỔ (BR-D1); mã đợt chính là tên phiếu. Mức
	theo dõi vật lý là (vật tư, lô): khi một lô được nhận qua nhiều đợt, số đã
	xuất được phân bổ cho đợt CŨ NHẤT trước (FIFO) trong phạm vi lô đó — quy
	ước PHÂN TÍCH, không phải bút toán, sổ kho không đổi.

	NL-8.2 (đợt có phiếu bị đảo): SL nhập của đợt phải TRỪ ĐÚNG phần đã đảo
	TÍNH TỚI `den_ngay` — không phải trạng thái `da_dao` HIỆN TẠI của phiếu
	(I-2, review E4 phần B). `da_dao` là một cờ KHÔNG CÓ NGÀY: nó đúng cho
	"báo cáo chạy hôm nay", nhưng một báo cáo dựng lại cho MỘT KỲ ĐÃ ĐÓNG
	(ví dụ in lại báo cáo tháng 6 vào tháng 8, sau khi một đợt của tháng 6 bị
	huỷ vào tháng 8) sẽ bị chính bút toán huỷ đó viết lại lịch sử: đợt biến
	mất khỏi một báo cáo mà, TẠI THỜI ĐIỂM `den_ngay` của báo cáo đó, đợt vẫn
	còn nguyên — vi phạm đúng nguyên tắc "báo cáo lịch sử phải bất biến" mà
	`the_kho_rows()`/`nxt_data()` (đầu file) đã đặt ra cho sổ chính, giờ áp
	dụng cho lớp phân tích theo đợt này.

	Sửa: lấy SL nhập GỐC của đợt (không lọc `da_dao`, giống hệt cách sổ NHẬP
	chính vẫn cộng dòng gốc — xem docstring đầu file), rồi TRỪ ĐI đúng phần đã
	đảo mà chính bút toán đảo đó có `ngay <= den_ngay` (tra qua
	`Customer Stock Receipt.phieu_goc`, vì phiếu đảo là một CHỨNG TỪ RIÊNG,
	không cùng `chung_tu` với đợt gốc). Một đợt bị huỷ SAU `den_ngay` vẫn hiện
	đủ SL nhập gốc; bị huỷ TRƯỚC/ĐÚNG `den_ngay` thì ròng về 0 (huỷ luôn đảo
	TOÀN BỘ phiếu — Frappe không có huỷ một phần) và bị loại khỏi danh sách
	(`sl_nhap_rong <= EPS`), không phải hiện một dòng 0.

	Với sổ XUẤT thì đơn giản hơn — CỘNG mọi dòng bất kể `da_dao`, không cần
	tra theo `phieu_goc`: một phiếu xuất bị huỷ để lại dòng gốc âm VÀ dòng đảo
	dương của CHÍNH NÓ, cả hai đều đã có `ngay <= den` được lọc thẳng ở SQL
	(dòng đảo mang `ngay` = ngày huỷ thật, không phải ngày phiếu gốc) — cộng
	cả hai cho ra đúng 0 tại đúng mốc thời gian nó xảy ra, tự triệt tiêu theo
	cấu trúc mà không cần biết đợt nào.

	`tu_ngay`/`den_ngay` là KỲ hiển thị: đợt được hiển thị khi ngày nhận nằm
	trong kỳ, nhưng chuỗi phân bổ FIFO và tổng đã xuất vẫn phải chạy qua MỌI
	đợt (kể cả trước `tu_ngay`) và mọi lượt xuất tới `den_ngay` — một đợt cũ
	hơn tu_ngay vẫn tiêu thụ trước theo đúng nguyên tắc FIFO dù nó không được
	vẽ ra màn hình. "Ngày báo cáo" dùng để tính tuổi tồn chính là `den_ngay`.
	"""
	tu = frappe.utils.getdate(tu_ngay)
	den = frappe.utils.getdate(den_ngay)
	if tu > den:
		frappe.throw("Từ ngày phải trước hoặc bằng Đến ngày.", frappe.ValidationError)

	filters_nhap = {
		"kho": kho, "chung_tu_type": "Customer Stock Receipt",
		"so_luong": [">", EPS], "ngay": ["<=", den],
	}
	if vat_tu:
		filters_nhap["vat_tu"] = vat_tu
	nhap_entries = frappe.get_all(
		"Customer Stock Ledger Entry", filters=filters_nhap,
		fields=["chung_tu", "vat_tu", "so_lo", "han_su_dung", "ngay", "so_luong", "gia_tri"],
	)

	filters_xuat = {"kho": kho, "chung_tu_type": "Customer Stock Issue", "ngay": ["<=", den]}
	if vat_tu:
		filters_xuat["vat_tu"] = vat_tu
	xuat_entries = frappe.get_all(
		"Customer Stock Ledger Entry", filters=filters_xuat,
		fields=["vat_tu", "so_lo", "so_luong"],
	)
	xuat_pool: dict[tuple, float] = {}
	for e in xuat_entries:
		key = (e["vat_tu"], e["so_lo"])
		# so_luong ÂM cho lượt xuất gốc, DƯƠNG cho dòng đảo bù trừ (xem
		# docstring) — trừ so_luong ra khỏi pool cộng dồn đúng cả hai chiều.
		xuat_pool[key] = xuat_pool.get(key, 0.0) - float(e["so_luong"])

	dot_map: dict[tuple, dict] = {}
	for e in nhap_entries:
		key = (e["chung_tu"], e["vat_tu"], e["so_lo"])
		d = dot_map.setdefault(key, {
			"dot": e["chung_tu"], "vat_tu": e["vat_tu"], "lo": e["so_lo"],
			"han_su_dung": e["han_su_dung"], "ngay_nhan": e["ngay"],
			"sl_nhap": 0.0, "gia_tri_nhap": 0.0,
		})
		d["sl_nhap"] += float(e["so_luong"])
		d["gia_tri_nhap"] += float(e["gia_tri"] or 0)

	dot_names = {d["dot"] for d in dot_map.values()}

	# I-2: trừ đúng phần đã đảo TÍNH TỚI den_ngay. Phiếu đảo là chứng từ
	# RIÊNG (chung_tu khác đợt gốc), tra ngược qua phieu_goc; chỉ những phiếu
	# đảo đã xảy ra TRONG kỳ báo cáo (ngay <= den) mới được trừ — một phiếu
	# đảo lập SAU den_ngay không được phép "viết lại" một báo cáo đã đóng.
	if dot_names:
		dao_receipts = frappe.get_all(
			"Customer Stock Receipt",
			filters={
				"kho": kho, "loai_nhap": voucher.LOAI_DAO,
				"phieu_goc": ["in", list(dot_names)], "ngay": ["<=", den],
			},
			fields=["name", "phieu_goc"],
		)
		if dao_receipts:
			goc_cua = {r["name"]: r["phieu_goc"] for r in dao_receipts}
			dao_entries = frappe.get_all(
				"Customer Stock Ledger Entry",
				filters={
					"kho": kho, "chung_tu_type": "Customer Stock Receipt",
					"chung_tu": ["in", list(goc_cua)],
				},
				fields=["chung_tu", "vat_tu", "so_lo", "so_luong", "gia_tri"],
			)
			for e in dao_entries:
				key = (goc_cua[e["chung_tu"]], e["vat_tu"], e["so_lo"])
				d = dot_map.get(key)
				if not d:
					continue
				# so_luong/gia_tri của dòng đảo đã mang dấu ÂM sẵn (xem
				# `_he_so_dau`) — CỘNG thẳng vào là trừ đúng phần đã đảo.
				d["sl_nhap"] += float(e["so_luong"])
				d["gia_tri_nhap"] += float(e["gia_tri"] or 0)

	# Loại đợt đã ròng về 0 (huỷ toàn bộ trong kỳ) — không hiện dòng 0.
	dot_map = {k: d for k, d in dot_map.items() if d["sl_nhap"] > EPS}

	receipt_names = {d["dot"] for d in dot_map.values()}
	receipts = {}
	if receipt_names:
		receipts = {
			r["name"]: r for r in frappe.get_all(
				"Customer Stock Receipt", filters={"name": ["in", list(receipt_names)]},
				fields=["name", "loai_nhap", "ncc", "so_chung_tu_ncc", "delivery_note"],
			)
		}
	ncc_names = {r["ncc"] for r in receipts.values() if r.get("ncc")}
	ncc_ten = {}
	if ncc_names:
		ncc_ten = dict(frappe.get_all(
			"Customer Supplier", filters={"name": ["in", list(ncc_names)]},
			fields=["name", "ten_ncc"], as_list=True,
		))

	nguong = _nguong_cham_luan_chuyen()

	by_lot: dict[tuple, list] = {}
	for d in dot_map.values():
		by_lot.setdefault((d["vat_tu"], d["lo"]), []).append(d)

	out = []
	for lot_key, dots in by_lot.items():
		# FIFO: đợt nhận SỚM NHẤT tiêu thụ trước. `dot` (tên phiếu) là
		# tiebreak khi hai đợt trùng ngày nhận — autoname tăng đơn điệu nên
		# tên phiếu cũ hơn luôn "nhỏ" hơn theo thứ tự chuỗi trong cùng kho/năm.
		dots.sort(key=lambda d: (frappe.utils.getdate(d["ngay_nhan"]), d["dot"]))
		con_lai_pool = xuat_pool.get(lot_key, 0.0)
		for d in dots:
			sl_nhap = _r(d["sl_nhap"])
			da_xuat = _r(min(max(con_lai_pool, 0.0), sl_nhap))
			con_lai_pool -= da_xuat
			con_lai = _r(sl_nhap - da_xuat)
			ngay_nhan = frappe.utils.getdate(d["ngay_nhan"])

			if ngay_nhan < tu or ngay_nhan > den:
				continue

			rc = receipts.get(d["dot"]) or {}
			if rc.get("loai_nhap") == "Mua ngoài (NCC khác)" and rc.get("ncc"):
				nguon_dot = ncc_ten.get(rc["ncc"], rc["ncc"])
				chung_tu_ncc = rc.get("so_chung_tu_ncc") or ""
			else:
				nguon_dot = "Miyano"
				chung_tu_ncc = rc.get("delivery_note") or ""
			if nguon and nguon_dot != nguon:
				continue

			tuoi = (den - ngay_nhan).days
			pct = round(da_xuat / sl_nhap * 100, 2) if sl_nhap > EPS else 0.0
			out.append({
				"dot": d["dot"],
				"ngay_nhan": ngay_nhan,
				"nguon": nguon_dot,
				"chung_tu": chung_tu_ncc,
				"vat_tu": d["vat_tu"],
				"lo": d["lo"],
				"han_su_dung": d["han_su_dung"],
				"sl_nhap": sl_nhap,
				"gia_tri_nhap": _r(d["gia_tri_nhap"]),
				"da_xuat": da_xuat,
				"con_lai": con_lai,
				"tuoi_ton_ngay": tuoi,
				"pct_tieu_thu": pct,
				# Chỉ đợt CÒN TỒN mới đáng gắn cờ chậm luân chuyển — một đợt
				# đã tiêu thụ hết (con_lai=0) không còn nằm kho nữa, tuổi tồn
				# của nó không còn ý nghĩa cảnh báo (khớp AC US-E4.7: PNK-001
				# còn 0 -> không cờ dù tuổi tồn cũng đã vượt ngưỡng).
				"cham_luan_chuyen": bool(nguong > 0 and con_lai > EPS and tuoi > nguong),
			})

	return sorted(out, key=lambda r: (r["vat_tu"], r["lo"], r["ngay_nhan"], r["dot"]))


def bao_cao_cap_phat_rows(
	kho: str, tu_ngay, den_ngay, khoa_phong: str | None = None, vat_tu: str | None = None,
) -> dict:
	"""Báo cáo cấp phát theo khoa phòng — US-E8.5, UC-56, BR-CP4.

	Đọc từ SỔ KHO join qua PHIẾU XUẤT (khoa phòng/người nhận nằm trên đầu
	phiếu, không trên dòng sổ) — KHÔNG đổi schema `Customer Stock Ledger
	Entry` (BR-CP4 nói rõ điều này).

	LỆCH có chủ đích so với quy ước chung của module này (xem docstring đầu
	file: "da_dao=1 KHÔNG bị lọc khỏi bất kỳ tổng nào"): báo cáo NXT/thẻ kho
	trả lời câu hỏi lịch sử kế toán ("mọi biến động trong kỳ, kể cả phần đã
	bị đảo sau đó"), còn cấp phát trả lời câu hỏi nghiệp vụ khác — "khoa nào
	ĐANG THỰC SỰ giữ hàng đã cấp phát" — nên PRD E8 nói thẳng "Phiếu bị đảo
	không tính". Loại trừ HAI LỚP, cố ý không gộp làm một:
	  * `da_dao=0` ở tầng sổ — bỏ dòng GỐC của một phiếu đã bị huỷ;
	  * `loai_xuat == "Xuất sử dụng"` ở tầng phiếu — bỏ chính dòng BÙ TRỪ
	    (loai_xuat="Phiếu đảo", da_dao=0 vì bản thân nó không bị đảo) VÀ
	    mọi loại xuất khác (huỷ/trả lại/điều chỉnh — không phải cấp phát
	    thật). Chỉ lọc một lớp sẽ để lọt lớp còn lại.

	Nhóm theo `khoa_phong`; dòng KHÔNG gắn khoa (phiếu tạo khi kho chưa bật
	`bat_buoc_khoa_phong`, hoặc kho chưa từng bật) tách thành nhóm riêng
	`khoa_phong=None`/`ten_hien_thi="Chưa gắn khoa"` — KHÔNG bị loại khỏi
	báo cáo và KHÔNG lẫn vào khoa nào khác (yêu cầu tường minh của US-E8.5:
	"đừng giấu nó, đó là dữ liệu thật").

	Một dòng "dong" = một (phiếu, vật tư), CỘNG DỒN qua mọi lô của vật tư đó
	trên cùng phiếu — dvt/tên vật tư không phụ thuộc lô, và bộ số chuẩn của
	PRD ("Găng M 8 hộp") không tách theo lô.

	F-4 (review E8, CHẶN) — mỗi dòng còn mang thêm `noi_nhan` (field free-text
	CŨ, tồn tại từ trước E8 trên chính đầu phiếu, KHÔNG liên quan schema sổ)
	để nhóm "Chưa gắn khoa" còn CỨU ĐƯỢC dữ liệu: một phiếu chưa chọn
	khoa_phong (kho chưa bật bắt buộc, hoặc thủ kho quen gõ tự do) rất có thể
	đã ghi đúng tên khoa vào `noi_nhan` — ẩn nó đi sẽ biến "Chưa gắn khoa"
	thành "không biết gì cả", trong khi thật ra biết một phần.
	"""
	tu = frappe.utils.getdate(tu_ngay)
	den = frappe.utils.getdate(den_ngay)
	if tu > den:
		frappe.throw("Từ ngày phải trước hoặc bằng Đến ngày.", frappe.ValidationError)

	filters = {
		"kho": kho, "chung_tu_type": "Customer Stock Issue",
		"da_dao": 0, "ngay": ["between", [tu, den]],
	}
	if vat_tu:
		filters["vat_tu"] = vat_tu
	entries = frappe.get_all(
		"Customer Stock Ledger Entry", filters=filters,
		fields=["chung_tu", "vat_tu", "so_luong", "gia_tri", "ngay"],
	)
	if not entries:
		return {"tong_gia_tri": 0.0, "nhom": []}

	issue_names = {e["chung_tu"] for e in entries}
	issues = {
		r["name"]: r for r in frappe.get_all(
			"Customer Stock Issue", filters={"name": ["in", list(issue_names)]},
			fields=["name", "loai_xuat", "khoa_phong", "nguoi_nhan", "noi_nhan"],
		)
	}
	vat_tu_names = {e["vat_tu"] for e in entries}
	vt_info = {
		r["name"]: r for r in frappe.get_all(
			"Customer Warehouse Item", filters={"name": ["in", list(vat_tu_names)]},
			fields=["name", "ten_vat_tu", "dvt"],
		)
	} if vat_tu_names else {}

	agg: dict[tuple, dict] = {}
	for e in entries:
		iss = issues.get(e["chung_tu"])
		if not iss or iss["loai_xuat"] != "Xuất sử dụng":
			continue
		kp = iss["khoa_phong"] or None
		if khoa_phong and kp != khoa_phong:
			continue
		key = (iss["name"], e["vat_tu"])
		row = agg.setdefault(key, {
			"khoa_phong": kp, "phieu": iss["name"], "vat_tu": e["vat_tu"],
			"ngay": e["ngay"], "nguoi_nhan": iss["nguoi_nhan"] or "",
			"noi_nhan": iss["noi_nhan"] or "",
			"sl": 0.0, "gia_tri": 0.0,
		})
		# so_luong/gia_tri của dòng XUẤT mang dấu ÂM (xem docstring
		# ledger.post_lines) — đảo dấu để hiển thị số dương cho người dùng.
		row["sl"] += -float(e["so_luong"])
		row["gia_tri"] += -float(e["gia_tri"] or 0)

	nhom_map: dict = {}
	for row in agg.values():
		kp = row["khoa_phong"]
		nhom = nhom_map.setdefault(kp, {"gia_tri": 0.0, "dong": []})
		info = vt_info.get(row["vat_tu"]) or {}
		nhom["dong"].append({
			"phieu": row["phieu"],
			"ngay": row["ngay"],
			"vat_tu": info.get("ten_vat_tu") or row["vat_tu"],
			"dvt": info.get("dvt") or "",
			"sl": _r(row["sl"]),
			"gia_tri": round(row["gia_tri"], 2),
			"nguoi_nhan": row["nguoi_nhan"],
			"noi_nhan": row["noi_nhan"],
		})
		nhom["gia_tri"] += row["gia_tri"]

	tong_gia_tri = sum(v["gia_tri"] for v in nhom_map.values())

	ten_khoa = {}
	ten_can_tra = [kp for kp in nhom_map if kp]
	if ten_can_tra:
		ten_khoa = dict(frappe.get_all(
			"Customer Department", filters={"name": ["in", ten_can_tra]},
			fields=["name", "ten_khoa_phong"], as_list=True,
		))

	nhom_out = []
	for kp, v in nhom_map.items():
		gia_tri = round(v["gia_tri"], 2)
		nhom_out.append({
			"khoa_phong": kp,
			"ten_hien_thi": ten_khoa.get(kp, kp) if kp else "Chưa gắn khoa",
			"gia_tri": gia_tri,
			"pct": round(gia_tri / tong_gia_tri * 100, 1) if tong_gia_tri > EPS else 0.0,
			"dong": sorted(v["dong"], key=lambda r: (r["ngay"], r["phieu"])),
		})
	# Khoa có tên sắp trước theo bảng chữ cái; nhóm "Chưa gắn khoa" LUÔN ở
	# cuối — nó không phải một khoa để so tên, và US-E8.5 muốn nó "tách
	# riêng", đứng lẫn giữa danh sách khoa sẽ trông như một khoa thật.
	nhom_out.sort(key=lambda r: (r["khoa_phong"] is None, r["ten_hien_thi"]))

	return {"tong_gia_tri": round(tong_gia_tri, 2), "nhom": nhom_out}


def build_xlsx(columns: list[tuple[str, str]], rows: list[dict], sheet_title: str) -> bytes:
	"""Dựng .xlsx với ĐÚNG bộ cột `columns`, theo ĐÚNG thứ tự — dùng chung cho
	cả ba loại báo cáo. Giá trị ghi thẳng kiểu gốc (số, ngày) chứ không format
	thành chuỗi hiển thị, để file mở lại được bằng công thức/pivot ngay, đúng
	nguyên tắc round-tripping-spreadsheets."""
	wb = Workbook()
	ws = wb.active
	ws.title = sheet_title[:31]  # giới hạn cứng của Excel cho tên sheet
	ws.append([label for label, _ in columns])
	for cell in ws[1]:
		cell.font = Font(bold=True)
	for row in rows:
		ws.append([row.get(field, "") for _, field in columns])
	for i, (label, _) in enumerate(columns, start=1):
		ws.column_dimensions[get_column_letter(i)].width = max(12, min(30, len(label) + 4))
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()
