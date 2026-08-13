"""Import danh mục vật tư + tồn đầu kỳ cho kho khách hàng, có preview.

Một bộ cột DUY NHẤT (COLUMNS) dùng chung cho cả ba việc: sinh file mẫu, đọc
file preview, đọc file commit — đúng nguyên tắc của round-tripping-spreadsheets:
"export, template, and import share one column config". Nếu sau này cần thêm
cột, chỉ sửa ở đây.

preview và commit đều gọi chung parse_workbook(): commit KHÔNG tin bất cứ dữ
liệu dòng nào mà client gửi lên, nó tự đọc lại và validate lại y hệt preview
đã làm, đúng nguyên tắc của previewing-imports-before-writing.
"""

import io
import unicodedata
from datetime import date, datetime

import frappe
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from miyano_portal.kho import ledger

# (nhãn cột hiển thị, tên field nội bộ) — ĐÚNG THỨ TỰ đề bài yêu cầu cho file mẫu.
COLUMNS = [
	("Mã vật tư", "ma_vat_tu"),
	("Tên vật tư", "ten_vat_tu"),
	("ĐVT", "dvt"),
	("Số lô", "so_lo"),
	("Hạn sử dụng", "han_su_dung"),
	("Số lượng", "so_luong"),
	("Đơn giá", "don_gia"),
	("Quy cách", "quy_cach"),
	("Nhóm", "nhom"),
]

# Các cột bắt buộc phải có GIÁ TRỊ trên mỗi dòng. `so_lo` cố ý KHÔNG bắt buộc:
# thiếu thì nhận mặc định LOT_KHONG_CO, giống hành vi của voucher.fill_item_details()
# cho phiếu nhập nhập tay.
REQUIRED_FIELDS = {"ma_vat_tu", "ten_vat_tu", "dvt", "so_luong", "don_gia"}


def build_aliases(columns: list[tuple[str, str]]) -> dict[str, str]:
	"""Bảng nhận diện tiêu đề cột: nhãn hiển thị (đã NFC + trim + lower) và
	chính tên field đều trỏ về field. Nhờ vậy header gõ lệch hoa/thường hay
	đảo thứ tự cột vẫn nhận ra được.

	Là HÀM chứ không phải hằng số module: ba bộ cột (tồn đầu kỳ, danh mục vật
	tư, dòng phiếu) dùng chung đúng một cơ chế nhận diện này.
	"""
	aliases: dict[str, str] = {}
	for label, field in columns:
		aliases[unicodedata.normalize("NFC", label).strip().lower()] = field
		aliases[field] = field
	return aliases


def _norm(value) -> str:
	if value is None:
		return ""
	return unicodedata.normalize("NFC", str(value)).strip()


def _fold(value) -> str:
	return _norm(value).lower()


def _coerce_date(value) -> tuple[str | None, str | None]:
	"""Trả (giá trị ISO hoặc None, lỗi hoặc None)."""
	if value in (None, ""):
		return None, None
	if isinstance(value, (datetime, date)):
		return frappe.utils.getdate(value).strftime("%Y-%m-%d"), None
	s = str(value).strip()
	if not s:
		return None, None
	for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
		try:
			return datetime.strptime(s, fmt).strftime("%Y-%m-%d"), None
		except ValueError:
			continue
	return None, f"Hạn sử dụng không hợp lệ: '{s}' (dùng dd/mm/yyyy hoặc yyyy-mm-dd)"


def _coerce_num(value) -> tuple[float | None, str | None]:
	if isinstance(value, (int, float)):
		return float(value), None
	s = str(value).strip().replace(" ", "").replace(",", "")
	try:
		return float(s), None
	except ValueError:
		return None, f"'{value}' không phải số"


def build_template_bytes() -> bytes:
	"""File mẫu .xlsx: đúng 9 cột theo thứ tự COLUMNS, kèm một dòng ví dụ.

	Dòng ví dụ dùng mã KHÁCH TỰ THÊM (không trùng bất kỳ Item nào của Miyano)
	để việc tải mẫu xuống rồi nạp lại ngay (không sửa gì) luôn thành công,
	không phụ thuộc catalog Item thật có tồn tại trên site hay không.
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = "Tồn đầu kỳ"
	ws.append([label for label, _ in COLUMNS])
	for cell in ws[1]:
		cell.font = Font(bold=True)
	ws.append([
		"VT-001", "Bông y tế 500g", "Cuộn", "LO-VIDU-01",
		date(2027, 12, 31), 100, 15000, "Cuộn 500g", "Vật tư tiêu hao",
	])
	for i, width in enumerate([16, 30, 10, 14, 14, 12, 14, 16, 16], start=1):
		ws.column_dimensions[get_column_letter(i)].width = width
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


def read_header(ws, columns, required_fields) -> tuple[int, dict[str, int]]:
	"""Tìm dòng tiêu đề trong 5 dòng đầu và ánh xạ field -> chỉ số cột.

	`columns`/`required_fields` là tham số chứ không phải hằng số module: hàm
	này phục vụ cả ba bộ cột. `required_fields` ở đây là các cột BẮT BUỘC PHẢI
	CÓ MẶT trong header — khác với "bắt buộc có giá trị ở mỗi dòng", việc đó
	do từng hàm parse tự kiểm.
	"""
	aliases = build_aliases(columns)
	header_row = None
	header_cells = None
	for r, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5)), start=1):
		if any(c.value not in (None, "") for c in row_cells):
			header_row = r
			header_cells = row_cells
			break
	if header_row is None:
		frappe.throw("Tệp trống, không có dữ liệu.", frappe.ValidationError)

	col_index: dict[str, int] = {}
	for idx, cell in enumerate(header_cells, start=1):
		field = aliases.get(_fold(cell.value))
		if field:
			col_index[field] = idx

	missing = [label for label, field in columns if field in required_fields and field not in col_index]
	if missing:
		frappe.throw(
			"Tệp thiếu cột bắt buộc: " + ", ".join(missing) + ". Vui lòng tải lại tệp mẫu.",
			frappe.ValidationError,
		)
	return header_row, col_index


def _read_header(ws) -> tuple[int, dict[str, int]]:
	"""Bộ cột tồn đầu kỳ. Giữ lại để nơi gọi cũ và test cũ không phải đổi."""
	return read_header(ws, COLUMNS, REQUIRED_FIELDS)


def _cell_value(row_cells, col: int):
	idx = col - 1
	return row_cells[idx].value if idx < len(row_cells) else None


def _match_vat_tu(kho: str, ma_vat_tu: str) -> tuple[str, str, str | None]:
	"""Phân loại một mã vật tư: đã có trong kho / khớp Item Miyano / mã riêng.

	So khớp KHÔNG PHÂN BIỆT HOA THƯỜNG, đã trim — thực hiện tường minh bằng
	lower() trong SQL/Python thay vì trông cậy collation của DB (có thể đổi
	giữa các site). Với mã khớp Item, ghi lại CHÍNH TẢ CHUẨN từ DB (item_code
	thật), không phải cách người dùng gõ trong file.

	Trả về (match_type, item_code, vat_tu_name):
	  * "existing" — đã có Customer Warehouse Item trong kho này, vat_tu_name
	    là name của bản ghi đó (dùng lại, không tạo mới).
	  * "miyano"   — khớp Item.item_code của Miyano, vat_tu_name là None (sẽ
	    tạo mới lúc commit, với item_code = mã Item thật).
	  * "private"  — mã khách tự thêm, item_code rỗng, vat_tu_name là None.
	"""
	fold = ma_vat_tu.strip().lower()
	existing = frappe.db.sql(
		"""select name, item_code from `tabCustomer Warehouse Item`
		   where kho=%s and lower(ma_vat_tu)=%s limit 1""",
		(kho, fold),
	)
	if existing:
		name, item_code = existing[0]
		return "existing", item_code or "", name
	item_row = frappe.db.sql(
		"select item_code from `tabItem` where lower(item_code)=%s limit 1", (fold,)
	)
	if item_row:
		return "miyano", item_row[0][0], None
	return "private", "", None


def mo_workbook(content: bytes):
	"""Mở nội dung .xlsx và trả về sheet đang hoạt động.

	Tách riêng vì cả ba đường import đều cần đúng một thông điệp tiếng Việt khi
	tệp hỏng — openpyxl ném lỗi tiếng Anh nêu chi tiết nội bộ của tệp zip.
	"""
	try:
		wb = load_workbook(io.BytesIO(content), data_only=True)
	except Exception:
		frappe.throw(
			"Tệp không đúng định dạng .xlsx hoặc đã hỏng. Vui lòng dùng tệp mẫu.",
			frappe.ValidationError,
		)
	return wb.active


def _chan_neu_da_nhap_ton_dau(kho: str) -> None:
	"""BR-K21 / NL-4.4 / US-E4.3: tồn đầu kỳ chỉ nhập được MỘT LẦN cho mỗi kho.

	Chặn NGAY TỪ BƯỚC UPLOAD (preview), không phải sau khi đã đọc/xử lý file —
	đặt ở ĐẦU parse_workbook() nên cả kho_import_preview lẫn kho_import_commit
	(gọi lại parse_workbook trước khi ghi) đều tự động được chặn qua đúng một
	chỗ. Chỉ tính phiếu ĐÃ GHI SỔ (docstatus=1) — "đã commit" theo đúng nghĩa
	của BR-K21; một phiếu nháp bị bỏ dở không tính là đã nhập.
	"""
	ngay_da_nhap = frappe.db.get_value(
		"Customer Stock Receipt",
		{"kho": kho, "loai_nhap": "Tồn đầu kỳ", "docstatus": 1},
		"ngay",
	)
	if ngay_da_nhap:
		frappe.throw(
			f"Kho đã nhập tồn đầu kỳ ngày {frappe.utils.formatdate(ngay_da_nhap)}. "
			"Dùng phiếu Điều chỉnh kiểm kê cho chênh lệch.",
			frappe.ValidationError,
		)


def parse_workbook(content: bytes, kho: str) -> dict:
	"""Đọc và validate toàn bộ file, KHÔNG GHI GÌ vào database.

	Trả verdict theo khuôn mẫu của previewing-imports-before-writing: tổng số
	dòng, số dòng hợp lệ/lỗi, tóm tắt phân loại, và danh sách chi tiết cho cả
	hai nhóm — mỗi dòng lỗi mang đủ số dòng trong file gốc (1-based, tính cả
	header) và MỌI lý do lỗi trên dòng đó, không chỉ lý do đầu tiên.
	"""
	_chan_neu_da_nhap_ton_dau(kho)
	ws = mo_workbook(content)
	header_row, col_index = _read_header(ws)

	rows_ok: list[dict] = []
	rows_error: list[dict] = []

	for line, row_cells in enumerate(
		ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1
	):
		raw = {field: _cell_value(row_cells, col) for field, col in col_index.items()}
		if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw.values()):
			continue  # dòng trắng hoàn toàn — bỏ qua, không tính vào total

		errors: list[str] = []
		ma_vat_tu = _norm(raw.get("ma_vat_tu"))
		ten_vat_tu = _norm(raw.get("ten_vat_tu"))
		dvt = _norm(raw.get("dvt"))
		so_lo = _norm(raw.get("so_lo")) or ledger.LOT_KHONG_CO
		quy_cach = _norm(raw.get("quy_cach"))
		nhom = _norm(raw.get("nhom"))

		if not ma_vat_tu:
			errors.append("Thiếu Mã vật tư")
		if not ten_vat_tu:
			errors.append("Thiếu Tên vật tư")
		if not dvt:
			errors.append("Thiếu ĐVT")

		han_su_dung, han_err = _coerce_date(raw.get("han_su_dung"))
		if han_err:
			errors.append(han_err)

		so_luong_raw = raw.get("so_luong")
		so_luong = None
		if so_luong_raw in (None, ""):
			errors.append("Thiếu Số lượng")
		else:
			so_luong, num_err = _coerce_num(so_luong_raw)
			if num_err:
				errors.append(f"Số lượng không hợp lệ: {num_err}")
			elif so_luong <= 0:
				errors.append("Số lượng phải lớn hơn 0")

		don_gia_raw = raw.get("don_gia")
		don_gia = None
		if don_gia_raw in (None, ""):
			errors.append("Thiếu Đơn giá")
		else:
			don_gia, num_err = _coerce_num(don_gia_raw)
			if num_err:
				errors.append(f"Đơn giá không hợp lệ: {num_err}")
			elif don_gia < 0:
				errors.append("Đơn giá không được âm")

		if errors:
			rows_error.append({
				"line": line,
				"ma_vat_tu": ma_vat_tu or f"(dòng {line})",
				"errors": errors,
			})
			continue

		match_type, item_code, vat_tu_name = _match_vat_tu(kho, ma_vat_tu)
		rows_ok.append({
			"line": line,
			"ma_vat_tu": ma_vat_tu,
			"ten_vat_tu": ten_vat_tu,
			"dvt": dvt,
			"so_lo": so_lo,
			"han_su_dung": han_su_dung,
			"so_luong": so_luong,
			"don_gia": don_gia,
			"quy_cach": quy_cach,
			"nhom": nhom,
			"match_type": match_type,
			"item_code": item_code,
			"vat_tu": vat_tu_name,
		})

	summary_key = {"existing": "existing_in_kho", "miyano": "matched_miyano", "private": "private_new"}
	summary = {"existing_in_kho": 0, "matched_miyano": 0, "private_new": 0}
	for row in rows_ok:
		summary[summary_key[row["match_type"]]] += 1

	return {
		"total": len(rows_ok) + len(rows_error),
		"ok_count": len(rows_ok),
		"error_count": len(rows_error),
		"summary": summary,
		"rows_ok": rows_ok,
		"rows_error": rows_error,
	}


def commit_workbook(content: bytes, kho: str) -> dict:
	"""Đọc lại TỪ ĐẦU trên server (không tin dữ liệu client gửi), rồi ghi.

	Tất-cả-hoặc-không: bất kỳ dòng lỗi nào cũng khiến KHÔNG GÌ được ghi. Phần
	ghi thật (tạo Vật Tư Kho Khách còn thiếu + một Phiếu Nhập Kho) được bọc
	trong một savepoint riêng — lỗi bất ngờ ở bước ghi (ví dụ ràng buộc unique,
	hoặc _ensure_non_negative) khiến state quay lại đúng trước khi savepoint
	được đặt, không để lại một phần dữ liệu nửa vời.
	"""
	parsed = parse_workbook(content, kho)
	if parsed["error_count"]:
		first = parsed["rows_error"][0]
		frappe.throw(
			f"Tệp có {parsed['error_count']} dòng lỗi trong tổng số {parsed['total']} dòng "
			f"(ví dụ dòng {first['line']}: {'; '.join(first['errors'])}). "
			"Vui lòng sửa và tải lại — chưa có dữ liệu nào được ghi.",
			frappe.ValidationError,
		)
	if not parsed["rows_ok"]:
		frappe.throw("Tệp không có dòng dữ liệu hợp lệ nào để nhập.", frappe.ValidationError)

	ngay_bat_dau = frappe.db.get_value("Customer Warehouse", kho, "ngay_bat_dau")

	sp = "kho_import_commit_sp"
	frappe.db.savepoint(sp)
	try:
		code_to_vat_tu: dict[str, str] = {}
		for row in parsed["rows_ok"]:
			if row["vat_tu"]:
				continue  # đã có Vật Tư Kho Khách — dùng lại, không tạo mới
			fold = row["ma_vat_tu"].strip().lower()
			if fold in code_to_vat_tu:
				row["vat_tu"] = code_to_vat_tu[fold]
				continue
			cwi = frappe.get_doc({
				"doctype": "Customer Warehouse Item",
				"kho": kho,
				"ma_vat_tu": row["ma_vat_tu"],
				"ten_vat_tu": row["ten_vat_tu"],
				"dvt": row["dvt"],
				"item_code": row["item_code"] or None,
				"quy_cach": row["quy_cach"] or None,
				"nhom": row["nhom"] or None,
			})
			cwi.insert(ignore_permissions=True)
			code_to_vat_tu[fold] = cwi.name
			row["vat_tu"] = cwi.name

		receipt = frappe.get_doc({
			"doctype": "Customer Stock Receipt",
			"kho": kho,
			"ngay": ngay_bat_dau or frappe.utils.today(),
			"loai_nhap": "Tồn đầu kỳ",
			"dien_giai": "Nhập tồn đầu kỳ từ tệp Excel qua cổng khách hàng.",
			"items": [{
				"vat_tu": row["vat_tu"],
				"so_lo": row["so_lo"],
				"han_su_dung": row["han_su_dung"],
				"so_luong": row["so_luong"],
				"don_gia": row["don_gia"],
			} for row in parsed["rows_ok"]],
		})
		receipt.insert(ignore_permissions=True)
		receipt.submit()
	except Exception:
		frappe.db.rollback(save_point=sp)
		raise

	return {
		"receipt": receipt.name,
		"created_items": len(code_to_vat_tu),
		"rows_written": len(parsed["rows_ok"]),
		"summary": parsed["summary"],
	}
