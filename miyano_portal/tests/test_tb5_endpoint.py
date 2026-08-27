"""Logic + endpoint danh mục thiết bị.

Ca quan trọng nhất KHÔNG phải "gửi bậy thì lỗi" mà là "gửi bậy thì bị ÉP về
đúng": một nhân viên khoa A gửi kèm khoa B vẫn tạo ra máy thuộc khoa A. Trả
lỗi cũng chấp nhận được về mặt an toàn, nhưng ép là hành vi đã chốt
(BR-TB-6) và giống `portal_context.khoa_phong_cho_don()` đang chạy — TRỪ
hai điểm module này KHÔNG dùng lại hàm đó (xem docstring `_khoa_ep_theo_
phien` trong `kho/thiet_bi.py`).

Bốn lớp test trong file này (Task 6, 7, 8, 11) dùng CHUNG lớp nền
`_NenThietBi` định nghĩa ngay dưới đây — ba task sau kế thừa, không dựng lại
fixture riêng.
"""

import re

import frappe
from frappe.tests.utils import FrappeTestCase

from miyano_portal.api import kho as kho_api
from miyano_portal.kho import dong_phieu
from miyano_portal.kho import thiet_bi as thiet_bi_mod

KHACH = "ZZTB5 Benh Vien"
KHACH_KHAC = "ZZTB5 Benh Vien Khac"


class _NenThietBi(FrappeTestCase):
	"""Nền dùng chung cho Task 6 (logic), 7 (endpoint), 8 (phiếu xuất/nhận
	máy), 11 (cột Excel). Đọc `_don()` trước khi sửa — lọc ĐÚNG hai bệnh
	viện ZZTB5 của bộ test này, không filter mặc định khớp-tất-cả (site
	erptest.local mang dữ liệu demo thật của nhiều bệnh viện/bộ test khác)."""

	def setUp(self):
		frappe.set_user("Administrator")
		self.khach = KHACH
		self.khach_khac = KHACH_KHAC
		self._don()
		self.addCleanup(self._don)

		frappe.get_doc({
			"doctype": "Customer", "customer_name": KHACH,
			"customer_type": "Company", "customer_group": "All Customer Groups",
			"territory": "All Territories",
			"custom_ma_ngan": "ZZT5", "custom_cho_phep_mua_le": 1,
		}).insert(ignore_permissions=True)
		frappe.get_doc({
			"doctype": "Customer", "customer_name": KHACH_KHAC,
			"customer_type": "Company", "customer_group": "All Customer Groups",
			"territory": "All Territories",
			"custom_ma_ngan": "ZZT5K", "custom_cho_phep_mua_le": 1,
		}).insert(ignore_permissions=True)

		self.kho = frappe.get_doc({
			"doctype": "Customer Warehouse", "customer": KHACH,
			"ten_kho": "ZZTB5 Kho", "ma_kho": "ZZTB5",
			"ngay_bat_dau": frappe.utils.add_days(frappe.utils.today(), -30),
		}).insert(ignore_permissions=True)
		self.kho_khac = frappe.get_doc({
			"doctype": "Customer Warehouse", "customer": KHACH_KHAC,
			"ten_kho": "ZZTB5 Kho Khac", "ma_kho": "ZZTB5K",
			"ngay_bat_dau": frappe.utils.add_days(frappe.utils.today(), -30),
		}).insert(ignore_permissions=True)

		self.kp_a = frappe.get_doc({
			"doctype": "Customer Department", "customer": KHACH,
			"ten_khoa_phong": "ZZTB5 Khoa A", "ma_khoa": "ZZT5A",
		}).insert(ignore_permissions=True)
		self.kp_b = frappe.get_doc({
			"doctype": "Customer Department", "customer": KHACH,
			"ten_khoa_phong": "ZZTB5 Khoa B", "ma_khoa": "ZZT5B",
		}).insert(ignore_permissions=True)

		# self.ql/nv_a/nv_b PHẢI là CHUỖI EMAIL (không phải Document) — mọi
		# test gọi frappe.set_user(self.nv_a) rồi truyền thẳng self.nv_a làm
		# tham số `user` cho thiet_bi_mod.*; pham_vi_don()/get_portal_member()
		# tra `Portal Member` bằng {"user": user} nên một Document ở đó sẽ
		# không khớp gì và ném PermissionError cho mọi test, kể cả ca đúng.
		self.tv_ql = self._tv("zztb5.ql@demo.miyano", KHACH, "Quản lý", None)
		self.ql = self.tv_ql.user
		self.tv_nv_a = self._tv("zztb5.nv.a@demo.miyano", KHACH, "Nhân viên khoa", self.kp_a.name)
		self.nv_a = self.tv_nv_a.user
		self.tv_nv_b = self._tv("zztb5.nv.b@demo.miyano", KHACH, "Nhân viên khoa", self.kp_b.name)
		self.nv_b = self.tv_nv_b.user

		self.may_a = frappe.get_doc({
			"doctype": "Customer Equipment", "customer": KHACH,
			"ma_thiet_bi": "ZZTB5-MAY-A", "ten_thiet_bi": "May khoa A",
			"khoa_phong": self.kp_a.name,
		}).insert(ignore_permissions=True)
		self.may_b = frappe.get_doc({
			"doctype": "Customer Equipment", "customer": KHACH,
			"ma_thiet_bi": "ZZTB5-MAY-B", "ten_thiet_bi": "May khoa B",
			"khoa_phong": self.kp_b.name,
		}).insert(ignore_permissions=True)
		self.may_chung = frappe.get_doc({
			"doctype": "Customer Equipment", "customer": KHACH,
			"ma_thiet_bi": "ZZTB5-MAY-DC", "ten_thiet_bi": "May dung chung",
		}).insert(ignore_permissions=True)
		# Bệnh viện KHÁC — có thật, để các test cách ly liên viện (gan_vao_
		# vat_tu, và Task 7/8/11) có fixture thật thay vì phải tự dựng lại.
		self.may_benh_vien_khac = frappe.get_doc({
			"doctype": "Customer Equipment", "customer": KHACH_KHAC,
			"ma_thiet_bi": "ZZTB5K-MAY", "ten_thiet_bi": "May benh vien khac",
		}).insert(ignore_permissions=True)

		self.vat_tu = frappe.get_doc({
			"doctype": "Customer Warehouse Item", "kho": self.kho.name,
			"ma_vat_tu": "ZZTB5-HC1", "ten_vat_tu": "Hoa chat 1", "dvt": "Hop",
		}).insert(ignore_permissions=True)
		self.vat_tu_khac_kho = frappe.get_doc({
			"doctype": "Customer Warehouse Item", "kho": self.kho_khac.name,
			"ma_vat_tu": "ZZTB5K-HC1", "ten_vat_tu": "Hoa chat kho khac", "dvt": "Hop",
		}).insert(ignore_permissions=True)

		self.lo = "ZZTB5-LO1"
		phieu = frappe.get_doc({
			"doctype": "Customer Stock Receipt", "kho": self.kho.name,
			"ngay": frappe.utils.today(), "loai_nhap": "Nhập khác",
			"nguoi_giao": "ZZTB5 Nguoi giao",
			"items": [{
				"vat_tu": self.vat_tu.name, "so_lo": self.lo,
				"han_su_dung": "2030-01-01", "so_luong": 100, "don_gia": 1000,
			}],
		})
		phieu.insert(ignore_permissions=True)
		phieu.submit()
		self.phieu_nhap = phieu

	def _tv(self, email, khach, vai_tro, khoa_phong):
		"""Dựng User + Contact + Portal Member — sao y `_tv()` của
		`test_tb4_cach_ly.py` (đã kiểm chứng đúng, không viết lại logic
		Contact "mồ côi")."""
		if not frappe.db.exists("User", email):
			u = frappe.get_doc({
				"doctype": "User", "email": email, "first_name": "ZZ",
				"user_type": "Website User", "send_welcome_email": 0,
			})
			u.append("roles", {"role": "Customer"})
			u.insert(ignore_permissions=True)
			ten_contact = f"{khach}-{email}"
			if not frappe.db.exists("Contact", ten_contact):
				ct = frappe.get_doc({"doctype": "Contact", "first_name": khach, "user": email})
				ct.name = ten_contact
				ct.append("email_ids", {"email_id": email, "is_primary": 1})
				ct.append("links", {"link_doctype": "Customer", "link_name": khach})
				ct.insert(ignore_permissions=True)
		return frappe.get_doc({
			"doctype": "Portal Member", "user": email, "customer": khach,
			"vai_tro": vai_tro, "khoa_phong": khoa_phong,
		}).insert(ignore_permissions=True)

	def tearDown(self):
		"""Chạy TRƯỚC addCleanup(self._don) (thứ tự unittest: tearDown() rồi
		mới tới cleanup functions) — phiên có thể còn đang là Website User từ
		một `frappe.set_user()` giữa test, và một số bước dọn (Contact) đi
		qua `frappe.delete_doc()` sẽ ném PermissionError nếu không phải
		Administrator."""
		frappe.set_user("Administrator")
		super().tearDown()

	def _don(self):
		"""Dọn CHỈ dữ liệu của hai bệnh viện ZZTB5 (KHACH, KHACH_KHAC) — xem
		`test_tb4_cach_ly.py::_don()` làm khuôn mẫu đúng. TUYỆT ĐỐI không
		filter mặc định khớp-tất-cả."""
		khach = [KHACH, KHACH_KHAC]
		khos = frappe.get_all(
			"Customer Warehouse", filters={"customer": ["in", khach]}, pluck="name"
		) or [""]
		vat_tu = frappe.get_all(
			"Customer Warehouse Item", filters={"kho": ["in", khos]}, pluck="name"
		) or [""]
		phieu_nhap = frappe.get_all(
			"Customer Stock Receipt", filters={"kho": ["in", khos]}, pluck="name"
		) or [""]
		phieu_xuat = frappe.get_all(
			"Customer Stock Issue", filters={"kho": ["in", khos]}, pluck="name"
		) or [""]
		frappe.db.delete("Customer Warehouse Item Equipment", {"parent": ["in", vat_tu]})
		frappe.db.delete("Customer Stock Receipt Item", {"parent": ["in", phieu_nhap]})
		frappe.db.delete("Customer Stock Issue Item", {"parent": ["in", phieu_xuat]})
		frappe.db.delete("Customer Stock Ledger Entry", {"kho": ["in", khos]})
		frappe.db.delete("Customer Stock Lot Balance", {"kho": ["in", khos]})
		frappe.db.delete("Customer Stock Receipt", {"kho": ["in", khos]})
		frappe.db.delete("Customer Stock Issue", {"kho": ["in", khos]})
		frappe.db.delete("Customer Warehouse Item", {"kho": ["in", khos]})
		frappe.db.delete("Customer Equipment", {"customer": ["in", khach]})
		emails = [
			"zztb5.ql@demo.miyano", "zztb5.nv.a@demo.miyano", "zztb5.nv.b@demo.miyano",
		]
		frappe.db.delete("Portal Member", {"user": ["in", emails]})
		for khach_ten in khach:
			for email in emails:
				ct = f"{khach_ten}-{email}"
				if frappe.db.exists("Contact", ct):
					frappe.delete_doc("Contact", ct, force=True, ignore_permissions=True)
		frappe.db.delete("Customer Department", {"customer": ["in", khach]})
		frappe.db.delete("Customer Warehouse", {"customer": ["in", khach]})
		frappe.db.delete("Customer", {"name": ["in", khach]})


class TestThietBiLogic(_NenThietBi):
	def test_nhan_vien_khoa_chi_thay_may_khoa_minh_va_may_chung(self):
		frappe.set_user(self.nv_a)
		ten = {r["name"] for r in thiet_bi_mod.list_rows(self.khach, self.nv_a)}
		self.assertIn(self.may_a.name, ten)
		self.assertIn(self.may_chung.name, ten)
		self.assertNotIn(self.may_b.name, ten)

	def test_quan_ly_thay_tat_ca(self):
		frappe.set_user(self.ql)
		ten = {r["name"] for r in thiet_bi_mod.list_rows(self.khach, self.ql)}
		self.assertEqual(ten, {self.may_a.name, self.may_b.name, self.may_chung.name})

	def test_loc_tang_hai_theo_vat_tu(self):
		self.vat_tu.set("may_su_dung", [{"thiet_bi": self.may_a.name}])
		self.vat_tu.save(ignore_permissions=True)
		frappe.set_user(self.ql)
		ten = {
			r["name"] for r in thiet_bi_mod.list_rows(
				self.khach, self.ql, vat_tu=self.vat_tu.name
			)
		}
		self.assertEqual(ten, {self.may_a.name})

	def test_bang_may_trong_thi_khong_loc_tang_hai(self):
		frappe.set_user(self.ql)
		ten = {
			r["name"] for r in thiet_bi_mod.list_rows(
				self.khach, self.ql, vat_tu=self.vat_tu.name
			)
		}
		self.assertEqual(len(ten), 3)

	def test_nhan_vien_khoa_gui_khoa_khac_thi_bi_ep_ve_khoa_minh(self):
		"""BR-TB-6 — ÉP, không phải tin."""
		frappe.set_user(self.nv_a)
		ra = thiet_bi_mod.save(self.khach, self.nv_a, {
			"ma_thiet_bi": "ZZTB5-NEW", "ten_thiet_bi": "Máy mới",
			"khoa_phong": self.kp_b.name,
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", ra["name"], "khoa_phong"),
			self.kp_a.name,
		)

	def test_nhan_vien_khoa_khong_sua_duoc_may_khoa_khac(self):
		"""BR-TB-7."""
		frappe.set_user(self.nv_a)
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.save(self.khach, self.nv_a, {
				"name": self.may_b.name, "ten_thiet_bi": "Đổi trộm",
			})

	def test_nhan_vien_khoa_khong_sua_duoc_may_dung_chung(self):
		"""BR-TB-8b — thấy và chọn được, nhưng không sửa được."""
		frappe.set_user(self.nv_a)
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.save(self.khach, self.nv_a, {
				"name": self.may_chung.name, "ten_thiet_bi": "Đổi trộm",
			})

	def test_quan_ly_dieu_chuyen_duoc_may_sang_khoa_khac(self):
		"""BR-TB-8."""
		frappe.set_user(self.ql)
		thiet_bi_mod.save(self.khach, self.ql, {
			"name": self.may_a.name, "khoa_phong": self.kp_b.name,
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", self.may_a.name, "khoa_phong"),
			self.kp_b.name,
		)

	def test_sua_khong_gui_khoa_phong_thi_giu_nguyen(self):
		"""Vòng sửa theo advisor: `save()` chỉ ÉP lại khoa_phong khi client
		THỰC SỰ gửi khoá đó. Nếu không, một sửa không liên quan (đổi tên) sẽ
		không được đè `khoa_phong` về None/nơi khác — quan trọng nhất với máy
		DÙNG CHUNG (`khoa_phong` rỗng), nơi một sửa vô tình gửi thiếu khoá có
		thể âm thầm mở rộng/thu hẹp phạm vi nhìn thấy nếu bị gán lại."""
		frappe.set_user(self.ql)
		thiet_bi_mod.save(self.khach, self.ql, {
			"name": self.may_a.name, "ten_thiet_bi": "May khoa A doi ten",
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", self.may_a.name, "khoa_phong"),
			self.kp_a.name,
		)

	def test_truong_tao_nhanh_chua_khoa_phong(self):
		"""Canary cho cảnh báo trong task-11-report.md: `kho_thiet_bi_tao_
		nhanh` an toàn CHỈ VÌ `khoa_phong` chưa có trong tuple này (`tao_
		nhanh()` ép khoa theo phiên bằng `_khoa_ep_theo_phien`, không đọc
		payload client cho trường đó — xem `test_tao_nhanh_nhan_vien_khoa_
		van_bi_ep_khoa` ở trên). Nếu về sau ai thêm `khoa_phong` vào tuple
		này mà quên thêm guard tương ứng, ca này phải đỏ TRƯỚC khi kênh dò
		liên khoa (BR-TB-6) tái diễn — nó không tự sửa được lỗ, chỉ ngăn lỗ
		đó lọt qua âm thầm."""
		self.assertNotIn("khoa_phong", thiet_bi_mod.TRUONG_TAO_NHANH)

	def test_tao_nhanh_van_validate_day_du(self):
		""""Nhanh" nói về SỐ Ô, không nói về độ chặt."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.ValidationError):
			thiet_bi_mod.tao_nhanh(self.khach, self.ql, {
				"ten_thiet_bi": "Máy X", "ma_thiet_bi": self.may_a.ma_thiet_bi,
			})

	def test_tao_nhanh_dien_du_sau_o(self):
		frappe.set_user(self.ql)
		ra = thiet_bi_mod.tao_nhanh(self.khach, self.ql, {
			"ten_thiet_bi": "Máy Cobas", "ma_thiet_bi": "COBAS-01",
			"hang_san_xuat": "Roche", "xuat_xu": "Thuỵ Sĩ", "so_serial": "SN-9",
		})
		doc = frappe.get_doc("Customer Equipment", ra["name"])
		self.assertEqual(doc.hang_san_xuat, "Roche")
		self.assertEqual(doc.xuat_xu, "Thuỵ Sĩ")

	def test_tao_nhanh_nhan_vien_khoa_van_bi_ep_khoa(self):
		"""Đối chứng cho vòng sửa split-assignment: `tao_nhanh()` KHÔNG nhận
		`khoa_phong` từ client (không có trong TRUONG_TAO_NHANH) — nhưng máy
		tạo ra bởi Nhân viên khoa vẫn phải thuộc khoa của họ, không phải rơi
		vào "dùng chung" vì thiếu khoá trong payload."""
		frappe.set_user(self.nv_a)
		ra = thiet_bi_mod.tao_nhanh(self.khach, self.nv_a, {
			"ten_thiet_bi": "May nhanh khoa A", "ma_thiet_bi": "ZZTB5-NHANH-A",
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", ra["name"], "khoa_phong"),
			self.kp_a.name,
		)

	def test_gan_vao_vat_tu_goi_hai_lan_khong_sinh_dong_thu_hai(self):
		thiet_bi_mod.gan_vao_vat_tu(self.vat_tu.name, self.may_a.name)
		thiet_bi_mod.gan_vao_vat_tu(self.vat_tu.name, self.may_a.name)
		self.vat_tu.reload()
		self.assertEqual(len(self.vat_tu.may_su_dung), 1)

	def test_gan_vao_vat_tu_khac_benh_vien_bi_chan(self):
		"""`gan_vao_vat_tu()` không nhận `customer`/`user` — phải tự suy tenant
		từ hai đầu (kho của vật tư -> customer; customer của máy) và CHẶN khi
		lệch, không âm thầm gắn máy bệnh viện khác vào vật tư của bệnh viện
		này."""
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.gan_vao_vat_tu(self.vat_tu.name, self.may_benh_vien_khac.name)

	# -- Review vòng 1 (Important #1) — fail-closed chốt TẠI RANH GIỚI
	# thiet_bi.py, không chỉ ở pham_vi_don()/thiet_bi_query (đã chốt ở Task
	# 5, KHÔNG kiểm ranh giới module này). Đi vòng qua PortalMember.
	# validate() bằng db.set_value, đúng kịch bản đã xác nhận có thật
	# (test_tb4_cach_ly.py::test_nhan_vien_khoa_active_ma_khoa_phong_rong_
	# fail_closed).

	def test_nhan_vien_khoa_active_khoa_rong_bi_chan_o_moi_ham(self):
		"""Important #1 — nếu `_khoa_ep_theo_phien`/`_chan_sua_ngoai_pham_vi`/
		`list_rows` từng quay lại đọc thẳng `vai_tro`/`khoa_phong` (hoặc bọc
		`pham_vi_don()` trong try/except nuốt lỗi), test này phải đỏ — xem
		"Vòng sửa 1" trong task-6-report.md cho kết quả thực nghiệm đột biến."""
		frappe.db.set_value("Portal Member", self.tv_nv_a.name, "khoa_phong", "")
		frappe.set_user(self.nv_a)
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.list_rows(self.khach, self.nv_a)
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.save(self.khach, self.nv_a, {
				"ma_thiet_bi": "ZZTB5-BROKEN", "ten_thiet_bi": "May hong khoa",
			})
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.tao_nhanh(self.khach, self.nv_a, {
				"ten_thiet_bi": "May hong khoa nhanh", "ma_thiet_bi": "ZZTB5-BROKEN2",
			})

	def test_nhan_vien_khoa_active_khoa_rong_bi_chan_khi_sua(self):
		"""Task 7, mang từ Task 6 sang — `_chan_sua_ngoai_pham_vi()` mới được
		test ở nhánh TẠO MỚI (test ngay phía trên). Ghim thêm nhánh SỬA (`save()`
		có truyền `name`) cùng kịch bản Nhân viên khoa `active=1`/`khoa_phong=""`
		đi vòng qua validate() bằng `db.set_value`. Hành vi hôm nay đã fail-closed
		(save() gọi `_chan_sua_ngoai_pham_vi()` -> `pham_vi_don()` -> ném
		PermissionError trước khi chạm ghi) nhưng chưa có test nào ghim lại."""
		frappe.db.set_value("Portal Member", self.tv_nv_a.name, "khoa_phong", "")
		frappe.set_user(self.nv_a)
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.save(self.khach, self.nv_a, {
				"name": self.may_a.name, "ten_thiet_bi": "Đổi trộm khi khoa rỗng",
			})

	# -- Review vòng 1 (Important #2) — ra_dict() tự kiểm tenant --------------

	def test_ra_dict_tra_dung_may_cua_minh(self):
		ra = thiet_bi_mod.ra_dict(self.may_a.name, self.khach)
		self.assertEqual(ra["name"], self.may_a.name)

	def test_ra_dict_tu_choi_may_benh_vien_khac(self):
		with self.assertRaises(frappe.PermissionError):
			thiet_bi_mod.ra_dict(self.may_benh_vien_khac.name, self.khach)

	def test_ra_dict_khong_ton_tai_va_may_benh_vien_khac_cung_thong_diep(self):
		"""Không phân biệt "không tồn tại" với "của bệnh viện khác" — phân
		biệt là lộ ra một docname bệnh viện khác có thật."""
		with self.assertRaises(frappe.PermissionError) as cm1:
			thiet_bi_mod.ra_dict("ZZTB5-KHONG-TON-TAI", self.khach)
		with self.assertRaises(frappe.PermissionError) as cm2:
			thiet_bi_mod.ra_dict(self.may_benh_vien_khac.name, self.khach)
		self.assertEqual(str(cm1.exception), str(cm2.exception))

	# -- Review vòng 1 (Important #3) — vat_tu không còn là oracle -----------

	def test_loc_tang_hai_tu_choi_vat_tu_benh_vien_khac(self):
		"""`vat_tu` của bệnh viện khác (có thật, `vat_tu_khac_kho`) phải cho
		kết quả GIỐNG HỆT một `vat_tu` không tồn tại — không còn phân biệt
		được, không còn là oracle dò tồn tại xuyên bệnh viện."""
		frappe.set_user(self.ql)
		ten_khac_vien = {
			r["name"] for r in thiet_bi_mod.list_rows(
				self.khach, self.ql, vat_tu=self.vat_tu_khac_kho.name
			)
		}
		ten_khong_ton_tai = {
			r["name"] for r in thiet_bi_mod.list_rows(
				self.khach, self.ql, vat_tu="ZZTB5-KHONG-TON-TAI"
			)
		}
		self.assertEqual(ten_khac_vien, ten_khong_ton_tai)
		self.assertEqual(len(ten_khac_vien), 3)


class TestThietBiEndpoint(_NenThietBi):
	"""Task 7 — bốn endpoint cổng nối `kho/thiet_bi.py` (Task 6). Không dựng
	fixture riêng, kế thừa nguyên `_NenThietBi` (xem docstring lớp đó)."""

	def test_endpoint_khong_nhan_customer_tu_client(self):
		"""Chữ ký hàm KHÔNG được có tham số customer/kho/user — nguyên tắc bất
		di bất dịch ở đầu api/kho.py. Test đọc chữ ký để một PR sau không thêm
		vào cho tiện."""
		import inspect
		for ten in ("kho_thiet_bi_list", "kho_thiet_bi_save",
		            "kho_thiet_bi_tao_nhanh", "kho_vat_tu_gan_thiet_bi"):
			tham_so = set(inspect.signature(getattr(kho_api, ten)).parameters)
			self.assertFalse(
				tham_so & {"customer", "kho", "user"},
				f"{ten} nhận định danh từ client",
			)

	def test_list_qua_endpoint_loc_dung_theo_phien(self):
		frappe.set_user(self.nv_a)
		ten = {r["name"] for r in kho_api.kho_thiet_bi_list()}
		self.assertNotIn(self.may_b.name, ten)

	def test_may_benh_vien_khac_gan_vao_vat_tu_bi_chan(self):
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_vat_tu_gan_thiet_bi(self.vat_tu.name, self.may_benh_vien_khac.name)

	def test_vat_tu_benh_vien_khac_bi_chan(self):
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_vat_tu_gan_thiet_bi(self.vat_tu_khac_kho.name, self.may_a.name)

	def test_gan_vao_vat_tu_qua_endpoint_thanh_cong(self):
		"""Đối chứng ca đúng cho ba test "bị chặn" ở trên — endpoint không chỉ
		biết chặn mà còn phải cho một cặp vật tư/máy CÙNG bệnh viện đi qua.

		Kèm chốt hình dạng: `gan_vao_vat_tu()` trả `vat_tu.ra_dict()` — dữ liệu
		VẬT TƯ, không phải thiết bị — và trường `may_su_dung` bên trong là
		DANH SÁCH DICT `{"thiet_bi": docname, "ten_thiet_bi": tên}`, khác hẳn
		hình dạng `tao()`/`sua()` NHẬN VÀO (danh sách docname trần). Một caller
		nối thẳng response này vào payload của `kho_vat_tu_sua` mà không tự
		rút `thiet_bi` ra trước sẽ gửi sai hình dạng."""
		frappe.set_user(self.ql)
		ra = kho_api.kho_vat_tu_gan_thiet_bi(self.vat_tu.name, self.may_a.name)
		self.assertEqual(ra["name"], self.vat_tu.name)
		self.assertEqual(ra["may_su_dung"], [
			{"thiet_bi": self.may_a.name, "ten_thiet_bi": self.may_a.ten_thiet_bi}
		])

	def test_loc_theo_vat_tu_cua_kho_khac_bi_chan(self):
		"""`vat_tu` là định danh do client gửi — phải qua guard TRƯỚC khi dùng
		làm bộ lọc, nếu không nó thành một kênh dò dữ liệu kho khác."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_thiet_bi_list(vat_tu=self.vat_tu_khac_kho.name)

	def test_loi_khong_lo_ten_lop_ngoai_le(self):
		"""Decorator _thiet_bi_action phải dịch mọi lỗi lạ sang tiếng Việt.

		Ca này (giữ nguyên theo brief) KHÔNG tự nó chứng minh gì: guard
		`_thiet_bi_cua_khach` chặn "KHONG-CO-THAT" bằng `PermissionError` —
		nhánh mà `_action` CHUYỂN TIẾP NGUYÊN VẸN (không dịch), không phải
		nhánh `except Exception` cần kiểm. Xem
		`test_loi_la_duoc_dich_sang_tieng_viet` ngay dưới cho ca thật sự chạm
		nhánh dịch lỗi."""
		frappe.set_user(self.ql)
		try:
			kho_api.kho_thiet_bi_save({"name": "KHONG-CO-THAT"})
		except Exception as e:
			self.assertNotIn("Traceback", str(e))

	def test_loi_la_duoc_dich_sang_tieng_viet(self):
		"""Nhánh `except Exception` thật sự của `_thiet_bi_action` — một lỗi
		LẠ (không phải ValidationError/PermissionError của chính module) từ
		`thiet_bi_mod.list_rows()` phải được dịch sang một ValidationError
		tiếng Việt, không lộ tên lớp/thông điệp gốc."""
		from unittest.mock import patch

		frappe.set_user(self.ql)
		with patch.object(
			thiet_bi_mod, "list_rows",
			side_effect=RuntimeError("'NoneType' object has no attribute 'lft'"),
		):
			with self.assertRaises(frappe.ValidationError) as cm:
				kho_api.kho_thiet_bi_list()
		thong_diep = str(cm.exception)
		self.assertIn("thiết bị", thong_diep)
		self.assertNotIn("RuntimeError", thong_diep)
		self.assertNotIn("NoneType", thong_diep)

	def test_save_qua_endpoint_tao_moi(self):
		frappe.set_user(self.ql)
		ra = kho_api.kho_thiet_bi_save({
			"ten_thiet_bi": "May moi qua endpoint", "ma_thiet_bi": "ZZTB5-EP-NEW",
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", ra["name"], "customer"), self.khach
		)

	def test_save_sua_may_benh_vien_khac_bi_chan(self):
		"""`name` trong payload cũng là định danh do client gửi — phải qua
		guard trước khi save() chạm doc, đúng nguyên tắc đầu file."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_thiet_bi_save({
				"name": self.may_benh_vien_khac.name, "ten_thiet_bi": "Đổi trộm qua endpoint",
			})

	def test_save_khoa_phong_benh_vien_khac_bi_chan(self):
		"""Vòng sửa 1 (Important #1) — `khoa_phong` trong payload cũng là
		định danh do client gửi, phải qua `_khoa_cua_khach()` TRƯỚC khi chạm
		Link field `Customer Equipment.khoa_phong` (get_invalid_links() +
		validate() của doctype). Khoa của MỘT bệnh viện khác phải bị chặn
		ngay ở tầng endpoint, không rơi xuống để lộ oracle phân biệt "không
		tồn tại" / "khác bệnh viện" qua thông điệp Frappe/controller."""
		kp_khac = frappe.get_doc({
			"doctype": "Customer Department", "customer": KHACH_KHAC,
			"ten_khoa_phong": "ZZTB5K Khoa", "ma_khoa": "ZZT5KX",
		}).insert(ignore_permissions=True)
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_thiet_bi_save({
				"ten_thiet_bi": "May gan khoa vien khac", "ma_thiet_bi": "ZZTB5-EP-KPX",
				"khoa_phong": kp_khac.name,
			})

	def test_save_khoa_phong_dict_khong_bi_hieu_thanh_filters(self):
		"""Cùng khuôn `test_vat_tu_dict_khong_bi_hieu_thanh_filters` — dựng
		dict khớp CHÍNH `self.kp_a` (khoa của người gọi) để chứng minh có lỗ
		thật nếu thiếu ép `str()`, không phải một dict ngẫu nhiên."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_thiet_bi_save({
				"ten_thiet_bi": "May khoa dict", "ma_thiet_bi": "ZZTB5-EP-KPD",
				"khoa_phong": {"customer": self.khach},
			})

	def test_save_khoa_phong_cung_benh_vien_van_bi_ep_ve_khoa_minh(self):
		"""Đối chứng: guard mới KHÔNG phá BR-TB-6. Nhân viên khoa A gửi khoa B
		(CÙNG bệnh viện, qua được `_khoa_cua_khach()`) vẫn phải bị
		`_khoa_ep_theo_phien()` (trong `thiet_bi.save()`) ép về khoa A —
		guard chỉ chặn khoa của bệnh viện KHÁC, không nới quyền chọn khoa."""
		frappe.set_user(self.nv_a)
		ra = kho_api.kho_thiet_bi_save({
			"ten_thiet_bi": "May qua endpoint bi ep khoa", "ma_thiet_bi": "ZZTB5-EP-EP",
			"khoa_phong": self.kp_b.name,
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", ra["name"], "khoa_phong"),
			self.kp_a.name,
		)

	def test_tao_nhanh_qua_endpoint(self):
		frappe.set_user(self.nv_a)
		ra = kho_api.kho_thiet_bi_tao_nhanh({
			"ten_thiet_bi": "May nhanh qua endpoint", "ma_thiet_bi": "ZZTB5-EP-NHANH",
		})
		self.assertEqual(
			frappe.db.get_value("Customer Equipment", ra["name"], "khoa_phong"), self.kp_a.name
		)

	def test_vat_tu_dict_khong_bi_hieu_thanh_filters(self):
		"""Ép kiểu tại biên (mang từ Task 6 sang) — nếu một endpoint lỡ chuyển
		thẳng một payload chưa ép kiểu cho get_value(doctype, name, field),
		một `dict` sẽ bị hiểu là FILTERS chứ không phải docname.

		SỬA (vòng review 1, Important #2) — bản trước dùng khuôn
		try/except-không-assertRaises nên KHÔNG BAO GIỜ đỏ được: dict lọc
		`{"kho": ["!=", "KHONG-TON-TAI"]}` luôn khớp MỘT hàng nào đó (hoặc
		không khớp gì), cả hai nhánh đều không ném ngoại lệ nên test luôn
		xanh dù có ép `str()` hay không — đúng lỗi mà chính test này lẽ ra
		phải canh.

		Dựng dict khớp CHÍNH `self.vat_tu` (kho của người gọi) thay vì một
		dict ngẫu nhiên — đây là điều kiện để chứng minh có lỗ thật: nếu
		endpoint không ép `str()` trước `_vat_tu_cua_kho()`, dict này sẽ được
		`frappe.db.get_value` diễn giải thành FILTERS, khớp đúng vat_tu của
		CHÍNH người gọi, và lặng lẽ ĐI QUA guard — không phải một dict bất kỳ
		nào cũng làm được việc đó. Có ép `str()` (code hiện tại): dict bị ép
		thành một chuỗi vô nghĩa, không khớp docname nào, guard ném
		PermissionError."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_thiet_bi_list(vat_tu={"kho": self.kho.name})


class TestPhieuXuatNhanMay(_NenThietBi):
	"""Task 8 — `kho_phieu_xuat_save` nhận `thiet_bi` trên từng dòng và
	`thiet_bi_mac_dinh` ở header, trả cảnh báo mềm `canh_bao_thiet_bi` cho
	giao diện.

	SỬA (so với brief gốc, xem task-8-report.md): brief định KHÔNG kiểm sở
	hữu máy ở tầng endpoint, với điều kiện tự đặt ra "chấp nhận được NẾU
	controller chặn với CÙNG thông điệp cho mọi ca". Kiểm thực tế cho thấy
	SAI: `Document._validate_links()` (chạy TRƯỚC `validate()`) chết với
	`LinkValidationError` tiếng Anh cho máy KHÔNG TỒN TẠI, còn
	`_validate_thiet_bi()` chết với `ValidationError` tiếng Việt cho máy CÓ
	THẬT của bệnh viện khác — hai loại lỗi khác nhau, một oracle dò tồn tại
	docname. Endpoint giờ tự guard bằng `_thiet_bi_cua_khach()` (Task 7,
	dùng lại nguyên) cho cả `thiet_bi` từng dòng lẫn `thiet_bi_mac_dinh` —
	xem `test_may_khong_ton_tai_va_may_vien_khac_cung_mot_loi` cho bằng
	chứng oracle đã đóng."""

	def test_dong_ghi_dung_may(self):
		# `self.kp_a` (từ `_NenThietBi`) không gắn `kho` — đúng mô hình "khoa
		# thuộc bệnh viện" (Task 2/3). Nhưng `_validate_khoa_phong_thuoc_kho`
		# trong `customer_stock_issue.py` VẪN so trực tiếp `Customer
		# Department.kho == self.kho` (mô hình CŨ, ngoài phạm vi BR-TB — xem
		# đúng ghi chú tại `test_tb2_phieu_xuat.py::_khoa()`), nên một khoa
		# không gắn kho làm MỌI phiếu xuất trỏ tới nó ném lỗi "không thuộc
		# kho" — không liên quan gì tới máy. Gắn kho cho kp_a NGAY TRONG test
		# này (không sửa `_NenThietBi` dùng chung cho Task 6/7/11) để phiếu
		# lưu được, đúng khuôn `test_tb2_phieu_xuat.py::_khoa()`.
		frappe.db.set_value("Customer Department", self.kp_a.name, "kho", self.kho.name)
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"khoa_phong": self.kp_a.name, "thiet_bi_mac_dinh": self.may_a.name,
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
			           "so_luong": 2, "thiet_bi": self.may_a.name}],
		})
		doc = frappe.get_doc("Customer Stock Issue", ra["name"])
		self.assertEqual(doc.items[0].thiet_bi, self.may_a.name)
		self.assertEqual(doc.thiet_bi_mac_dinh, self.may_a.name)

	def test_may_ngoai_danh_muc_van_luu_duoc_kem_canh_bao(self):
		self.vat_tu.set("may_su_dung", [{"thiet_bi": self.may_a.name}])
		self.vat_tu.save(ignore_permissions=True)
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
			           "so_luong": 1, "thiet_bi": self.may_chung.name}],
		})
		self.assertTrue(ra["name"])
		self.assertTrue(ra["canh_bao_thiet_bi"])

	def test_khong_co_canh_bao_thi_khoa_van_ton_tai_va_rong(self):
		"""SPA đọc thẳng `ket_qua.canh_bao_thiet_bi` — thiếu khoá sẽ vỡ giao
		diện, nên khoá phải LUÔN có, kể cả khi không có cảnh báo nào."""
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo, "so_luong": 1}],
		})
		self.assertEqual(ra["canh_bao_thiet_bi"], [])

	def test_doc_lai_phieu_cung_co_khoa_canh_bao_rong(self):
		"""`_phieu_to_dict` còn được `kho_phieu_get` dùng lại trên một doc vừa
		`get_doc()` từ CSDL — doc đó KHÔNG có `flags.canh_bao_thiet_bi` (flag
		chỉ được set trong validate() của LẦN LƯU vừa rồi, không lưu xuống
		DB). Khoá vẫn phải là `[]`, không phải lỗi khi đọc `doc.flags` thiếu,
		cũng không phải vắng mặt."""
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo, "so_luong": 1}],
		})
		doc_lai = kho_api.kho_phieu_get("Customer Stock Issue", ra["name"])
		self.assertEqual(doc_lai["canh_bao_thiet_bi"], [])

	def test_may_benh_vien_khac_bi_chan_o_tang_endpoint(self):
		"""Đổi tên so với brief gốc ("...o_tang_controller") — chốt chặn thật
		sự nằm ở GUARD ENDPOINT (`_thiet_bi_cua_khach`), không phải
		controller (xem lý do ở docstring lớp). `assertRaises(frappe.
		PermissionError)` của brief ĐÚNG với code đã sửa: guard ném
		PermissionError, và `_phieu_action` chuyển tiếp nguyên vẹn (không
		dịch) — cùng khuôn `test_save_sua_may_benh_vien_khac_bi_chan`."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_phieu_xuat_save({
				"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
				"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
				           "so_luong": 1, "thiet_bi": self.may_benh_vien_khac.name}],
			})

	def test_thiet_bi_mac_dinh_may_benh_vien_khac_bi_chan(self):
		"""Mang từ Task 8 sang (việc còn thiếu, chỉ thị Task 11): chưa có ca
		nào ghim TRỰC TIẾP `thiet_bi_mac_dinh` nhận một docname THẬT của bệnh
		viện khác — ca dict-filter ngay dưới chỉ kiểm nhánh dict cho header,
		không kiểm nhánh docname thật này."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_phieu_xuat_save({
				"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
				"thiet_bi_mac_dinh": self.may_benh_vien_khac.name,
				"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo, "so_luong": 1}],
			})

	def test_may_khong_ton_tai_va_may_vien_khac_cung_mot_loi(self):
		"""Bằng chứng oracle đã đóng (lý do đổi thiết kế, xem docstring lớp
		và task-8-report.md): trước khi có guard, một máy KHÔNG TỒN TẠI và
		một máy CÓ THẬT của bệnh viện khác ra HAI loại lỗi/HAI thông điệp
		khác nhau (`LinkValidationError` tiếng Anh của
		`Document._validate_links()` so với `ValidationError` tiếng Việt của
		`_validate_thiet_bi()`) — đủ để dò tồn tại docname `Customer
		Equipment` xuyên bệnh viện. Với guard `_thiet_bi_cua_khach()` chặn
		TRƯỚC khi giá trị chạm `insert()`, cả hai ca phải ra ĐÚNG CÙNG một
		loại ngoại lệ và CÙNG một thông điệp."""
		frappe.set_user(self.ql)
		ket_qua = []
		for thiet_bi in ("TBK-KHONG-TON-TAI-999", self.may_benh_vien_khac.name):
			with self.assertRaises(Exception) as cm:
				kho_api.kho_phieu_xuat_save({
					"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
					"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
					           "so_luong": 1, "thiet_bi": thiet_bi}],
				})
			ket_qua.append((type(cm.exception), str(cm.exception)))
		self.assertEqual(ket_qua[0], ket_qua[1])
		self.assertEqual(ket_qua[0][0], frappe.PermissionError)

	def test_dong_thiet_bi_dict_khong_bi_hieu_thanh_filters(self):
		"""Cùng khuôn `test_save_khoa_phong_dict_khong_bi_hieu_thanh_filters`
		(Task 7) — nếu endpoint KHÔNG ép `str()` trước khi đưa `thiet_bi` của
		dòng vào `_thiet_bi_cua_khach()`/`doc.append()`, guard đó (dùng
		`frappe.db.get_value(doctype, thiet_bi, "customer")`) sẽ tự diễn
		giải một `dict` thành FILTERS, khớp một `Customer Equipment` THẬT
		rồi ÂM THẦM cho qua guard — dựng dict khớp CHÍNH `self.may_a` (máy
		của người gọi, không phải máy ngẫu nhiên) để đây là một lỗ THẬT: nó
		phải khớp được máy nào đó thì mới chứng minh có lỗ, một dict không
		khớp gì thì dù có ép `str()` hay không hai nhánh đều ném lỗi, không
		phân biệt được. Có `str()` (code hiện tại): dict bị ép thành chuỗi
		vô nghĩa, guard không tìm thấy máy nào khớp, ném `PermissionError`
		đúng thông điệp "Máy không thuộc đơn vị bạn."."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_phieu_xuat_save({
				"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
				"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
				           "so_luong": 1, "thiet_bi": {"ma_thiet_bi": self.may_a.ma_thiet_bi}}],
			})

	def test_thiet_bi_mac_dinh_dict_khong_bi_hieu_thanh_filters(self):
		"""Cùng lý do như test ngay trên, áp cho field HEADER
		`thiet_bi_mac_dinh` — cũng đi qua `_thiet_bi_cua_khach()` trước khi
		gán vào doc."""
		frappe.set_user(self.ql)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_phieu_xuat_save({
				"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
				"thiet_bi_mac_dinh": {"ma_thiet_bi": self.may_a.ma_thiet_bi},
				"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo, "so_luong": 1}],
			})


class TestExcelCotMaMay(_NenThietBi):
	"""Cột Mã máy trong file nhập phiếu xuất hàng loạt.

	Test dựng workbook THẬT bằng openpyxl rồi cho `doc_file` đọc bytes — đúng
	đường mà người dùng đi. Không gọi thẳng hàm nội bộ nào.
	"""

	def _file(self, ma_may):
		import io as _io
		import openpyxl
		wb = openpyxl.load_workbook(_io.BytesIO(dong_phieu.build_mau_xlsx("xuat")))
		ws = wb.active
		tieu_de = [c.value for c in ws[1]]
		dong = [""] * len(tieu_de)
		dong[tieu_de.index("Mã vật tư")] = self.vat_tu.ma_vat_tu
		dong[tieu_de.index("Số lô")] = self.lo
		dong[tieu_de.index("Số lượng")] = 2
		dong[tieu_de.index("Mã máy")] = ma_may
		ws.append(dong)
		buf = _io.BytesIO()
		wb.save(buf)
		return buf.getvalue()

	def test_file_mau_co_cot_ma_may(self):
		import io as _io
		import openpyxl
		wb = openpyxl.load_workbook(_io.BytesIO(dong_phieu.build_mau_xlsx("xuat")))
		self.assertIn("Mã máy", [c.value for c in wb.active[1]])

	def test_ma_may_khong_bat_buoc(self):
		"""Cột mới KHÔNG được vào REQUIRED — mọi file mẫu cũ đang lưu trên máy
		khách phải nạp lại được, nếu không đây là hồi quy chứ không phải tính
		năng."""
		self.assertNotIn("ma_thiet_bi", dong_phieu.REQUIRED["xuat"])

	def test_ma_dung_ra_docname(self):
		"""SỬA so với brief gốc (task-11-brief.md dùng literal "XN500-01" —
		đó là mã trong fixture của plan doc gốc/test_tb1_doctype.py, KHÔNG
		phải mã trong `_NenThietBi` thật của file này, nơi `may_a.ma_thiet_bi
		== "ZZTB5-MAY-A"`. Dùng thẳng `self.may_a.ma_thiet_bi` để test không
		phụ thuộc một hằng số trùng hợp."""
		ra = dong_phieu.doc_file(self._file(self.may_a.ma_thiet_bi), self.kho.name, "xuat")
		row = ra["rows"][0]
		self.assertEqual(row["thiet_bi"], self.may_a.name)
		self.assertEqual(row["loi"], [])

	def test_o_trong_la_khong_gan_may_khong_phai_loi(self):
		ra = dong_phieu.doc_file(self._file(""), self.kho.name, "xuat")
		row = ra["rows"][0]
		self.assertEqual(row["thiet_bi"], "")
		self.assertEqual(row["loi"], [])
		self.assertEqual(row["trang_thai"], "khop")

	def test_ma_sai_vao_loi_cua_dong_chu_khong_bi_bo_qua(self):
		"""Bỏ qua im lặng = ghi sổ thiếu máy mà người dùng tin là đã có."""
		ra = dong_phieu.doc_file(self._file("KHONG-CO"), self.kho.name, "xuat")
		row = ra["rows"][0]
		self.assertTrue(any("KHONG-CO" in x for x in row["loi"]))
		self.assertEqual(row["trang_thai"], "loi")
		self.assertEqual(row["thiet_bi"], "")

	def test_ma_may_benh_vien_khac_bao_loi_giong_ma_khong_ton_tai(self):
		"""Không được lộ ra rằng mã đó CÓ THẬT ở bệnh viện khác — thông điệp
		phải cùng khuôn, chỉ khác phần mã được trích dẫn."""
		la = dong_phieu.doc_file(
			self._file(self.may_benh_vien_khac.ma_thiet_bi), self.kho.name, "xuat"
		)["rows"][0]
		bia = dong_phieu.doc_file(self._file("HOAN-TOAN-BIA"), self.kho.name, "xuat")["rows"][0]
		# Không để hai vế trống cùng lúc biến phép so thành bất biến hình
		# thức — nếu tra cứu bị bỏ qua hoàn toàn, cả hai `loi` đều rỗng và
		# assertEqual dưới đây vẫn xanh dù không kiểm được gì.
		self.assertTrue(la["loi"])
		self.assertTrue(bia["loi"])
		chuan = lambda t: re.sub(r'"[^"]*"', '"X"', t)
		self.assertEqual(
			[chuan(x) for x in la["loi"]], [chuan(x) for x in bia["loi"]]
		)

	def test_ma_khong_phan_biet_hoa_thuong(self):
		"""`Customer Equipment.validate()` tự viết hoa `ma_thiet_bi` — mã gõ
		thường trong file vẫn phải khớp, đúng như ca `_match_vat_tu` (không
		phân biệt hoa thường) đã làm cho vật tư."""
		ra = dong_phieu.doc_file(
			self._file(self.may_a.ma_thiet_bi.lower()), self.kho.name, "xuat"
		)
		row = ra["rows"][0]
		self.assertEqual(row["thiet_bi"], self.may_a.name)
		self.assertEqual(row["loi"], [])

	def test_file_mau_cu_thieu_cot_ma_may_van_nap_duoc(self):
		"""Requirement #1 kiểm ở HÀNH VI, không chỉ ở hằng số REQUIRED: một
		file mẫu khách đã tải TRƯỚC Task 11 (không có cột "Mã máy" — mô
		phỏng bằng cách xoá hẳn cột đó khỏi workbook, không chỉ để trống) vẫn
		phải nạp được, dòng vẫn "khop" và không kèm lỗi nào."""
		import io as _io
		import openpyxl
		wb = openpyxl.load_workbook(_io.BytesIO(dong_phieu.build_mau_xlsx("xuat")))
		ws = wb.active
		tieu_de = [c.value for c in ws[1]]
		ws.delete_cols(tieu_de.index("Mã máy") + 1)
		tieu_de_moi = [c.value for c in ws[1]]
		self.assertNotIn("Mã máy", tieu_de_moi)
		dong = [""] * len(tieu_de_moi)
		dong[tieu_de_moi.index("Mã vật tư")] = self.vat_tu.ma_vat_tu
		dong[tieu_de_moi.index("Số lô")] = self.lo
		dong[tieu_de_moi.index("Số lượng")] = 2
		ws.append(dong)
		buf = _io.BytesIO()
		wb.save(buf)
		ra = dong_phieu.doc_file(buf.getvalue(), self.kho.name, "xuat")
		row = ra["rows"][0]
		self.assertEqual(row["trang_thai"], "khop")
		self.assertEqual(row["loi"], [])
		self.assertEqual(row["thiet_bi"], "")

	def test_doc_file_hoat_dong_duoi_phien_khach_hang_khong_phai_admin(self):
		"""`doc_file` tra `Customer Equipment` bằng `frappe.get_all` — hàm đó
		luôn chạy `ignore_permissions=True` nên hoạt động ĐÚNG dù không có
		DocPerm cho role Customer trên doctype này (Global Constraint 2 của
		kế hoạch). Test này neo điều đó bằng hành vi thay vì bằng tài liệu:
		nếu ai đó đổi `get_all` thành `get_list`, ca này phải đỏ vì phiên
		Website User không có quyền đọc `Customer Equipment`."""
		frappe.set_user(self.ql)
		ra = dong_phieu.doc_file(self._file(self.may_a.ma_thiet_bi), self.kho.name, "xuat")
		row = ra["rows"][0]
		self.assertEqual(row["thiet_bi"], self.may_a.name)
		self.assertEqual(row["loi"], [])


class TestExportGiuLaiMaMay(_NenThietBi):
	"""Task 11, phần xuất tệp: `export_rows`/`build_export_xlsx` dùng chung
	`COLUMNS["xuat"]` với `doc_file` — nếu không tự điền lại "Mã máy", một
	vòng xuất-rồi-nạp-lại (khách tải phiếu về sửa rồi nạp lại) sẽ âm thầm
	đánh rơi mọi gán máy đã lưu trên phiếu xuất."""

	def test_export_giu_lai_ma_may(self):
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo,
			           "so_luong": 1, "thiet_bi": self.may_chung.name}],
		})
		content = dong_phieu.build_export_xlsx("Customer Stock Issue", ra["name"])
		import io as _io
		import openpyxl
		wb = openpyxl.load_workbook(_io.BytesIO(content))
		ws = wb.active
		tieu_de = [c.value for c in ws[1]]
		idx = tieu_de.index("Mã máy")
		dong_du_lieu = list(ws.iter_rows(min_row=2, max_row=2))[0]
		self.assertEqual(dong_du_lieu[idx].value, self.may_chung.ma_thiet_bi)

		# Đóng trọn vòng: tệp vừa xuất phải NẠP LẠI đúng ra cùng một máy —
		# hai test độc lập (xuất viết đúng mã / đọc khớp đúng docname) không
		# tự chứng minh việc nối chúng lại vẫn đúng.
		ra_doc = dong_phieu.doc_file(content, self.kho.name, "xuat")
		self.assertEqual(ra_doc["rows"][0]["thiet_bi"], self.may_chung.name)
		self.assertEqual(ra_doc["rows"][0]["loi"], [])

	def test_export_dong_khong_co_may_de_trong_khong_loi(self):
		frappe.set_user(self.ql)
		ra = kho_api.kho_phieu_xuat_save({
			"ngay": frappe.utils.today(), "loai_xuat": "Xuất sử dụng",
			"items": [{"vat_tu": self.vat_tu.name, "so_lo": self.lo, "so_luong": 1}],
		})
		content = dong_phieu.build_export_xlsx("Customer Stock Issue", ra["name"])
		import io as _io
		import openpyxl
		wb = openpyxl.load_workbook(_io.BytesIO(content))
		ws = wb.active
		tieu_de = [c.value for c in ws[1]]
		idx = tieu_de.index("Mã máy")
		dong_du_lieu = list(ws.iter_rows(min_row=2, max_row=2))[0]
		self.assertFalse(dong_du_lieu[idx].value)
