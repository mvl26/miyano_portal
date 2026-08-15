"""Brief 2026-08-15 — phân trang toàn cổng.

Hai nhóm test theo đúng hai rủi ro brief nêu là nặng nhất, không lỗi nào
tự báo ra nếu hỏng:

1. Ba endpoint KIÊM HAI VAI (`kho_vat_tu_list`/`kho_ncc_list`/
   `kho_khoa_phong_list`) — không truyền `limit` phải trả ĐỦ danh sách (nuôi
   dropdown NhatKy.vue/BaoCaoNXT.vue); truyền `limit` mới cắt trang.
2. Xuất Excel (`kho_bao_cao_excel`) phải LUÔN xuất toàn bộ, không theo
   trang JSON đang xem.

Cộng thêm: `tong`/hình dạng dict của các endpoint đổi hình dạng
(`portal_order_history`, `portal_invoices`, `kho_phieu_list`), và tiebreak
thứ tự tất định.
"""

import io

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import load_workbook

from miyano_portal.api import kho as kho_api
from miyano_portal.api import portal
from miyano_portal.setup.seed_demo import COMPANY, PRICE_LIST
from miyano_portal.setup.seed_kho_demo import seed_kho_demo

BM_USER = "bvbm@demo.miyano"
BVBM = "Bệnh viện Bạch Mai"

SO_LUONG_DU_NHIEU = 55  # > mọi kích thước trang (10/20/50) đang có trong UI


def _tao_don_test(customer, workflow_state, ngay=None):
    """Dựng nhanh một Sales Order Draft cho test lọc trạng thái — cùng
    khuôn `_tao_so_bao_gia` (test_e6_mua_le.py): `insert(ignore_permissions
    =True)` rồi GHI ĐÈ `workflow_state` bằng `db.set_value` (BẪY: gán trước
    insert() bị workflow engine tự ghi đè state mặc định ngay khi insert)."""
    so = frappe.new_doc("Sales Order")
    so.customer = customer
    so.company = COMPANY
    so.transaction_date = ngay or frappe.utils.today()
    so.delivery_date = frappe.utils.add_days(so.transaction_date, 7)
    so.selling_price_list = PRICE_LIST
    so.custom_nguon_don = "Client Portal"
    so.custom_loai_don = "Mua lẻ"
    so.append("items", {"item_code": "VT0005", "qty": 1, "rate": 1000, "delivery_date": so.delivery_date})
    so.taxes = []
    so.taxes_and_charges = None
    so.insert(ignore_permissions=True)
    frappe.db.set_value("Sales Order", so.name, "workflow_state", workflow_state, update_modified=False)
    so.reload()
    return so


class TestKhoVatTuListKiemHaiVai(FrappeTestCase):
    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        frappe.db.delete("Customer Warehouse Item", {"kho": self.K, "ma_vat_tu": ["like", "PT-VT-%"]})
        for i in range(SO_LUONG_DU_NHIEU):
            frappe.get_doc({
                "doctype": "Customer Warehouse Item",
                "kho": self.K,
                "ma_vat_tu": f"PT-VT-{i:03d}",
                "ten_vat_tu": f"Vật tư phân trang {i:03d}",
                "dvt": "Cái",
            }).insert(ignore_permissions=True)

    def test_khong_truyen_limit_tra_du_danh_sach_cho_dropdown(self):
        """NhatKy.vue/BaoCaoNXT.vue gọi kho_vat_tu_list KHÔNG truyền limit
        để đổ dropdown lọc — endpoint phải trả BARE LIST đủ mọi vật tư, dù
        có > 50 dòng. Đây chính là bộ lọc brief cảnh báo sẽ hỏng ÂM THẦM
        nếu ai đó phân trang vô điều kiện."""
        frappe.set_user(BM_USER)
        try:
            rows = kho_api.kho_vat_tu_list()
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(rows, list, "không truyền limit phải trả list trần, không phải dict")
        ma_list = {r["ma_vat_tu"] for r in rows}
        self.assertGreaterEqual(
            len(rows), SO_LUONG_DU_NHIEU,
            "dropdown phải thấy ĐỦ vật tư, không bị cắt còn một trang",
        )
        for i in range(SO_LUONG_DU_NHIEU):
            self.assertIn(f"PT-VT-{i:03d}", ma_list)

    def test_truyen_limit_tra_dict_va_cat_dung_trang(self):
        frappe.set_user(BM_USER)
        try:
            trang_1 = kho_api.kho_vat_tu_list(limit=10, start=0)
            trang_2 = kho_api.kho_vat_tu_list(limit=10, start=10)
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(trang_1, dict, "có limit phải trả dict {rows, tong}")
        self.assertEqual(len(trang_1["rows"]), 10)
        self.assertGreaterEqual(trang_1["tong"], SO_LUONG_DU_NHIEU)
        self.assertEqual(len(trang_2["rows"]), 10)
        # Hai trang không được trùng dòng nào — order_by có tiebreak `name`.
        ten_1 = {r["name"] for r in trang_1["rows"]}
        ten_2 = {r["name"] for r in trang_2["rows"]}
        self.assertFalse(ten_1 & ten_2, "hai trang liền kề không được lẫn dòng")


class TestKhoNccListKiemHaiVai(FrappeTestCase):
    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        frappe.db.delete("Customer Supplier", {"kho": self.K, "ten_ncc": ["like", "NCC phân trang%"]})
        for i in range(SO_LUONG_DU_NHIEU):
            frappe.get_doc({
                "doctype": "Customer Supplier",
                "kho": self.K,
                "ten_ncc": f"NCC phân trang {i:03d}",
            }).insert(ignore_permissions=True)

    def test_khong_truyen_limit_tra_du_danh_sach_cho_dropdown(self):
        frappe.set_user(BM_USER)
        try:
            rows = kho_api.kho_ncc_list()
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(rows, list)
        self.assertGreaterEqual(len(rows), SO_LUONG_DU_NHIEU)

    def test_truyen_limit_tra_dict_va_cat_dung_trang(self):
        frappe.set_user(BM_USER)
        try:
            ket_qua = kho_api.kho_ncc_list(limit=20, start=0)
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(ket_qua, dict)
        self.assertEqual(len(ket_qua["rows"]), 20)
        self.assertGreaterEqual(ket_qua["tong"], SO_LUONG_DU_NHIEU)


class TestKhoKhoaPhongListKiemHaiVai(FrappeTestCase):
    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        frappe.db.delete("Customer Department", {"kho": self.K, "ten_khoa_phong": ["like", "Khoa phân trang%"]})
        for i in range(SO_LUONG_DU_NHIEU):
            frappe.get_doc({
                "doctype": "Customer Department",
                "kho": self.K,
                "ten_khoa_phong": f"Khoa phân trang {i:03d}",
            }).insert(ignore_permissions=True)

    def test_khong_truyen_limit_tra_du_danh_sach_cho_dropdown(self):
        frappe.set_user(BM_USER)
        try:
            rows = kho_api.kho_khoa_phong_list()
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(rows, list)
        self.assertGreaterEqual(len(rows), SO_LUONG_DU_NHIEU)

    def test_truyen_limit_tra_dict_va_cat_dung_trang(self):
        frappe.set_user(BM_USER)
        try:
            ket_qua = kho_api.kho_khoa_phong_list(limit=50, start=0)
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(ket_qua, dict)
        self.assertEqual(len(ket_qua["rows"]), 50)
        self.assertGreaterEqual(ket_qua["tong"], SO_LUONG_DU_NHIEU)


class TestKhoTonPhanTrang(FrappeTestCase):
    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        self.VT = self.kho["vt_bm"]
        frappe.db.delete("Customer Stock Ledger Entry", {"kho": self.K})
        frappe.db.delete("Customer Stock Lot Balance", {"kho": self.K})

    def test_khong_truyen_limit_van_tra_list_tran(self):
        frappe.set_user(BM_USER)
        try:
            rows = kho_api.kho_ton()
        finally:
            frappe.set_user("Administrator")
        self.assertIsInstance(rows, list)


class TestBaoCaoNxtPhanTrang(FrappeTestCase):
    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        frappe.db.delete("Customer Stock Ledger Entry", {"kho": self.K})
        frappe.db.delete("Customer Stock Lot Balance", {"kho": self.K})
        frappe.db.delete("Customer Warehouse Item", {"kho": self.K, "ma_vat_tu": ["like", "PT-NXT-%"]})
        today = frappe.utils.getdate(frappe.utils.today())
        self.tu_ngay = frappe.utils.add_days(today, -10)
        self.den_ngay = today
        for i in range(15):
            vt = frappe.get_doc({
                "doctype": "Customer Warehouse Item",
                "kho": self.K, "ma_vat_tu": f"PT-NXT-{i:03d}",
                "ten_vat_tu": f"Vật tư NXT phân trang {i:03d}", "dvt": "Cái",
            }).insert(ignore_permissions=True)
            doc = frappe.get_doc({
                "doctype": "Customer Stock Receipt",
                "kho": self.K, "ngay": today, "loai_nhap": "Nhập khác",
                "nguoi_giao": "Test",
                "items": [{
                    "vat_tu": vt.name, "so_lo": f"LO-{i}", "so_luong": 10, "don_gia": 1000,
                }],
            })
            doc.insert(ignore_permissions=True)
            doc.submit()

    def test_muc_vat_tu_phan_trang_dung(self):
        frappe.set_user(BM_USER)
        try:
            ket_qua = kho_api.kho_bao_cao_nxt(
                tu_ngay=self.tu_ngay, den_ngay=self.den_ngay, limit=5, start=0,
            )
        finally:
            frappe.set_user("Administrator")
        self.assertEqual(ket_qua["muc"], "vat_tu")
        self.assertEqual(len(ket_qua["rows"]), 5)
        self.assertGreaterEqual(ket_qua["tong"], 15)

    def test_muc_lo_khong_bi_anh_huong_boi_limit(self):
        """Màn chi tiết (bung MỘT vật tư xuống lô) KHÔNG phân trang — theo
        giả định đã chốt với chủ dự án. `limit` bị bỏ qua khi có `vat_tu`."""
        vat_tu_name = frappe.db.get_value(
            "Customer Warehouse Item", {"kho": self.K, "ma_vat_tu": "PT-NXT-000"}, "name"
        )
        frappe.set_user(BM_USER)
        try:
            ket_qua = kho_api.kho_bao_cao_nxt(
                tu_ngay=self.tu_ngay, den_ngay=self.den_ngay, vat_tu=vat_tu_name, limit=1,
            )
        finally:
            frappe.set_user("Administrator")
        self.assertEqual(ket_qua["muc"], "lo")
        self.assertNotIn("tong", ket_qua, "mức lô là chi tiết, không có khái niệm tong/trang")


class TestBaoCaoExcelLuonXuatToanBo(FrappeTestCase):
    """DoD brief 2026-08-15: xuất Excel/PDF của BaoCaoNXT LUÔN xuất TOÀN
    BỘ, không theo trang JSON đang xem."""

    def setUp(self):
        self.kho = seed_kho_demo()
        self.K = self.kho["kho_bm"]
        frappe.db.delete("Customer Stock Ledger Entry", {"kho": self.K})
        frappe.db.delete("Customer Stock Lot Balance", {"kho": self.K})
        frappe.db.delete("Customer Warehouse Item", {"kho": self.K, "ma_vat_tu": ["like", "PT-XLS-%"]})
        today = frappe.utils.getdate(frappe.utils.today())
        self.tu_ngay = frappe.utils.add_days(today, -10)
        self.den_ngay = today
        # 12 vật tư có phát sinh — nhiều hơn kích thước trang nhỏ (10) test
        # sẽ dùng cho endpoint JSON, để phân biệt được "trang" với "toàn bộ".
        for i in range(12):
            vt = frappe.get_doc({
                "doctype": "Customer Warehouse Item",
                "kho": self.K, "ma_vat_tu": f"PT-XLS-{i:03d}",
                "ten_vat_tu": f"Vật tư xuất Excel {i:03d}", "dvt": "Cái",
            }).insert(ignore_permissions=True)
            doc = frappe.get_doc({
                "doctype": "Customer Stock Receipt",
                "kho": self.K, "ngay": today, "loai_nhap": "Nhập khác",
                "nguoi_giao": "Test",
                "items": [{
                    "vat_tu": vt.name, "so_lo": f"LO-{i}", "so_luong": 10, "don_gia": 1000,
                }],
            })
            doc.insert(ignore_permissions=True)
            doc.submit()

    def test_export_khong_bi_gioi_han_boi_page_size_cua_man_json(self):
        frappe.set_user(BM_USER)
        try:
            # Màn hình đang xem TRANG NHỎ (5 dòng/trang) — endpoint JSON
            # phải trả đúng 5, nhưng nút xuất Excel không được biết/không
            # được dùng con số này.
            trang_json = kho_api.kho_bao_cao_nxt(
                tu_ngay=self.tu_ngay, den_ngay=self.den_ngay, limit=5, start=0,
            )
            self.assertEqual(len(trang_json["rows"]), 5)
            self.assertGreaterEqual(trang_json["tong"], 12)

            kho_api.kho_bao_cao_excel(
                loai="nxt", tu_ngay=str(self.tu_ngay), den_ngay=str(self.den_ngay),
            )
            content = frappe.local.response.filecontent
        finally:
            frappe.local.response.clear()
            frappe.set_user("Administrator")

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        so_dong_du_lieu = ws.max_row - 1  # trừ dòng tiêu đề
        self.assertGreaterEqual(
            so_dong_du_lieu, 12,
            "xuất Excel phải có ĐỦ toàn bộ 12 vật tư, không chỉ 5 dòng của trang JSON",
        )
        self.assertEqual(so_dong_du_lieu, trang_json["tong"], "xuất Excel phải khớp đúng tong, không phải page size")


class TestPortalOrderHistoryVaInvoicesHinhDangMoi(FrappeTestCase):
    def test_portal_order_history_tra_dict_co_tong(self):
        frappe.set_user(BM_USER)
        try:
            ket_qua = portal.portal_order_history()
        finally:
            frappe.set_user("Administrator")
        self.assertIn("rows", ket_qua)
        self.assertIn("tong", ket_qua)
        self.assertIsInstance(ket_qua["rows"], list)
        self.assertGreaterEqual(ket_qua["tong"], len(ket_qua["rows"]))

    def test_portal_invoices_tra_dict_co_tong_va_thong_ke_cong_no(self):
        frappe.set_user(BM_USER)
        try:
            ket_qua = portal.portal_invoices()
        finally:
            frappe.set_user("Administrator")
        for key in ("rows", "tong", "qua_han_thanh_toan", "sap_den_han_so_luong"):
            self.assertIn(key, ket_qua)


class TestKhoPhieuListHinhDangMoi(FrappeTestCase):
    def test_luon_tra_dict(self):
        seed_kho_demo()
        frappe.set_user(BM_USER)
        try:
            ket_qua = kho_api.kho_phieu_list("nhap")
        finally:
            frappe.set_user("Administrator")
        self.assertIn("rows", ket_qua)
        self.assertIn("tong", ket_qua)


class TestPortalOrderHistoryLocTrangThaiPhiaServer(FrappeTestCase):
    """Hồi quy do đợt phân trang (brief 2026-08-16): chip lọc trạng thái ở
    Orders.vue lọc PHÍA CLIENT trên đúng MỘT trang đã tải — khách chọn "xem
    10" rồi bấm "Chờ xác nhận" thấy TRỐNG dù có đơn khớp ở trang 2. Test này
    dựng đúng kịch bản đó: 12 đơn KHÔNG khớp bộ lọc (creation MỚI HƠN, nên
    đứng đầu danh sách không lọc) rồi 3 đơn "Chờ xác nhận" (creation CŨ HƠN,
    rơi vào "trang 2" nếu không lọc)."""

    def test_loc_trang_thai_tim_thay_don_o_ngoai_trang_dau_khong_loc(self):
        frappe.set_user(BM_USER)
        try:
            muc_tieu = [_tao_don_test(BVBM, "Chờ khách đồng ý") for _ in range(3)]
            for _ in range(12):
                _tao_don_test(BVBM, "Khách huỷ")

            # Không lọc, trang đầu (limit=10) — đúng hành vi cũ: 12 đơn "Đã
            # huỷ" mới hơn lấp đầy, cả 3 đơn mục tiêu bị đẩy ra "trang 2".
            khong_loc = portal.portal_order_history(limit=10, start=0)
            ten_khong_loc = {r["name"] for r in khong_loc["rows"]}
            for so in muc_tieu:
                self.assertNotIn(
                    so.name, ten_khong_loc,
                    "phát hiện sai kịch bản test: đơn mục tiêu phải nằm ngoài trang đầu KHÔNG lọc",
                )

            # Có lọc — server phải tìm ra đủ 3 đơn NGAY Ở TRANG ĐẦU, không
            # phụ thuộc chỗ chúng đứng trong danh sách không lọc.
            co_loc = portal.portal_order_history(limit=10, start=0, trang_thai="Chờ xác nhận")
            ten_co_loc = {r["name"] for r in co_loc["rows"]}
            for so in muc_tieu:
                self.assertIn(so.name, ten_co_loc, "đơn khớp bộ lọc phải tìm được dù ở 'trang 2' nếu không lọc")
            for r in co_loc["rows"]:
                self.assertEqual(r["status_vi"], "Chờ xác nhận", "trang lọc không được lẫn đơn khác trạng thái")
        finally:
            frappe.set_user("Administrator")

    def test_tong_dem_dung_theo_bo_loc_dang_ap(self):
        frappe.set_user(BM_USER)
        try:
            for _ in range(3):
                _tao_don_test(BVBM, "Chờ khách đồng ý")

            # Ground truth độc lập: lấy TOÀN BỘ đơn (limit lớn, không lọc),
            # tự đếm `status_vi == "Chờ xác nhận"` bằng đúng field production
            # đã tính — không dựa vào con số cố định vì site có dữ liệu demo
            # sẵn khác thuộc cùng khách hàng.
            tat_ca = portal.portal_order_history(limit=1000, start=0)["rows"]
            dem_that = sum(1 for r in tat_ca if r["status_vi"] == "Chờ xác nhận")
            self.assertGreaterEqual(dem_that, 3)

            trang1 = portal.portal_order_history(limit=1, start=0, trang_thai="Chờ xác nhận")
            self.assertEqual(
                trang1["tong"], dem_that,
                "tong phải đếm ĐÚNG bộ lọc đang áp, không phải tổng toàn bộ đơn",
            )
            self.assertEqual(len(trang1["rows"]), 1, "limit=1 phải cắt đúng 1 dòng dù tong lớn hơn")
        finally:
            frappe.set_user("Administrator")

    def test_khong_truyen_trang_thai_giu_nguyen_hanh_vi_cu(self):
        """Mọi caller cũ (test_e2e_flow.py, test_tracking.py, test_e6_mua_
        le.py, Dashboard.vue) gọi KHÔNG truyền `trang_thai` — phải KHÔNG lọc
        gì, y hệt trước bản vá này."""
        frappe.set_user(BM_USER)
        try:
            khong_truyen = portal.portal_order_history(limit=1000, start=0)
            rong = portal.portal_order_history(limit=1000, start=0, trang_thai=None)
            self.assertEqual(khong_truyen["tong"], rong["tong"])
            self.assertEqual({r["name"] for r in khong_truyen["rows"]}, {r["name"] for r in rong["rows"]})
        finally:
            frappe.set_user("Administrator")


DEBIT_TO = "Debtors - MYN"
INCOME_ACCOUNT = "Sales - MYN"
COST_CENTER = "Main - MYN"


def _tao_hoa_don_test(customer, no=True):
    """Sales Invoice SUBMIT, `update_stock=0` (không cần Delivery Note) —
    cùng khuôn `_tao_si` (test_e7_hddt.py). `no=True` để lại `outstanding_
    amount > 0` (chưa thanh toán); `no=False` thanh toán đủ ngay để loại
    khỏi KPI "chưa thanh toán"."""
    si = frappe.new_doc("Sales Invoice")
    si.company = COMPANY
    si.customer = customer
    si.posting_date = frappe.utils.today()
    si.set_posting_time = 1
    si.debit_to = DEBIT_TO
    si.update_stock = 0
    si.append("items", {
        "item_code": "VT0005", "qty": 1, "rate": 1000,
        "income_account": INCOME_ACCOUNT, "cost_center": COST_CENTER,
    })
    si.insert(ignore_permissions=True)
    si.submit()
    if not no:
        # `Payment Entry.set_missing_values` gọi thẳng `frappe.has_permission
        # (throw=True)` bất kể `ignore_permissions` của `insert()` — BẪY
        # riêng của doctype này (khác Sales Order/Sales Invoice). Tạo tạm
        # bằng Administrator rồi trả lại phiên BM_USER cho phần còn lại của
        # test.
        nguoi_dang_goi = frappe.session.user
        frappe.set_user("Administrator")
        try:
            pe = frappe.get_doc({
                "doctype": "Payment Entry", "payment_type": "Receive", "company": COMPANY,
                "party_type": "Customer", "party": customer, "paid_amount": si.grand_total,
                "received_amount": si.grand_total, "paid_from": DEBIT_TO,
                "paid_to": "Cash - MYN", "references": [{
                    "reference_doctype": "Sales Invoice", "reference_name": si.name,
                    "allocated_amount": si.grand_total,
                }],
            })
            pe.insert(ignore_permissions=True)
            pe.submit()
        finally:
            frappe.set_user(nguoi_dang_goi)
    si.reload()
    return si


class TestPortalDashboardKpiKhongPhuThuocPhanTrang(FrappeTestCase):
    """Brief 2026-08-16 — Dashboard.vue trước đây suy 3 ô KPI (đơn chờ xác
    nhận/đang giao, hoá đơn chưa thanh toán) từ DANH SÁCH ĐÃ PHÂN TRANG
    (`portal_order_history`/`portal_invoices`, limit mặc định 20/10 tuỳ nơi
    gọi). Với `limit=10`, khách có 15 đơn "Chờ xác nhận" thật chỉ thấy ô
    hiện 3-10 — hỏng NGAY màn đầu tiên khách nhìn thấy mỗi lần đăng nhập."""

    def test_kpi_dem_toan_bo_khong_phu_thuoc_limit_cua_danh_sach(self):
        frappe.set_user(BM_USER)
        try:
            SO_LUONG = 15  # > mọi limit Dashboard.vue từng dùng (10/20)
            for _ in range(SO_LUONG):
                _tao_don_test(BVBM, "Chờ khách đồng ý")

            # Ground truth độc lập — cùng cách lỗi 1 đã canh: đếm tay trên
            # TOÀN BỘ danh sách (limit lớn), không dựa số cố định vì site có
            # dữ liệu demo sẵn khác thuộc cùng khách hàng.
            tat_ca = portal.portal_order_history(limit=1000, start=0)["rows"]
            dem_that = sum(1 for r in tat_ca if r["status_vi"] == "Chờ xác nhận")
            self.assertGreaterEqual(dem_that, SO_LUONG)

            # Trang ĐẦU limit=10 (đúng cỡ Dashboard.vue cũ dùng) không thể
            # đủ chỗ cho toàn bộ — canh giữ kịch bản bug (KPI giả thấp) còn
            # tái hiện được nếu ai đó lại suy KPI từ danh sách phân trang.
            trang_dau = portal.portal_order_history(limit=10, start=0)["rows"]
            dem_trang_dau = sum(1 for r in trang_dau if r["status_vi"] == "Chờ xác nhận")
            self.assertLess(
                dem_trang_dau, dem_that,
                "phát hiện sai kịch bản test: trang đầu limit=10 phải KHÔNG đủ để canh KPI",
            )

            kpi = portal.portal_dashboard_kpi()
            self.assertEqual(
                kpi["don_cho_xac_nhan"], dem_that,
                "KPI phải đếm TOÀN BỘ đơn của khách, không phải chỉ trang đầu",
            )
        finally:
            frappe.set_user("Administrator")

    def test_kpi_hoa_don_chua_thanh_toan_dem_toan_bo(self):
        frappe.set_user(BM_USER)
        try:
            SO_LUONG = 12  # > limit mặc định `portal_invoices` (20 vẫn ổn,
            # nhưng Dashboard.vue cũ có thể gọi limit nhỏ hơn — canh > 10)
            for _ in range(SO_LUONG):
                _tao_hoa_don_test(BVBM, no=True)
            _tao_hoa_don_test(BVBM, no=False)  # đã thanh toán — KHÔNG được đếm

            tat_ca = frappe.get_list(
                "Sales Invoice", filters={"customer": BVBM, "outstanding_amount": [">", 0]},
                fields=["count(name) as n"],
            )[0].n
            self.assertGreaterEqual(tat_ca, SO_LUONG)

            kpi = portal.portal_dashboard_kpi()
            self.assertEqual(
                kpi["hoa_don_chua_thanh_toan"], tat_ca,
                "KPI hoá đơn chưa TT phải đếm TOÀN BỘ, không chỉ trang đầu portal_invoices",
            )
        finally:
            frappe.set_user("Administrator")
