"""Import/export danh mục vật tư qua cổng."""

import io
from unittest.mock import patch

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import Workbook, load_workbook

from miyano_portal.kho import vat_tu as vat_tu_mod
from miyano_portal.setup.seed_kho_demo import seed_kho_demo

HEADERS = [label for label, _ in vat_tu_mod.DANH_MUC_COLUMNS]


def _xlsx(rows, headers=None):
	wb = Workbook()
	ws = wb.active
	ws.append(headers if headers is not None else HEADERS)
	for r in rows:
		ws.append(r)
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


class TestDanhMucFile(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		self.kho_bm = self.kho["kho_bm"]
		self._da_tao = []

	def tearDown(self):
		for name in self._da_tao:
			if frappe.db.exists("Customer Warehouse Item", name):
				frappe.delete_doc("Customer Warehouse Item", name, force=True, ignore_permissions=True)

	def test_preview_khong_ghi_gi(self):
		truoc = frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm})
		vat_tu_mod.parse_danh_muc(
			_xlsx([["BM-NEW-01", "Vật tư mới", "Cái", "", "", "", 1]]), self.kho_bm
		)
		self.assertEqual(frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm}), truoc)

	def test_ma_moi_thi_tao_ma_da_co_thi_cap_nhat(self):
		content = _xlsx([
			["BM-NEW-02", "Vật tư mới 2", "Cái", "", "Gói 10", "Tiêu hao", 1],
			["MYN-GLOVE-M", "Găng tay ĐỔI TÊN", "Hộp", "", "", "", 1],
		])
		parsed = vat_tu_mod.parse_danh_muc(content, self.kho_bm)
		self.assertEqual(parsed["error_count"], 0)
		self.assertEqual(parsed["summary"], {"tao_moi": 1, "cap_nhat": 1})

		kq = vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		moi = frappe.db.get_value("Customer Warehouse Item", {"kho": self.kho_bm, "ma_vat_tu": "BM-NEW-02"})
		self._da_tao.append(moi)
		self.assertEqual(kq, {"tao_moi": 1, "cap_nhat": 1})
		self.assertEqual(
			frappe.db.get_value("Customer Warehouse Item", self.kho["vt_bm"], "ten_vat_tu"),
			"Găng tay ĐỔI TÊN",
		)

	def test_mot_dong_loi_thi_khong_ghi_gi(self):
		truoc = frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm})
		content = _xlsx([
			["BM-NEW-03", "Hợp lệ", "Cái", "", "", "", 1],
			["", "Thiếu mã", "Cái", "", "", "", 1],
		])
		with self.assertRaises(frappe.ValidationError):
			vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		self.assertEqual(frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm}), truoc)

	def test_doi_dvt_vat_tu_da_phat_sinh_la_dong_loi(self):
		self._bao_dam_co_phat_sinh(self.kho["vt_bm"])
		parsed = vat_tu_mod.parse_danh_muc(
			_xlsx([["MYN-GLOVE-M", "Găng tay y tế size M", "Cái", "", "", "", 1]]), self.kho_bm
		)
		self.assertEqual(parsed["error_count"], 1)
		self.assertIn("ĐVT", " ".join(parsed["rows_error"][0]["errors"]))

	def test_tat_vat_tu_con_ton_la_dong_loi(self):
		self._bao_dam_co_phat_sinh(self.kho["vt_bm"])
		parsed = vat_tu_mod.parse_danh_muc(
			_xlsx([["MYN-GLOVE-M", "Găng tay y tế size M", "Hộp", "", "", "", 0]]), self.kho_bm
		)
		self.assertEqual(parsed["error_count"], 1)
		self.assertIn("còn tồn", " ".join(parsed["rows_error"][0]["errors"]))

	def test_round_trip_xuat_roi_nap_lai_khong_doi_du_lieu(self):
		content = vat_tu_mod.build_danh_muc_xlsx(self.kho_bm)
		ws = load_workbook(io.BytesIO(content), data_only=True).active
		self.assertEqual([c.value for c in ws[1]], HEADERS)
		parsed = vat_tu_mod.parse_danh_muc(content, self.kho_bm)
		self.assertEqual(parsed["error_count"], 0)
		self.assertEqual(parsed["summary"]["tao_moi"], 0)

	def test_dvt_lech_hoa_thuong_khong_bi_chan_va_khong_doi(self):
		# ĐVT hiện tại của MYN-GLOVE-M là "Hộp" (seed_kho_demo). File ghi lại
		# "hộp" — chỉ khác hoa/thường, KHÔNG phải một phép đổi ĐVT thật — trên
		# một vật tư đã có phát sinh. Rào ĐVT ở parse_danh_muc so sánh fold nên
		# không được chặn dòng này; nếu commit_danh_muc lỡ gửi "dvt" xuống
		# sua() (vốn so sánh case-sensitive) thì sẽ ném lỗi giữa vòng ghi dù
		# preview báo sạch — đúng lỗi đã tìm thấy và sửa ở round trước.
		self._bao_dam_co_phat_sinh(self.kho["vt_bm"])
		content = _xlsx([["MYN-GLOVE-M", "Găng tay y tế size M", "hộp", "", "", "", 1]])
		parsed = vat_tu_mod.parse_danh_muc(content, self.kho_bm)
		self.assertEqual(parsed["error_count"], 0)

		kq = vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		self.assertEqual(kq, {"tao_moi": 0, "cap_nhat": 1})
		self.assertEqual(
			frappe.db.get_value("Customer Warehouse Item", self.kho["vt_bm"], "dvt"),
			"Hộp",
		)

	def test_tao_moi_dang_dung_0_thi_tao_o_trang_thai_tat(self):
		# tao() luôn tạo active=1 theo mặc định — dòng file xin 'Đang dùng=0'
		# cho một mã hoàn toàn mới không được lặng lẽ bỏ qua giá trị đó.
		content = _xlsx([["BM-NEW-TAT", "Vật tư mới nhưng tắt", "Cái", "", "", "", 0]])
		kq = vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		self.assertEqual(kq, {"tao_moi": 1, "cap_nhat": 0})
		moi = frappe.db.get_value(
			"Customer Warehouse Item", {"kho": self.kho_bm, "ma_vat_tu": "BM-NEW-TAT"}
		)
		self._da_tao.append(moi)
		self.assertEqual(frappe.db.get_value("Customer Warehouse Item", moi, "active"), 0)

	def test_dang_dung_chi_co_khoang_trang_la_dang_dung(self):
		# Một ô "Đang dùng" chỉ chứa khoảng trắng (dán từ Excel/Sheets rất hay
		# dính) phải được coi là TRỐNG, tức "đang dùng" — không được fold về
		# "" rồi rơi nhầm vào nhánh FALSE.
		content = _xlsx([["BM-NEW-WS", "Vật tư mới, ô Đang dùng dính khoảng trắng", "Cái", "", "", "", " "]])
		vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		moi = frappe.db.get_value(
			"Customer Warehouse Item", {"kho": self.kho_bm, "ma_vat_tu": "BM-NEW-WS"}
		)
		self._da_tao.append(moi)
		self.assertEqual(frappe.db.get_value("Customer Warehouse Item", moi, "active"), 1)

	def test_ma_trung_trong_cung_tep_la_dong_loi(self):
		# Hai dòng cùng một mã (kể cả chỉ lệch hoa/thường) trong CÙNG một tệp:
		# xử lý độc lập từng dòng sẽ khiến cả hai đều thành 'tao_moi' hợp lệ ở
		# preview, rồi lúc ghi dòng sau âm thầm bị tao() trả về đúng bản ghi
		# dòng trước vừa tạo (không lỗi, không ghi đè, nội dung dòng sau biến
		# mất). Coi đây là LỖI ngay ở bước parse, không đoán dòng nào thắng.
		truoc = frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm})
		content = _xlsx([
			["BM-DUP-01", "Vật tư trùng mã 1", "Cái", "", "", "", 1],
			["bm-dup-01", "Vật tư trùng mã 2", "Cái", "", "", "", 1],
		])
		parsed = vat_tu_mod.parse_danh_muc(content, self.kho_bm)
		self.assertEqual(parsed["error_count"], 2)
		self.assertIn("trùng", " ".join(parsed["rows_error"][0]["errors"]))
		self.assertIn("trùng", " ".join(parsed["rows_error"][1]["errors"]))

		with self.assertRaises(frappe.ValidationError):
			vat_tu_mod.commit_danh_muc(content, self.kho_bm)
		self.assertEqual(frappe.db.count("Customer Warehouse Item", {"kho": self.kho_bm}), truoc)

	def _bao_dam_co_phat_sinh(self, vat_tu):
		if vat_tu_mod.co_phat_sinh(vat_tu):
			return
		doc = frappe.get_doc({
			"doctype": "Customer Stock Receipt",
			"kho": self.kho_bm,
			"ngay": frappe.utils.today(),
			"loai_nhap": "Nhập khác",
			"items": [{"vat_tu": vat_tu, "so_lo": "LO-DM-01", "so_luong": 5, "don_gia": 1000}],
		})
		doc.insert(ignore_permissions=True)
		doc.submit()


from miyano_portal.api import kho as kho_api

BM_USER = "bvbm@demo.miyano"


class TestDanhMucEndpoint(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		frappe.set_user(BM_USER)

	def tearDown(self):
		frappe.set_user("Administrator")

	def test_export_dat_response_dung_dinh_dang(self):
		kho_api.kho_vat_tu_export()
		self.assertEqual(frappe.local.response.filename, "danh_muc_vat_tu.xlsx")
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent)

	def test_loi_ngoai_du_kien_khi_xuat_thanh_cau_tieng_viet_ve_vat_tu(self):
		"""Ba endpoint danh mục phải nằm dưới lớp bọc _action, VÀ lớp bọc phải
		gọi đúng tên thứ đang xử lý.

		Trước bản sửa này chúng không được bọc: một lỗi ngoài dự kiến đi thẳng
		ra khách dưới dạng traceback tiếng Anh. Bọc bằng _phieu_action cũ thì
		khách đang nhập DANH MỤC lại đọc được câu "khi xử lý phiếu" và đi tìm
		phiếu nào vừa hỏng.
		"""
		with patch.object(vat_tu_mod, "build_danh_muc_xlsx", side_effect=ValueError("boom")):
			with self.assertRaises(frappe.ValidationError) as ctx:
				kho_api.kho_vat_tu_export()
		self.assertIn("vật tư", str(ctx.exception))
		self.assertNotIn("phiếu", str(ctx.exception))

	def test_loi_ngoai_du_kien_khi_nap_danh_muc_cung_duoc_boc(self):
		with patch.object(vat_tu_mod, "commit_danh_muc", side_effect=ValueError("boom")):
			f = frappe.get_doc({
				"doctype": "File", "file_name": "danh_muc.xlsx",
				"content": _xlsx([["BM-NEW-09", "Vật tư mới", "Cái", "", "", "", 1]]),
				"is_private": 1,
			}).insert(ignore_permissions=True)
			with self.assertRaises(frappe.ValidationError) as ctx:
				kho_api.kho_vat_tu_import_commit(f.file_url)
		self.assertIn("vật tư", str(ctx.exception))

	def test_boc_khong_nuot_co_che_tra_tep(self):
		"""Lớp bọc chỉ chuyển tiếp giá trị trả về; frappe.local.response mà
		endpoint export đặt vào phải còn nguyên sau khi đi qua nó."""
		frappe.local.response.clear()
		kho_api.kho_vat_tu_export()
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent)

	def test_preview_file_cua_nguoi_khac_bi_chan(self):
		f = frappe.get_doc({
			"doctype": "File", "file_name": "cua_nguoi_khac.xlsx",
			"content": "x", "is_private": 1,
		}).insert(ignore_permissions=True)
		frappe.db.set_value("File", f.name, "owner", "Administrator")
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_vat_tu_import_preview(f.file_url)
