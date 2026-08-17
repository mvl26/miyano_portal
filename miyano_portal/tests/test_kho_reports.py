"""Phase 5 — báo cáo Nhập-Xuất-Tồn, thẻ kho, cảnh báo hạn dùng, xuất Excel.

Mọi ngày dùng trong file này được TÍNH TƯƠNG ĐỐI so với frappe.utils.today(),
không hardcode một ngày tuyệt đối: dự án này đã dính "date rot" hai lần
(xem progress.md, Task 4 & 5) vì test hardcode ngày trong khi validate_ngay/
so sánh "hôm nay" trôi theo thời gian chạy thật. Ở đây còn một lý do thứ hai
mạnh hơn: phiếu đảo do on_cancel() sinh ra LUÔN mang `ngay = frappe.utils.today()`
(xem customer_stock_receipt.py/_tao_phieu_dao và customer_stock_issue.py cùng
tên hàm) — một kỳ báo cáo hardcode trong quá khứ sẽ không bao giờ chứa được
phiếu đảo của phép huỷ diễn ra lúc test chạy.
"""

import io
import re
from pathlib import Path

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import load_workbook

from miyano_portal.api import kho as kho_api
from miyano_portal.kho import ledger, reports
from miyano_portal.setup.seed_kho_demo import seed_kho_demo

BM_USER = "bvbm@demo.miyano"
PXN_USER = "pxnabc@demo.miyano"


def _today():
	return frappe.utils.getdate(frappe.utils.today())


def _iso(d):
	return frappe.utils.getdate(d).strftime("%Y-%m-%d")


def _norm_cell(v):
	"""Hai đặc điểm của openpyxl khiến so sánh cell-by-cell "ngây thơ" báo lệch
	dù dữ liệu giống hệt nhau — chuẩn hoá TRƯỚC khi so sánh:

	1. Đọc lại một ô đã ghi `datetime.date` ra `datetime.datetime` (Excel
	   không phân biệt hai kiểu này ở tầng lưu trữ).
	2. Đọc lại một ô đã ghi CHUỖI RỖNG `''` ra `None` (openpyxl không ghi gì
	   cho chuỗi rỗng, nên đọc lại là một ô trống thật sự) — nhiều field
	   "không áp dụng" của NHAT_KY_COLUMNS/DOT_COLUMNS (nguồn/đợt của dòng
	   xuất, chứng từ NCC khi nguồn là Miyano...) cố ý là `''`, không phải
	   `None`. Chỉ chuẩn hoá `None -> ''`, KHÔNG đụng các giá trị falsy khác
	   (0, 0.0, False) — những giá trị đó có ý nghĩa số/logic thật, không
	   phải "trống"."""
	import datetime
	if isinstance(v, datetime.datetime):
		return v.date()
	if v is None:
		return ""
	return v


def _nhap(kho, vat_tu, so_luong, don_gia, ngay, so_lo="LO-A", han="2030-01-01"):
	doc = frappe.get_doc({
		"doctype": "Customer Stock Receipt",
		"kho": kho, "ngay": _iso(ngay), "loai_nhap": "Nhập khác",
		"nguoi_giao": "Trần Văn Giao",
		"items": [{
			"vat_tu": vat_tu, "so_lo": so_lo, "han_su_dung": han,
			"so_luong": so_luong, "don_gia": don_gia,
		}],
	})
	doc.insert(ignore_permissions=True)
	doc.submit()
	return doc


def _xuat(kho, vat_tu, so_luong, ngay, so_lo="LO-A"):
	doc = frappe.get_doc({
		"doctype": "Customer Stock Issue",
		"kho": kho, "ngay": _iso(ngay), "loai_xuat": "Xuất sử dụng",
		"noi_nhan": "Khoa Nội", "nguoi_nhan": "Y tá Lan",
		"items": [{"vat_tu": vat_tu, "so_lo": so_lo, "so_luong": so_luong}],
	})
	doc.insert(ignore_permissions=True)
	doc.submit()
	return doc


class _KhoBmTestCase(FrappeTestCase):
	"""Base: seed + dọn sạch sổ/tồn của kho BM trước mỗi test.

	FrappeTestCase chỉ rollback ở cuối CLASS, không phải cuối từng test method
	(bài học đã ghi trong nhiều file test khác của module này) — vì các báo
	cáo ở đây CỘNG DỒN trên toàn bộ sổ của kho, một dòng rác từ test trước sẽ
	âm thầm cộng vào tổng của test sau. Dọn ở setUp, đúng khuôn
	TestKhoBatBienGiaTri/TestKhoLedger.
	"""

	def setUp(self):
		self.kho = seed_kho_demo()
		self.K = self.kho["kho_bm"]
		self.VT = self.kho["vt_bm"]
		frappe.db.delete("Customer Stock Ledger Entry", {"kho": self.K})
		frappe.db.delete("Customer Stock Lot Balance", {"kho": self.K})


class TestKhoBaoCaoNXT(_KhoBmTestCase):
	def test_eight_columns_balance_with_receipt_issue_and_cancellation(self):
		"""Kỳ chứa cả ba sự kiện: một khoản tồn đầu, một phiếu nhập bị huỷ
		trong chính kỳ đó, và một phiếu xuất. Phiếu đảo do huỷ sinh ra mang
		ngay=today() (xem docstring đầu file), nên kỳ PHẢI trải rộng qua
		"hôm nay" để chứa được nó.
		"""
		today = _today()
		tu_ngay = frappe.utils.add_days(today, -20)
		den_ngay = frappe.utils.add_days(today, 5)

		# Tồn đầu kỳ: 50 @ 40.000, trước tu_ngay.
		_nhap(self.K, self.VT, 50, 40000, frappe.utils.add_days(today, -30))
		# Trong kỳ: nhập 100 @ 50.000 (sẽ bị huỷ), rồi xuất 30.
		pn = _nhap(self.K, self.VT, 100, 50000, frappe.utils.add_days(today, -15))
		_xuat(self.K, self.VT, 30, frappe.utils.add_days(today, -10))
		# Huỷ phiếu nhập NGAY BÂY GIỜ -> phiếu đảo -100 @ 50.000, ngay=today().
		pn.reload()
		pn.cancel()

		rows = reports.nxt_item_rows(self.K, tu_ngay, den_ngay)
		row = next(r for r in rows if r["vat_tu"] == self.VT)

		self.assertEqual(row["ton_dau_sl"], 50)
		self.assertEqual(row["ton_dau_tt"], 2_000_000)
		self.assertEqual(row["nhap_sl"], 100)
		self.assertEqual(row["nhap_tt"], 5_000_000)
		# xuất = 30 (phiếu xuất thường) + 100 (phiếu đảo của phiếu nhập, dấu âm
		# trong sổ nên rơi vào nhánh xuất) = 130.
		self.assertEqual(row["xuat_sl"], 130)
		self.assertEqual(row["xuat_tt"], 6_400_000)
		# Hằng đẳng thức bắt buộc, cả số lượng lẫn giá trị.
		self.assertEqual(row["ton_dau_sl"] + row["nhap_sl"] - row["xuat_sl"], row["ton_cuoi_sl"])
		self.assertEqual(row["ton_dau_tt"] + row["nhap_tt"] - row["xuat_tt"], row["ton_cuoi_tt"])
		self.assertEqual(row["ton_cuoi_sl"], 20)
		self.assertEqual(row["ton_cuoi_tt"], 600_000)

		# Khớp với tồn thật của lô sau toàn bộ chuỗi sự kiện (bằng chứng độc
		# lập: báo cáo không chỉ "tự khớp nó", mà khớp cache tồn theo lô).
		bal = ledger.get_lot_balance(self.K, self.VT, "LO-A")
		self.assertAlmostEqual(float(bal["so_luong"]), row["ton_cuoi_sl"], delta=0.01)
		self.assertAlmostEqual(
			float(bal["so_luong"]) * float(bal["don_gia"]), row["ton_cuoi_tt"], delta=0.01
		)

	def test_opening_balance_correct_when_period_starts_mid_history(self):
		"""Hai lần nhập TRƯỚC tu_ngay, giá khác nhau, KHÔNG có phát sinh trong
		kỳ. Tồn đầu phải là tổng thật của sổ trước tu_ngay — không phải 0,
		không phải chỉ lần nhập gần nhất."""
		today = _today()
		_nhap(self.K, self.VT, 40, 30000, frappe.utils.add_days(today, -50))
		_nhap(self.K, self.VT, 60, 60000, frappe.utils.add_days(today, -40))

		tu_ngay = frappe.utils.add_days(today, -30)
		den_ngay = frappe.utils.add_days(today, 5)
		rows = reports.nxt_item_rows(self.K, tu_ngay, den_ngay)
		row = next(r for r in rows if r["vat_tu"] == self.VT)

		self.assertEqual(row["ton_dau_sl"], 100)
		self.assertEqual(row["ton_dau_tt"], 40 * 30000 + 60 * 60000)
		self.assertEqual(row["nhap_sl"], 0)
		self.assertEqual(row["xuat_sl"], 0)
		# Không phát sinh trong kỳ -> tồn cuối = tồn đầu, không phải bịa ra 0.
		self.assertEqual(row["ton_cuoi_sl"], row["ton_dau_sl"])
		self.assertEqual(row["ton_cuoi_tt"], row["ton_dau_tt"])

	def test_item_with_movement_but_zero_closing_still_appears(self):
		"""Nhập rồi xuất hết sạch trong cùng kỳ: tồn cuối = 0 nhưng dòng vật
		tư đó KHÔNG được biến mất khỏi báo cáo."""
		today = _today()
		tu_ngay = frappe.utils.add_days(today, -10)
		den_ngay = frappe.utils.add_days(today, 5)
		_nhap(self.K, self.VT, 25, 20000, frappe.utils.add_days(today, -5))
		_xuat(self.K, self.VT, 25, frappe.utils.add_days(today, -3))

		rows = reports.nxt_item_rows(self.K, tu_ngay, den_ngay)
		row = next((r for r in rows if r["vat_tu"] == self.VT), None)
		self.assertIsNotNone(row, "vật tư có phát sinh trong kỳ nhưng tồn cuối 0 đã bị lọc mất")
		self.assertEqual(row["nhap_sl"], 25)
		self.assertEqual(row["xuat_sl"], 25)
		self.assertEqual(row["ton_cuoi_sl"], 0)
		self.assertEqual(row["ton_cuoi_tt"], 0)

	def test_reversal_lands_in_the_period_it_was_posted_not_the_original(self):
		"""Phiếu nhập ở kỳ A, huỷ ở kỳ B (sau đó): kỳ A vẫn giữ NGUYÊN số nhập
		gộp (không bị hồi tố trừ đi bởi cái huỷ diễn ra sau); kỳ B nhận đúng
		dòng đảo và tự cân bằng."""
		today = _today()
		ky_a_tu = frappe.utils.add_days(today, -60)
		ky_a_den = frappe.utils.add_days(today, -31)
		ky_b_tu = frappe.utils.add_days(today, -30)
		ky_b_den = frappe.utils.add_days(today, 5)

		pn = _nhap(self.K, self.VT, 100, 50000, frappe.utils.add_days(today, -45))
		pn.reload()
		pn.cancel()  # đảo hôm nay -> rơi vào kỳ B, không phải kỳ A

		rows_a = reports.nxt_item_rows(self.K, ky_a_tu, ky_a_den)
		row_a = next(r for r in rows_a if r["vat_tu"] == self.VT)
		self.assertEqual(row_a["ton_dau_sl"], 0)
		self.assertEqual(row_a["nhap_sl"], 100, "kỳ gốc phải giữ nguyên số nhập gộp")
		self.assertEqual(row_a["xuat_sl"], 0)
		self.assertEqual(row_a["ton_cuoi_sl"], 100)
		self.assertEqual(row_a["ton_cuoi_tt"], 5_000_000)

		rows_b = reports.nxt_item_rows(self.K, ky_b_tu, ky_b_den)
		row_b = next(r for r in rows_b if r["vat_tu"] == self.VT)
		self.assertEqual(row_b["ton_dau_sl"], 100, "tồn đầu kỳ B kế thừa đúng tồn cuối kỳ A")
		self.assertEqual(row_b["ton_dau_tt"], 5_000_000)
		self.assertEqual(row_b["xuat_sl"], 100, "dòng đảo phải xuất hiện Ở KỲ B")
		self.assertEqual(row_b["xuat_tt"], 5_000_000)
		self.assertEqual(row_b["ton_cuoi_sl"], 0)
		self.assertEqual(row_b["ton_cuoi_tt"], 0)

	def test_lot_drilldown_sums_to_item_row_on_all_eight_figures(self):
		today = _today()
		tu_ngay = frappe.utils.add_days(today, -10)
		den_ngay = frappe.utils.add_days(today, 5)
		_nhap(self.K, self.VT, 40, 10000, frappe.utils.add_days(today, -8), so_lo="LO-A")
		_nhap(self.K, self.VT, 60, 20000, frappe.utils.add_days(today, -6), so_lo="LO-B")
		_xuat(self.K, self.VT, 10, frappe.utils.add_days(today, -4), so_lo="LO-A")

		item_row = next(
			r for r in reports.nxt_item_rows(self.K, tu_ngay, den_ngay) if r["vat_tu"] == self.VT
		)
		lot_rows = reports.nxt_lot_rows(self.K, self.VT, tu_ngay, den_ngay)
		self.assertEqual({r["so_lo"] for r in lot_rows}, {"LO-A", "LO-B"})

		for key in (
			"ton_dau_sl", "ton_dau_tt", "nhap_sl", "nhap_tt",
			"xuat_sl", "xuat_tt", "ton_cuoi_sl", "ton_cuoi_tt",
		):
			tong_lo = sum(r[key] for r in lot_rows)
			self.assertAlmostEqual(
				tong_lo, item_row[key], delta=0.01,
				msg=f"tổng lô của {key} ({tong_lo}) lệch dòng vật tư ({item_row[key]})",
			)

	def test_from_date_after_to_date_is_rejected(self):
		today = _today()
		with self.assertRaises(frappe.ValidationError):
			reports.nxt_item_rows(self.K, today, frappe.utils.add_days(today, -1))


class TestKhoTheKho(_KhoBmTestCase):
	def test_running_balance_matches_ledger_and_nxt_closing_at_every_row(self):
		today = _today()
		tu_ngay = frappe.utils.add_days(today, -15)
		den_ngay = frappe.utils.add_days(today, 5)

		_nhap(self.K, self.VT, 50, 10000, frappe.utils.add_days(today, -20))  # tồn đầu
		_nhap(self.K, self.VT, 30, 12000, frappe.utils.add_days(today, -10), so_lo="LO-A")
		_xuat(self.K, self.VT, 20, frappe.utils.add_days(today, -8), so_lo="LO-A")
		_nhap(self.K, self.VT, 15, 11000, frappe.utils.add_days(today, -2), so_lo="LO-B")

		card = reports.the_kho_rows(self.K, self.VT, tu_ngay, den_ngay)
		self.assertEqual(len(card), 3)

		expected_balance = [80, 60, 75]  # 50+30=80, 80-20=60, 60+15=75
		for row, expect in zip(card, expected_balance):
			self.assertEqual(row["ton_luy_ke"], expect)

		# Đối chiếu chéo với báo cáo N-X-T cho CÙNG khoảng ngày: dòng cuối
		# cùng của thẻ kho phải bằng đúng tồn cuối của báo cáo tổng hợp.
		nxt_row = next(
			r for r in reports.nxt_item_rows(self.K, tu_ngay, den_ngay) if r["vat_tu"] == self.VT
		)
		self.assertEqual(card[-1]["ton_luy_ke"], nxt_row["ton_cuoi_sl"])

	def test_rejects_vat_tu_belonging_to_another_customer(self):
		frappe.set_user(BM_USER)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_the_kho(
				vat_tu=self.kho["vt_pxn"],
				tu_ngay="2026-01-01", den_ngay="2026-12-31",
			)

	def test_own_item_reachable_positive_control(self):
		today = _today()
		_nhap(self.K, self.VT, 10, 10000, frappe.utils.add_days(today, -1))
		frappe.set_user(BM_USER)
		rows = kho_api.kho_the_kho(
			vat_tu=self.VT,
			tu_ngay=_iso(frappe.utils.add_days(today, -5)),
			den_ngay=_iso(frappe.utils.add_days(today, 5)),
		)
		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["ton_luy_ke"], 10)


class TestKhoCanhBaoHan(_KhoBmTestCase):
	def test_separates_expired_from_expiring_soon_and_orders_nearest_first(self):
		today = _today()
		_nhap(self.K, self.VT, 5, 1000, frappe.utils.add_days(today, -30),
			  so_lo="LO-HET-HAN", han=frappe.utils.add_days(today, -5))
		_nhap(self.K, self.VT, 7, 2000, frappe.utils.add_days(today, -20),
			  so_lo="LO-SAP-HET", han=frappe.utils.add_days(today, 30))
		# Ngoài cửa sổ mặc định 90 ngày -> PHẢI vắng mặt (positive control cho
		# việc bộ lọc thực sự loại trừ, không phải "trả về tất cả rồi sắp xếp").
		_nhap(self.K, self.VT, 9, 3000, frappe.utils.add_days(today, -10),
			  so_lo="LO-XA-HAN", han=frappe.utils.add_days(today, 100))

		rows = reports.canh_bao_han_rows(self.K, so_ngay=90)
		lo_thay = [r["so_lo"] for r in rows]
		self.assertIn("LO-HET-HAN", lo_thay)
		self.assertIn("LO-SAP-HET", lo_thay)
		self.assertNotIn("LO-XA-HAN", lo_thay, "lô hết hạn ngoài cửa sổ so_ngay phải bị loại")

		het_han = next(r for r in rows if r["so_lo"] == "LO-HET-HAN")
		sap_het = next(r for r in rows if r["so_lo"] == "LO-SAP-HET")
		self.assertEqual(het_han["trang_thai"], "Đã hết hạn")
		self.assertEqual(sap_het["trang_thai"], "Sắp hết hạn")
		self.assertLess(het_han["so_ngay_con_lai"], 0)
		self.assertGreater(sap_het["so_ngay_con_lai"], 0)

		# Nearest-first: đã hết hạn (ngày quá khứ) phải đứng TRƯỚC sắp hết hạn
		# (ngày tương lai) trong danh sách sắp theo han_su_dung tăng dần.
		idx_het = lo_thay.index("LO-HET-HAN")
		idx_sap = lo_thay.index("LO-SAP-HET")
		self.assertLess(idx_het, idx_sap)

	def test_lot_with_zero_stock_excluded(self):
		today = _today()
		_nhap(self.K, self.VT, 5, 1000, frappe.utils.add_days(today, -10),
			  so_lo="LO-BAN-HET", han=frappe.utils.add_days(today, 10))
		_xuat(self.K, self.VT, 5, frappe.utils.add_days(today, -1), so_lo="LO-BAN-HET")
		rows = reports.canh_bao_han_rows(self.K, so_ngay=90)
		self.assertFalse(any(r["so_lo"] == "LO-BAN-HET" for r in rows))


class TestKhoCanhBaoHanApiSoNgayCoercion(_KhoBmTestCase):
	"""`kho_api.kho_canh_bao_han` chưa được test qua ĐÚNG cổng portal ở đâu
	khác trong dự án (lớp trên chỉ gọi thẳng `reports.canh_bao_han_rows()`) —
	đây là chỗ phủ việc ép kiểu tham số `so_ngay` khi tới từ HTTP thật dưới
	dạng chuỗi (review Item 4)."""

	def test_so_ngay_accepts_string_param_like_real_http_request(self):
		today = _today()
		_nhap(self.K, self.VT, 5, 1000, frappe.utils.add_days(today, -10),
			  so_lo="LO-SAP-HET", han=frappe.utils.add_days(today, 20))
		frappe.set_user(BM_USER)
		try:
			rows = kho_api.kho_canh_bao_han(so_ngay="30")
		finally:
			frappe.set_user("Administrator")
		self.assertTrue(
			any(r["so_lo"] == "LO-SAP-HET" for r in rows),
			"so_ngay=\"30\" (chuỗi) phải được hiểu đúng như số nguyên 30",
		)

	def test_so_ngay_non_numeric_rejected_with_specific_vietnamese_message(self):
		frappe.set_user(BM_USER)
		try:
			with self.assertRaises(frappe.ValidationError) as ctx:
				kho_api.kho_canh_bao_han(so_ngay="khong-phai-so")
		finally:
			frappe.set_user("Administrator")
		msg = str(ctx.exception)
		self.assertIn("Số ngày không hợp lệ", msg)
		self.assertNotIn("Traceback", msg)

	def test_so_ngay_blank_string_falls_back_to_default(self):
		"""Chuỗi rỗng (ô input HTML trống gửi lên) phải rơi về mặc định 90,
		không bị coi là "không hợp lệ"."""
		frappe.set_user(BM_USER)
		try:
			rows = kho_api.kho_canh_bao_han(so_ngay="")
		finally:
			frappe.set_user("Administrator")
		self.assertIsInstance(rows, list)


class TestKhoBaoCaoExcel(_KhoBmTestCase):
	_LABEL_FILE = (
		Path(__file__).resolve().parents[2]
		/ "frontend" / "src" / "kho-bao-cao-columns.js"
	)

	def _js_array(self, export_name: str) -> list[str]:
		"""Đọc mảng `export const <export_name> = [ {label: '...', field: '...'}, ... ]`
		từ kho-bao-cao-columns.js và trả về DANH SÁCH NHÃN theo đúng thứ tự
		xuất hiện — cùng file mà BaoCaoNXT.vue import để vẽ tiêu đề bảng, nên
		đây là phép so sánh với chính artifact hiển thị, không phải một bản
		chép tay có thể lệch khỏi màn hình thật."""
		text = self._LABEL_FILE.read_text(encoding="utf-8")
		m = re.search(r"export const " + export_name + r"\s*=\s*\[(.*?)\n\]", text, re.S)
		self.assertIsNotNone(m, f"không tìm thấy {export_name} trong kho-bao-cao-columns.js")
		labels = re.findall(r"label:\s*'((?:[^'\\]|\\.)*)'", m.group(1))
		self.assertTrue(labels, f"{export_name}: không đọc được nhãn cột nào")
		return [i.encode().decode("unicode_escape") if "\\" in i else i for i in labels]

	def test_nxt_excel_columns_match_screen_labels_and_file_rereads(self):
		today = _today()
		tu_ngay = frappe.utils.add_days(today, -10)
		den_ngay = frappe.utils.add_days(today, 5)
		_nhap(self.K, self.VT, 20, 15000, frappe.utils.add_days(today, -2))

		screen_rows = reports.nxt_item_rows(self.K, tu_ngay, den_ngay)
		self.assertTrue(screen_rows)

		frappe.set_user(BM_USER)
		try:
			kho_api.kho_bao_cao_excel(
				loai="nxt", tu_ngay=_iso(tu_ngay), den_ngay=_iso(den_ngay),
			)
			content = frappe.local.response.filecontent
			self.assertEqual(
				frappe.local.response.content_type,
				"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			)
		finally:
			frappe.local.response.clear()
			frappe.set_user("Administrator")

		wb = load_workbook(io.BytesIO(content))
		ws = wb.active
		header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

		# Cột trên file PHẢI khớp cả nhãn Python (COLUMNS) lẫn nhãn hiển thị
		# trên màn hình (kho-bao-cao-columns.js) — hai nguồn độc lập cùng
		# phải đồng ý, đúng nguyên tắc round-tripping-spreadsheets.
		self.assertEqual(header, [label for label, _ in reports.NXT_COLUMNS])
		self.assertEqual(header, self._js_array("NXT_COLUMNS"))

		data_rows = list(ws.iter_rows(min_row=2, values_only=True))
		self.assertEqual(len(data_rows), len(screen_rows))
		field_order = [field for _, field in reports.NXT_COLUMNS]
		for excel_row, screen_row in zip(data_rows, screen_rows):
			for col_idx, field in enumerate(field_order):
				self.assertEqual(
					excel_row[col_idx], screen_row[field],
					msg=f"cột {field} lệch giữa Excel và màn hình",
				)

	def test_the_kho_excel_columns_match_js_labels(self):
		self.assertEqual(
			[label for label, _ in reports.THE_KHO_COLUMNS],
			self._js_array("THE_KHO_COLUMNS"),
		)

	def test_canh_bao_excel_columns_match_js_labels(self):
		self.assertEqual(
			[label for label, _ in reports.CANH_BAO_COLUMNS],
			self._js_array("CANH_BAO_COLUMNS"),
		)

	def test_nxt_lot_excel_columns_match_js_labels(self):
		self.assertEqual(
			[label for label, _ in reports.NXT_LOT_COLUMNS],
			self._js_array("NXT_LOT_COLUMNS"),
		)

	def test_nhat_ky_excel_columns_match_js_labels(self):
		"""Gap 2 (review E4 phần B): trước bản này kho_bao_cao_excel không
		nhận loai="nhat_ky" — hai nút Excel của bản mẫu (NhatKy.vue/
		BaoCaoNXT.vue) đã bị khoá cứng chờ backend."""
		self.assertEqual(
			[label for label, _ in reports.NHAT_KY_COLUMNS],
			self._js_array("NHAT_KY_COLUMNS"),
		)

	def test_dot_excel_columns_match_js_labels(self):
		self.assertEqual(
			[label for label, _ in reports.DOT_COLUMNS],
			self._js_array("DOT_COLUMNS"),
		)

	def test_cap_phat_thang_excel_columns_match_js_labels(self):
		"""Yêu cầu chủ đầu tư 2026-08-17. Tab "Cấp phát theo tháng" CÓ nút
		Excel (khác tab "Cấp phát theo khoa", cố tình không có) — nên bộ cột
		của nó phải nằm dưới cùng phép canh như năm bộ kia."""
		self.assertEqual(
			[label for label, _ in reports.CAP_PHAT_THANG_COLUMNS],
			self._js_array("CAP_PHAT_THANG_COLUMNS"),
		)

	def test_nhat_ky_excel_exports_beyond_the_fifty_row_screen_page(self):
		"""Excel KHÔNG được cắt theo trang 50 dòng như màn hình — NL-8.3 chỉ
		bắt buộc chọn kỳ, không giới hạn số dòng xuất. Dựng 55 dòng nhật ký
		(giống test_pagination_* của test_e4_nhat_ky.py) rồi khẳng định file
		xuất ra có ĐỦ 55 dòng, không phải 50."""
		today = _today()
		for i in range(55):
			_nhap(self.K, self.VT, 1, 1000, frappe.utils.add_days(today, -60 + i),
				  so_lo=f"LO-{i:02d}")

		tu_ngay, den_ngay = _iso(frappe.utils.add_days(today, -60)), _iso(today)
		frappe.set_user(BM_USER)
		try:
			kho_api.kho_bao_cao_excel(
				loai="nhat_ky", vat_tu=self.VT, tu_ngay=tu_ngay, den_ngay=den_ngay,
			)
			content = frappe.local.response.filecontent
		finally:
			frappe.local.response.clear()
			frappe.set_user("Administrator")

		wb = load_workbook(io.BytesIO(content))
		ws = wb.active
		header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
		self.assertEqual(header, [label for label, _ in reports.NHAT_KY_COLUMNS])
		data_rows = list(ws.iter_rows(min_row=2, values_only=True))
		self.assertEqual(len(data_rows), 55)

		# Cùng dữ liệu với reports.nhat_ky_rows_export() — không một đường
		# tính riêng nào khác cho Excel.
		screen_rows = reports.nhat_ky_rows_export(self.K, self.VT, tu_ngay, den_ngay)
		self.assertEqual(len(screen_rows), 55)
		field_order = [field for _, field in reports.NHAT_KY_COLUMNS]
		for excel_row, screen_row in zip(data_rows, screen_rows):
			for col_idx, field in enumerate(field_order):
				self.assertEqual(
					_norm_cell(excel_row[col_idx]), screen_row[field],
					msg=f"cột {field} lệch giữa Excel và nhat_ky_rows_export",
				)

	def test_dot_excel_matches_json_endpoint(self):
		today = _today()
		_nhap(self.K, self.VT, 100, 1000, frappe.utils.add_days(today, -30), so_lo="L1")
		tu_ngay, den_ngay = _iso(frappe.utils.add_days(today, -60)), _iso(today)

		screen_rows = reports.bao_cao_dot_rows(self.K, tu_ngay, den_ngay, vat_tu=self.VT)
		self.assertTrue(screen_rows)

		frappe.set_user(BM_USER)
		try:
			kho_api.kho_bao_cao_excel(
				loai="dot", vat_tu=self.VT, tu_ngay=tu_ngay, den_ngay=den_ngay,
			)
			content = frappe.local.response.filecontent
		finally:
			frappe.local.response.clear()
			frappe.set_user("Administrator")

		wb = load_workbook(io.BytesIO(content))
		ws = wb.active
		header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
		self.assertEqual(header, [label for label, _ in reports.DOT_COLUMNS])
		data_rows = list(ws.iter_rows(min_row=2, values_only=True))
		self.assertEqual(len(data_rows), len(screen_rows))
		field_order = [field for _, field in reports.DOT_COLUMNS]
		for excel_row, screen_row in zip(data_rows, screen_rows):
			for col_idx, field in enumerate(field_order):
				self.assertEqual(_norm_cell(excel_row[col_idx]), screen_row[field])

	def test_excel_export_rejects_vat_tu_of_another_customer(self):
		frappe.set_user(BM_USER)
		today = _today()
		try:
			with self.assertRaises(frappe.PermissionError):
				kho_api.kho_bao_cao_excel(
					loai="the_kho", vat_tu=self.kho["vt_pxn"],
					tu_ngay=_iso(frappe.utils.add_days(today, -30)),
					den_ngay=_iso(today),
				)
		finally:
			frappe.local.response.clear()
			frappe.set_user("Administrator")

	def test_excel_export_rejects_unknown_loai(self):
		frappe.set_user(BM_USER)
		try:
			with self.assertRaises(frappe.ValidationError):
				kho_api.kho_bao_cao_excel(loai="khong-hop-le")
		finally:
			frappe.local.response.clear()
			frappe.set_user("Administrator")


class TestKhoBaoCaoIsolation(_KhoBmTestCase):
	def test_nxt_drilldown_rejects_other_customers_item(self):
		frappe.set_user(BM_USER)
		today = _today()
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_bao_cao_nxt(
				tu_ngay=_iso(frappe.utils.add_days(today, -30)),
				den_ngay=_iso(today),
				vat_tu=self.kho["vt_pxn"],
			)

	def test_the_kho_rejects_other_customers_item(self):
		frappe.set_user(BM_USER)
		today = _today()
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_the_kho(
				vat_tu=self.kho["vt_pxn"],
				tu_ngay=_iso(frappe.utils.add_days(today, -30)),
				den_ngay=_iso(today),
			)

	def test_own_customer_data_reachable_positive_control(self):
		"""Kèm theo hai test phủ định ở trên: chứng minh cách ly không lỡ tay
		chặn luôn dữ liệu của chính chủ."""
		today = _today()
		_nhap(self.K, self.VT, 5, 1000, frappe.utils.add_days(today, -1))
		frappe.set_user(BM_USER)
		out = kho_api.kho_bao_cao_nxt(
			tu_ngay=_iso(frappe.utils.add_days(today, -5)),
			den_ngay=_iso(today),
		)
		self.assertTrue(any(r["vat_tu"] == self.VT for r in out["rows"]))

	def test_nxt_report_never_includes_other_customers_items(self):
		"""Positive control ngược: seed CẢ HAI khách rồi chứng minh báo cáo
		của BM không chứa vật tư của PXN — tránh lặp lại lỗi "test phủ định
		pass vì bên kia chưa có dữ liệu" đã ghi nhận nhiều lần trong dự án."""
		today = _today()
		_nhap(self.kho["kho_pxn"], self.kho["vt_pxn"], 99, 5000,
			  frappe.utils.add_days(today, -1))
		frappe.set_user(BM_USER)
		out = kho_api.kho_bao_cao_nxt(
			tu_ngay=_iso(frappe.utils.add_days(today, -30)),
			den_ngay=_iso(today),
		)
		self.assertTrue(all(r["vat_tu"] != self.kho["vt_pxn"] for r in out["rows"]))
