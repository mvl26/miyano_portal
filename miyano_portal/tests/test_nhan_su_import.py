"""Nhập nhân sự bệnh viện từ Excel (Task 15) — tests/test_nhan_su_import.py.

Mỗi test dưới đây canh MỘT chốt cụ thể của `miyano_portal/api/nhan_su.py`.
Chốt lớn nhất là bước XEM TRƯỚC: tạo tài khoản đăng nhập là việc khó lùi,
nên "xem trước không ghi gì" được đo bằng cách đếm bản ghi của CẢ NĂM thứ
mà đường cấp tài khoản đụng tới (User, Contact, User Permission, Portal
Member, Customer Department), không phải chỉ hai cái cuối.

`FrappeTestCase` chỉ rollback MỘT LẦN cho cả LỚP, nên bản ghi do test method
này tạo vẫn còn nguyên khi method sau chạy — vì vậy mọi thứ phải được dọn
TƯỜNG MINH ở setUp (dọn trước, không chỉ dọn sau: một lần chạy đứt gánh không
được để lại tài khoản đăng nhập được trên site).

ĐÍNH CHÍNH (vòng sửa 1): bản đầu của file này ghi lý do là *"`User.insert()`
tự commit bên trong"*. **SAI** — `test_ghi_hong_giua_chung_khong_de_lai_gi`
bên dưới đo thẳng điều đó: hai tài khoản được tạo rồi lần ghi vỡ ở người thứ
ba, và điểm lưu (savepoint) cuốn sạch cả hai. Không có `commit` nào trong
đường ghi. Việc dọn tường minh vẫn cần (lý do ở trên), nhưng lý do cũ thì
sai, và người sau sẽ tin nó.
"""

import hashlib
import io

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils.password import check_password, update_password
from openpyxl import Workbook, load_workbook

from miyano_portal.api import nhan_su as nhan_su_api
from miyano_portal.api import portal as portal_api

CUST_A = "ZZTEST NS Bệnh viện A"
CUST_B = "ZZTEST NS Bệnh viện B"
CUST_KHONG_MA_NGAN = "ZZTEST NS Bệnh viện C"
DOMAIN = "@zztest.miyano"

HOA = f"hoa{DOMAIN}"
BINH = f"binh{DOMAIN}"
CUC = f"cuc{DOMAIN}"
DUNG = f"dung{DOMAIN}"

HEADERS = [label for label, _ in nhan_su_api.COLUMNS]

def _xlsx_bytes(rows, headers=None):
	wb = Workbook()
	ws = wb.active
	ws.append(headers if headers is not None else HEADERS)
	for row in rows:
		ws.append(row)
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


# VÒNG SỬA 2 (04/09/2026): các bài kiểm CHUẨN HOÁ số điện thoại cần một số
# CỐ ĐỊNH lặp lại nhiều lần (nhiều cách viết cùng chuẩn hoá về MỘT số) —
# không dùng được `_sdt_gia_theo_email()` (băm theo email, mỗi email ra một
# số khác nhau) cho các bài đó. Số CŨ của các bài này là "0912345678" —
# TRÙNG ĐÚNG số một tài khoản demo (`quanly.demoe2e@miyano-test.vn`) để lại
# trên site dùng chung `erptest.local` sau lượt chạy thử tay 04/09/2026. Từ
# khi `_phan_tich()` biết soi trùng với `User` đã tồn tại (VÒNG SỬA 2, xem
# `api/nhan_su.py::_kiem_trung_dien_thoai`), các bài dùng "0912345678" không
# còn kiểm đúng thứ chúng viết ra để kiểm (CHUẨN HOÁ) nữa — chúng đổi hướng
# thành bài kiểm TRÙNG SỐ (đã có lớp riêng, `TestDienThoaiTrung`, bên dưới).
# Đổi sang một số KHÔNG nằm trong bất kỳ tệp mẫu, tài liệu demo, hay dữ liệu
# chạy thử nào đã biết — bài test không được phụ thuộc vào việc site dùng
# chung có đang "sạch" demo data hay không.
SDT_CHUAN_HOA_TEST = "0938271099"


def _sdt_gia_theo_email(email: str) -> str:
	"""SĐT giả HỢP LỆ và DUY NHẤT theo email — `User.mobile_no` có ràng buộc
	UNIQUE ở tầng database; một hằng số dùng chung cho mọi dòng của `_row()`
	sẽ vỡ ngay khi tệp có từ hai người trở lên (`test_ghi_tao_du_tai_khoan...`
	tạo ba tài khoản trong một lần commit)."""
	so = int(hashlib.md5(email.encode()).hexdigest(), 16) % 10**8
	return f"09{so:08d}"


def _row(ho_ten, email, ten_khoa="", ma_khoa="", vai_tro="Nhân viên khoa", dien_thoai=None):
	if dien_thoai is None:
		dien_thoai = _sdt_gia_theo_email(email)
	return [ho_ten, email, ten_khoa, ma_khoa, vai_tro, dien_thoai]


TEP_HOP_LE = [
	_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý"),
	_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC"),
	_row("Lê Thị Cúc", CUC, "Huyết học", "HUYETHOC"),
]


class _NhanSuTestBase(FrappeTestCase):
	def setUp(self):
		frappe.set_user("Administrator")
		self._created_files = []
		self._don_sach()
		self.addCleanup(self._don_sach)
		self.addCleanup(frappe.set_user, "Administrator")
		self._tao_khach(CUST_A, "ZZNSA")
		self._tao_khach(CUST_B, "ZZNSB")

	def tearDown(self):
		frappe.set_user("Administrator")
		for name in self._created_files:
			try:
				frappe.delete_doc("File", name, ignore_permissions=True, force=True)
			except Exception:
				pass

	def _don_sach(self):
		frappe.set_user("Administrator")
		khach = [CUST_A, CUST_B, CUST_KHONG_MA_NGAN]
		emails = frappe.get_all("User", filters={"name": ["like", f"%{DOMAIN}"]}, pluck="name")
		if emails:
			frappe.db.delete("User Permission", {"user": ["in", emails]})
			frappe.db.delete("Portal Member", {"user": ["in", emails]})
			frappe.db.delete("User", {"name": ["in", emails]})
		# Contact PHẢI xoá qua delete_doc, KHÔNG phải frappe.db.delete: mỗi
		# tài khoản mới sinh ra HAI Contact (Frappe tự tạo một cái ở
		# `user.py::create_contact`, `portal_provision` tạo thêm bản ghi liên
		# hệ của cổng), và mỗi cái mang bảng con `Contact Email`. Xoá thẳng
		# bảng cha để lại dòng con mồ côi; lần chạy sau, Contact mới trùng
		# tên "hút" lại đám con đó và vỡ ở "Only one Email ID can be set as
		# primary" — một cái bẫy chỉ nổ ở LẦN CHẠY THỨ HAI.
		mo_coi = frappe.get_all("Contact Email", filters={"email_id": ["like", f"%{DOMAIN}"]}, pluck="parent")
		lien_he = frappe.get_all("Contact", filters={"user": ["like", f"%{DOMAIN}"]}, pluck="name")
		for ten in set(mo_coi) | set(lien_he):
			frappe.delete_doc(
				"Contact", ten, force=True, ignore_permissions=True, ignore_missing=True
			)
			frappe.db.delete("Contact Email", {"parent": ten})
			frappe.db.delete("Contact Phone", {"parent": ten})
			frappe.db.delete("Dynamic Link", {"parenttype": "Contact", "parent": ten})
		frappe.db.delete("Portal Member", {"customer": ["in", khach]})
		frappe.db.delete("Customer Department", {"customer": ["in", khach]})
		frappe.db.delete("Dynamic Link", {"parenttype": "Contact", "link_name": ["in", khach]})
		frappe.db.delete("Customer", {"name": ["in", khach]})
		# Commit việc dọn: nếu một lần chạy TRƯỚC đã kịp commit ở đâu đó (hoặc
		# bị giết giữa chừng), rác của nó nằm ngoài transaction hiện tại và chỉ
		# biến mất thật khi lần dọn này được commit — cú rollback cuối lớp
		# không phân biệt được "xoá rác cũ" với "xoá dữ liệu test mới", nó
		# hoàn tác cả hai. KHÔNG phải vì `User.insert()` tự commit: nó không
		# commit, xem đính chính ở đầu file.
		frappe.db.commit()

	def _tao_khach(self, ten, ma_ngan=None):
		doc = frappe.get_doc({
			"doctype": "Customer", "customer_name": ten, "customer_type": "Company",
			"customer_group": "All Customer Groups", "territory": "All Territories",
		})
		doc.insert(ignore_permissions=True)
		if ma_ngan:
			frappe.db.set_value("Customer", doc.name, "custom_ma_ngan", ma_ngan)
		return doc.name

	def _upload(self, content: bytes, filename="nhan_su.xlsx", user="Administrator", rieng_tu=1):
		frappe.set_user(user)
		file_doc = frappe.get_doc({
			"doctype": "File", "file_name": filename, "is_private": rieng_tu, "content": content,
		})
		file_doc.insert(ignore_permissions=True)
		self._created_files.append(file_doc.name)
		return file_doc

	def _upload_cong_khai(self, content: bytes, filename="nhan_su_cong_khai.xlsx"):
		return self._upload(content, filename=filename, rieng_tu=0)

	def _counts(self):
		"""Đếm CẢ NĂM thứ mà đường cấp tài khoản ghi ra."""
		khach = [CUST_A, CUST_B, CUST_KHONG_MA_NGAN]
		return (
			frappe.db.count("User", {"name": ["like", f"%{DOMAIN}"]}),
			frappe.db.count("Contact", {"user": ["like", f"%{DOMAIN}"]}),
			frappe.db.count("User Permission", {"user": ["like", f"%{DOMAIN}"]}),
			frappe.db.count("Portal Member", {"customer": ["in", khach]}),
			frappe.db.count("Customer Department", {"customer": ["in", khach]}),
		)

	def _dong(self, ket_qua, email):
		for row in ket_qua["rows"]:
			if row["email"] == email:
				return row
		self.fail(f"Không thấy dòng {email} trong kết quả: {ket_qua['rows']}")


class TestXemTruoc(_NhanSuTestBase):
	def test_xem_truoc_khong_ghi_gi(self):
		"""Chốt chính của task: bước xem trước phải LIỆT KÊ mà KHÔNG GHI."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		before = self._counts()

		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		self.assertEqual(before, self._counts(), "xem trước không được ghi gì vào database")
		self.assertEqual(ket_qua["total"], 3)
		self.assertEqual(ket_qua["so_tao_moi"], 3)
		for email in (HOA, BINH, CUC):
			self.assertEqual(self._dong(ket_qua, email)["trang_thai"], "tao_moi")

	def test_xem_truoc_noi_ro_se_tao_khoa_moi(self):
		"""QĐ-G20: khoa chưa có thì tự tạo, nhưng phải NÓI RÕ — gõ nhầm tên
		khoa không được lặng lẽ đẻ ra khoa rác."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		# Hai dòng cùng trỏ về một khoa mới -> chỉ báo tạo MỘT khoa.
		self.assertEqual(len(ket_qua["khoa_se_tao"]), 1, ket_qua["khoa_se_tao"])
		self.assertEqual(ket_qua["khoa_se_tao"][0]["ten_khoa_phong"], "Huyết học")
		self.assertEqual(ket_qua["khoa_se_tao"][0]["ma_khoa"], "HUYETHOC")
		self.assertTrue(self._dong(ket_qua, BINH)["khoa_moi"])

	def test_khoa_da_co_thi_dung_lai_khong_tao_them(self):
		kp = frappe.get_doc({
			"doctype": "Customer Department", "customer": CUST_A,
			"ten_khoa_phong": "Huyết học", "ma_khoa": "HUYETHOC",
		}).insert(ignore_permissions=True)
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		self.assertEqual(ket_qua["khoa_se_tao"], [])
		self.assertEqual(self._dong(ket_qua, BINH)["khoa"], kp.name)
		self.assertFalse(self._dong(ket_qua, BINH)["khoa_moi"])


class TestChotHopLe(_NhanSuTestBase):
	def test_nhan_vien_khoa_thieu_khoa_bi_tu_choi_kem_so_dong(self):
		"""Đúng trạng thái kẹt mà task này sinh ra để dẹp."""
		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý"),
			_row("Trần Văn Bình", BINH),  # Nhân viên khoa, KHÔNG có khoa
		]))
		before = self._counts()
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertEqual(dong["line"], 3)  # 1 = header, 2 = Hoa, 3 = Bình
		self.assertIn("khoa", " ".join(dong["errors"]).lower())
		self.assertEqual(ket_qua["so_tu_choi"], 1)

		# Tất-cả-hoặc-không: một dòng lỗi thì KHÔNG dòng nào được ghi.
		with self.assertRaises(frappe.ValidationError) as cm:
			nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertIn("dòng 3", str(cm.exception))
		self.assertIn("chưa có dữ liệu nào được ghi", str(cm.exception))
		self.assertEqual(before, self._counts())

	def test_quan_ly_duoc_de_trong_khoa(self):
		"""Quản lý nhìn toàn viện — bỏ trống khoa là HỢP LỆ, không phải lỗi."""
		f = self._upload(_xlsx_bytes([_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(self._dong(ket_qua, HOA)["trang_thai"], "tao_moi")

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		tv = frappe.db.get_value(
			"Portal Member", {"user": HOA}, ["customer", "vai_tro", "khoa_phong", "active"],
			as_dict=True,
		)
		self.assertEqual(tv.customer, CUST_A)
		self.assertEqual(tv.vai_tro, "Quản lý")
		self.assertFalse(tv.khoa_phong)
		self.assertEqual(tv.active, 1)

	def test_quan_ly_co_khoa_bi_tu_choi(self):
		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, "Huyết học", "HUYETHOC", "Quản lý"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(self._dong(ket_qua, HOA)["trang_thai"], "tu_choi")
		self.assertIn("Quản lý", " ".join(self._dong(ket_qua, HOA)["errors"]))

	def test_quan_ly_thu_hai_bi_tu_choi_kem_ten_quan_ly_dang_co(self):
		"""`_chan_hai_quan_ly` phải được chạy TRƯỚC ở bước xem trước — nếu
		không, người nhập chỉ biết khi commit nổ giữa chừng."""
		nhan_su_api.nhan_su_import_commit(
			CUST_A, self._upload(_xlsx_bytes([_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý")])).file_url
		)
		f = self._upload(_xlsx_bytes([_row("Phạm Văn Dũng", DUNG, vai_tro="Quản lý")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		dong = self._dong(ket_qua, DUNG)
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertIn(HOA, " ".join(dong["errors"]))

	def test_hai_dong_quan_ly_trong_cung_tep_bi_tu_choi(self):
		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý"),
			_row("Phạm Văn Dũng", DUNG, vai_tro="Quản lý"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(self._dong(ket_qua, HOA)["trang_thai"], "tao_moi")
		self.assertEqual(self._dong(ket_qua, DUNG)["trang_thai"], "tu_choi")

	def test_trung_email_trong_cung_tep_bi_tu_choi(self):
		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý"),
			_row("Hoa (gõ lại)", HOA, "Huyết học", "HUYETHOC"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(ket_qua["so_tu_choi"], 1)
		self.assertEqual(ket_qua["rows"][1]["trang_thai"], "tu_choi")
		self.assertIn("dòng 2", " ".join(ket_qua["rows"][1]["errors"]))

	def test_email_noi_bo_miyano_bi_tu_choi(self):
		"""Gõ nhầm email nhân viên Miyano vào tệp của bệnh viện thì
		`portal_provision` sẽ gắn User Permission trên Customer cho tài khoản
		nội bộ đó — người ấy mất tầm nhìn ở khắp ERPNext và gần như không ai
		lần ra vì sao. Chặn ở xem trước, đừng để ghi rồi mới biết."""
		noi_bo = f"noibo{DOMAIN}"
		u = frappe.get_doc({
			"doctype": "User", "email": noi_bo, "first_name": "Nhân viên Miyano",
			"user_type": "System User", "send_welcome_email": 0,
		})
		# PHẢI có một role mở được Desk: `User.validate()` tự hạ user_type
		# xuống "Website User" cho tài khoản không có role nào vào được Desk —
		# thiếu dòng này, "System User" trong fixture chỉ là chữ, và bài test
		# sẽ canh một trạng thái mà code không bao giờ gặp.
		u.append("roles", {"role": "Sales User"})
		u.insert(ignore_permissions=True)
		self.assertEqual(frappe.db.get_value("User", noi_bo, "user_type"), "System User")

		f = self._upload(_xlsx_bytes([_row("Nhân viên Miyano", noi_bo, vai_tro="Quản lý")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = ket_qua["rows"][0]
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertIn("System User", " ".join(dong["errors"]))

		with self.assertRaises(frappe.ValidationError) as cm:
			nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertIn("chưa có dữ liệu nào được ghi", str(cm.exception))
		self.assertFalse(frappe.db.exists("Portal Member", {"user": noi_bo}))
		self.assertFalse(frappe.db.exists(
			"User Permission", {"user": noi_bo, "allow": "Customer", "for_value": CUST_A}
		))

	def test_thieu_ma_ngan_thi_tu_choi_ca_tep_va_noi_ro(self):
		"""`_chan_thieu_ma_ngan` là chốt cấp KHÁCH HÀNG — nói ra ở xem trước,
		đừng để commit nổ."""
		self._tao_khach(CUST_KHONG_MA_NGAN)
		f = self._upload(_xlsx_bytes([_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_KHONG_MA_NGAN, f.file_url)

		self.assertIn("Mã ngắn", " ".join(ket_qua["loi_toan_tep"]))
		self.assertEqual(self._dong(ket_qua, BINH)["trang_thai"], "tu_choi")


class TestGhiThat(_NhanSuTestBase):
	def test_ghi_tao_du_tai_khoan_khoa_phong_va_phan_quyen(self):
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		ket_qua = nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		self.assertEqual(ket_qua["so_tao_moi"], 3)
		# 3 User, SÁU Contact, 3 User Permission, 3 Portal Member, 1 khoa.
		# Sáu Contact là hành vi CÓ SẴN của `portal_provision` (không phải
		# của task này): Frappe tự tạo một Contact cho mỗi User mới
		# (`user.py::create_contact`), rồi `portal_provision` tạo thêm bản
		# ghi liên hệ riêng của cổng — cái NỐI Contact với Customer. Khẳng
		# định thẳng mối nối đó ở dưới để con số 6 không thành số ma.
		self.assertEqual(self._counts(), (3, 6, 3, 3, 1))
		lien_he = frappe.get_all("Contact", filters={"user": BINH}, pluck="name")
		self.assertEqual(len(lien_he), 2)
		self.assertTrue(frappe.db.exists("Dynamic Link", {
			"parenttype": "Contact", "parent": ["in", lien_he],
			"link_doctype": "Customer", "link_name": CUST_A,
		}))

		kp = frappe.db.get_value(
			"Customer Department", {"customer": CUST_A, "ma_khoa": "HUYETHOC"},
			["name", "ten_khoa_phong"], as_dict=True,
		)
		self.assertEqual(kp.ten_khoa_phong, "Huyết học")

		tv = frappe.db.get_value(
			"Portal Member", {"user": BINH}, ["customer", "vai_tro", "khoa_phong", "active"],
			as_dict=True,
		)
		self.assertEqual(tv.customer, CUST_A)
		self.assertEqual(tv.vai_tro, "Nhân viên khoa")
		self.assertEqual(tv.khoa_phong, kp.name)
		self.assertEqual(tv.active, 1, "tài khoản nhập từ tệp phải DÙNG ĐƯỢC ngay")

	def test_first_name_la_ten_nguoi_khong_phai_ten_benh_vien(self):
		"""QĐ-G21: `portal_provision` cũ đặt first_name = tên khách hàng —
		gốc của lỗi truy vết đã phải vá ở tầng hiển thị (97fd6e2). Ở đây
		tờ khai thắng."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		self.assertEqual(frappe.db.get_value("User", BINH, "first_name"), "Trần Văn Bình")
		self.assertNotEqual(frappe.db.get_value("User", BINH, "first_name"), CUST_A)

	def test_vai_tro_theo_to_khai_khong_theo_luat_ngam_nguoi_dau_tien(self):
		"""QĐ-G21 vế hai: `portal_provision` có luật ngầm "tài khoản đầu tiên
		của một bệnh viện là Quản lý". Tờ khai thắng — dòng đầu tệp là Nhân
		viên khoa thì nó PHẢI là Nhân viên khoa, dù bệnh viện chưa có ai."""
		f = self._upload(_xlsx_bytes([_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC")]))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		tv = frappe.db.get_value(
			"Portal Member", {"user": BINH}, ["vai_tro", "khoa_phong", "active"], as_dict=True,
		)
		self.assertEqual(tv.vai_tro, "Nhân viên khoa")
		self.assertTrue(tv.khoa_phong)
		self.assertEqual(tv.active, 1)

	def test_mat_khau_tra_ve_MOT_LAN_va_dat_that(self):
		"""QĐ-G19: Miyano đặt mật khẩu, màn hình hiện MỘT LẦN sau khi ghi để
		chép ra bàn giao. Không nằm trong tệp, không gửi email, không ghi log."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		ket_qua = nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		self.assertEqual(sorted(ket_qua["mat_khau"]), sorted([HOA, BINH, CUC]))
		# Mật khẩu trả về phải là mật khẩu THẬT của tài khoản vừa tạo —
		# check_password ném AuthenticationError nếu sai.
		check_password(BINH, ket_qua["mat_khau"][BINH])

		# Bắt đổi ở lần đăng nhập đầu: Frappe v15 chỉ có chính sách theo
		# SỐ NGÀY ở System Settings (`force_user_to_reset_password`), không
		# có cờ cho từng người — đặt mốc đổi mật khẩu về quá khứ để chính
		# sách đó (khi bật) chộp đúng những tài khoản này ngay lần đầu.
		moc = frappe.db.get_value("User", BINH, "last_password_reset_date")
		self.assertTrue(moc)
		self.assertLess(frappe.utils.getdate(moc), frappe.utils.getdate(frappe.utils.today()))

	def test_nhap_lai_cung_tep_khong_de_trung(self):
		"""QĐ-G22 — khớp theo email."""
		f1 = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f1.file_url)
		sau_lan_1 = self._counts()

		f2 = self._upload(_xlsx_bytes(TEP_HOP_LE), filename="nhan_su_lan_2.xlsx")
		xem_truoc = nhan_su_api.nhan_su_import_preview(CUST_A, f2.file_url)
		self.assertEqual(xem_truoc["so_bo_qua"], 3)
		self.assertEqual(self._dong(xem_truoc, BINH)["trang_thai"], "bo_qua")

		ket_qua = nhan_su_api.nhan_su_import_commit(CUST_A, f2.file_url)
		self.assertEqual(ket_qua["so_tao_moi"], 0)
		self.assertEqual(ket_qua["mat_khau"], {})
		self.assertEqual(sau_lan_1, self._counts(), "nhập lại cùng tệp không được đẻ trùng")


class TestEmailCuaKhachKhac(_NhanSuTestBase):
	def test_chi_bao_o_xem_truoc_khong_tao_khong_doi_mat_khau(self):
		"""QĐ-G23 — có thể là người thật làm ở hai nơi, cũng có thể là gõ
		nhầm: báo và để Miyano quyết, không tự xử."""
		nhan_su_api.nhan_su_import_commit(
			CUST_A, self._upload(_xlsx_bytes([_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý")])).file_url
		)
		mat_khau_cu = "ZZtest-MatKhau-Cu-1"
		update_password(HOA, mat_khau_cu)

		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý"),
			_row("Phạm Văn Dũng", DUNG, "Hồi sức", "HOISUC"),
		]), filename="nhan_su_b.xlsx")
		xem_truoc = nhan_su_api.nhan_su_import_preview(CUST_B, f.file_url)

		dong = self._dong(xem_truoc, HOA)
		self.assertEqual(dong["trang_thai"], "canh_bao")
		self.assertIn(CUST_A, " ".join(dong["errors"]))
		self.assertEqual(xem_truoc["so_canh_bao"], 1)

		# Cảnh báo KHÔNG chặn các dòng còn lại — nhưng dòng cảnh báo thì
		# không tạo gì, không đổi gì.
		ket_qua = nhan_su_api.nhan_su_import_commit(CUST_B, f.file_url)
		self.assertEqual(ket_qua["so_tao_moi"], 1)
		self.assertEqual(ket_qua["so_canh_bao"], 1)
		self.assertNotIn(HOA, ket_qua["mat_khau"])
		self.assertEqual(
			frappe.db.get_value("Portal Member", {"user": HOA}, "customer"), CUST_A
		)
		self.assertFalse(frappe.db.exists(
			"User Permission", {"user": HOA, "allow": "Customer", "for_value": CUST_B}
		))
		check_password(HOA, mat_khau_cu)  # mật khẩu của họ KHÔNG bị đổi


class TestCachLy(_NhanSuTestBase):
	def test_tep_cua_khach_A_tao_dung_nguoi_cua_A_va_khong_dung_gi_cua_B(self):
		"""QĐ-G18: khách hàng chọn TRÊN MÀN HÌNH, không nằm trong tệp — một
		cột "tên bệnh viện" gõ tay là đường nhập nhầm người sang viện khác."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		# VẾ DƯƠNG: đúng ba người đó thuộc về A.
		self.assertEqual(
			sorted(frappe.get_all(
				"Portal Member", filters={"customer": CUST_A}, pluck="user")),
			sorted([HOA, BINH, CUC]),
		)
		self.assertEqual(
			frappe.get_all("Customer Department", filters={"customer": CUST_A}, pluck="ten_khoa_phong"),
			["Huyết học"],
		)
		# VẾ ÂM: không gì của B bị đụng tới.
		self.assertEqual(frappe.get_all("Portal Member", filters={"customer": CUST_B}), [])
		self.assertEqual(frappe.get_all("Customer Department", filters={"customer": CUST_B}), [])
		for email in (HOA, BINH, CUC):
			self.assertFalse(frappe.db.exists(
				"User Permission", {"user": email, "allow": "Customer", "for_value": CUST_B}
			))

	def test_khach_cong_khong_goi_duoc(self):
		"""Đây là màn của nhân viên Miyano trong Desk — một tài khoản cổng
		(Website User) không được cấp tài khoản cho bất kỳ ai."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		frappe.set_user(HOA)
		with self.assertRaises(frappe.PermissionError):
			nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		with self.assertRaises(frappe.PermissionError):
			nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		frappe.set_user("Administrator")

	def test_khach_hang_khong_ton_tai_bi_chan(self):
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		with self.assertRaises(frappe.ValidationError) as cm:
			nhan_su_api.nhan_su_import_preview("ZZTEST KHONG CO KHACH NAY", f.file_url)
		self.assertIn("khách hàng", str(cm.exception).lower())


class TestTepMau(_NhanSuTestBase):
	def test_tai_mau_roi_nap_lai_ngay_thi_doc_duoc(self):
		"""Mẫu → xem trước phải đi lọt: tệp mẫu là thứ bệnh viện điền lên."""
		nhan_su_api.nhan_su_import_template()
		noi_dung = frappe.local.response.filecontent
		self.assertTrue(frappe.local.response.filename.endswith(".xlsx"))

		f = self._upload(noi_dung, filename="mau_tai_ve.xlsx")
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(ket_qua["loi_toan_tep"], [])
		self.assertEqual(ket_qua["so_tu_choi"], 0, ket_qua["rows"])
		self.assertGreaterEqual(ket_qua["total"], 1)

	def test_mau_co_sau_cot_va_hai_dong_vi_du_deu_co_so(self):
		"""Task 10, bài 1: tệp mẫu phải làm gương cho cột mới — bỏ trống ví dụ
		sẽ dạy người điền rằng để trống Số điện thoại là bình thường."""
		nhan_su_api.nhan_su_import_template()
		noi_dung = frappe.local.response.filecontent
		ws = load_workbook(io.BytesIO(noi_dung)).active

		header = [c.value for c in ws[1]]
		self.assertEqual(len(header), 6, header)
		self.assertEqual(header[5], "Số điện thoại")
		self.assertTrue(str(ws.cell(row=2, column=6).value or "").strip(), "dòng ví dụ 1 thiếu SĐT")
		self.assertTrue(str(ws.cell(row=3, column=6).value or "").strip(), "dòng ví dụ 2 thiếu SĐT")

		# Tải mẫu xuống rồi nạp lại ngay phải lọt xem trước KHÔNG một cảnh
		# báo nào — nếu ví dụ nào bỏ trống SĐT, bài `so_canh_bao == 0` này đỏ.
		f = self._upload(noi_dung, filename="mau_6_cot.xlsx")
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertEqual(ket_qua["so_tu_choi"], 0, ket_qua["rows"])
		self.assertEqual(ket_qua["so_canh_bao"], 0, ket_qua["rows"])
		self.assertEqual(ket_qua["so_tao_moi"], 2)


class TestDienThoai(_NhanSuTestBase):
	"""Task 10 — cột Số điện thoại. Bốn quyết định điều phối (QĐ-1..4) của
	brief, một bài (lớp con) cho mỗi mảnh hành vi khác nhau."""

	def test_thieu_han_cot_dien_thoai_bao_loi_header_khong_ghi_gi(self):
		"""Bài 2: tệp 5 cột cũ (chưa có Số điện thoại) phải bị chặn ngay ở
		bước ĐỌC HEADER — không lặng lẽ tạo cả viện không số."""
		f = self._upload(_xlsx_bytes(
			[["Nguyễn Thị Hoa", HOA, "", "", "Quản lý"]],
			headers=HEADERS[:5],
		))
		before = self._counts()
		with self.assertRaises(frappe.ValidationError) as cm:
			nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertIn("Số điện thoại", str(cm.exception))
		self.assertEqual(before, self._counts(), "báo lỗi header thì không được ghi gì")

	def test_o_trong_van_tao_tai_khoan_kem_canh_bao_o_ghi_chu_va_cap_tep(self):
		"""Bài 3 (QĐ-1, VÒNG SỬA 1 04/09/2026): chủ đầu tư chọn tường minh — ô
		trống chỉ CẢNH BÁO, KHÔNG hoãn cấp tài khoản. Brief gốc bảo dùng đúng
		trạng thái `CANH_BAO` sẵn có, nhưng `CANH_BAO` trong file này mang nghĩa
		"gắn cờ VÀ KHÔNG GHI" (`_ghi()`: `if trang_thai != TAO_MOI: continue`) —
		sai với ý chủ đầu tư. Sửa: dòng thiếu số vẫn `TAO_MOI`, cảnh báo dồn vào
		`ghi_chu` của dòng đó VÀ vào `canh_bao_toan_tep` cấp tệp (để người duyệt
		thấy ngay đầu màn, không phải dò từng dòng)."""
		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý", dien_thoai=""),
			_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC"),  # SĐT mặc định hợp lệ
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		dong_hoa = self._dong(ket_qua, HOA)
		self.assertEqual(dong_hoa["trang_thai"], "tao_moi", "thiếu số KHÔNG hoãn cấp tài khoản")
		self.assertNotEqual(dong_hoa["trang_thai"], "canh_bao")
		self.assertNotEqual(dong_hoa["trang_thai"], "tu_choi")
		self.assertIn("không gọi được", dong_hoa["ghi_chu"].lower())
		self.assertEqual(self._dong(ket_qua, BINH)["trang_thai"], "tao_moi")
		self.assertEqual(ket_qua["so_canh_bao"], 0)
		self.assertEqual(ket_qua["so_tao_moi"], 2)

		# Cấp TỆP: liệt kê TÊN người thiếu số — đây là thứ biến "cảnh báo" từng
		# dòng thành việc làm được, không phải chữ nằm im chờ ai đó dò ra.
		canh_bao_tep = " ".join(ket_qua["canh_bao_toan_tep"])
		self.assertIn("Nguyễn Thị Hoa", canh_bao_tep)
		self.assertIn(HOA, canh_bao_tep)

		# Vế QUAN TRỌNG NHẤT (suýt lọt ở vòng đầu): tài khoản THẬT SỰ được tạo,
		# kể cả người thiếu số.
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertTrue(frappe.db.exists("User", HOA), "thiếu số nhưng vẫn phải được tạo")
		self.assertTrue(frappe.db.exists("User", BINH))
		self.assertFalse(frappe.db.get_value("User", HOA, "mobile_no"))

	def test_sai_dinh_dang_thi_tu_choi_kem_ly_do(self):
		"""Bài 4 (QĐ-2): gõ sai (khác hẳn bỏ trống) là lỗi im lặng sẽ in ra
		trước mặt bệnh viện — phải TỪ CHỐI, không phải cảnh báo."""
		f = self._upload(_xlsx_bytes([
			_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC", dien_thoai="abc123"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertIn("điện thoại", " ".join(dong["errors"]).lower())

	def test_excel_nuot_so_0_dau_duoc_phuc_hoi(self):
		"""Bài 5 (QĐ-3) — cạm bẫy lớn nhất của cả task: ô định dạng SỐ trong
		Excel làm mất số 0 đứng đầu. Ghi số dưới dạng `int` (đúng hình
		openpyxl trả về khi ai đó gõ số điện thoại vào một ô định dạng số,
		không phải ô định dạng Văn bản)."""
		f = self._upload(_xlsx_bytes([
			_row(
				"Trần Văn Bình", BINH, "Huyết học", "HUYETHOC",
				dien_thoai=int(SDT_CHUAN_HOA_TEST),
			),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tao_moi", dong["errors"])
		self.assertEqual(dong["dien_thoai"], SDT_CHUAN_HOA_TEST)

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertEqual(frappe.db.get_value("User", BINH, "mobile_no"), SDT_CHUAN_HOA_TEST)

	def test_bon_cach_viet_hop_le_cong_excel_int_deu_ra_cung_mot_so_chuan(self):
		"""VÒNG SỬA 2 (coordinator, 04/09/2026): regex cũ áp THẲNG lên chuỗi
		thô nên bốn cách viết hợp lệ dưới đây — TIỀN TỐ QUỐC TẾ, KHOẢNG TRẮNG,
		DẤU CHẤM, GẠCH NGANG — đều bị TỪ CHỐI oan, và TỪ CHỐI kéo theo "một
		dòng bị từ chối là không ghi gì cả" — chặn cứng cả tệp, đúng cái bẫy
		QĐ-1 dựng ra để tránh, chỉ dịch sang QĐ-2. Cộng dạng `int` của QĐ-3 —
		tất cả năm cách viết phải chuẩn hoá về CÙNG một `mobile_no`."""
		cach_viet = {
			"tien_to_quoc_te": f"+84{SDT_CHUAN_HOA_TEST[1:]}",
			"khoang_trang": f"{SDT_CHUAN_HOA_TEST[:4]} {SDT_CHUAN_HOA_TEST[4:7]} {SDT_CHUAN_HOA_TEST[7:]}",
			"dau_cham": f"{SDT_CHUAN_HOA_TEST[:4]}.{SDT_CHUAN_HOA_TEST[4:7]}.{SDT_CHUAN_HOA_TEST[7:]}",
			"gach_ngang": f"{SDT_CHUAN_HOA_TEST[:4]}-{SDT_CHUAN_HOA_TEST[4:7]}-{SDT_CHUAN_HOA_TEST[7:]}",
			"excel_int": int(SDT_CHUAN_HOA_TEST),
		}
		for nhan, dien_thoai in cach_viet.items():
			with self.subTest(nhan):
				email = f"vs2_{nhan}{DOMAIN}"
				f = self._upload(_xlsx_bytes([
					_row("Người thử", email, "Huyết học", "HUYETHOC", dien_thoai=dien_thoai),
				]), filename=f"vs2_{nhan}.xlsx")
				ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
				dong = self._dong(ket_qua, email)
				self.assertEqual(
					dong["trang_thai"], "tao_moi",
					f'{nhan} ({dien_thoai!r}) bị từ chối oan: {dong["errors"]}',
				)
				self.assertEqual(
					dong["dien_thoai"], SDT_CHUAN_HOA_TEST, f"{nhan}: {dien_thoai!r}",
				)

	def test_gia_chuoi_con_chu_van_bi_tu_choi_sau_khi_lot_dau_phan_cach(self):
		"""Vế ÂM (coordinator yêu cầu giữ nguyên): lột dấu phân cách không
		được biến một chuỗi THẬT SỰ KHÔNG PHẢI số điện thoại thành hợp lệ —
		còn chữ cái sau khi lột thì vẫn phải TỪ CHỐI."""
		# 10 chữ số đúng như một số hợp lệ NẾU chữ "X" bị lột theo — cài đặt cố
		# ý để một phép lột "quá tay" (lột luôn cả chữ) sẽ ÂM THẦM biến chuỗi
		# này thành "0912345678" hợp lệ, thay vì lộ ra ở bước kiểm định dạng.
		f = self._upload(_xlsx_bytes([
			_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC", dien_thoai="091X2345678"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertIn("điện thoại", " ".join(dong["errors"]).lower())

	def test_so_noi_dia_084_khong_bi_hieu_nham_la_tien_to_quoc_te(self):
		"""BẪY của phép quy đổi tiền tố quốc tế (coordinator nêu đích danh):
		đầu số Vinaphone `084...` là số NỘI ĐỊA hợp lệ (10 số, có 0 đứng
		trước 84) — không được cắt "84" như tiền tố quốc tế. Số đó LUÔN bắt
		đầu bằng "0" nên tự nó đã an toàn với `startswith("84")`; vế thật sự
		cần một ĐỘ DÀI đúng hình quốc tế mới canh được (xem test kế bên)."""
		f = self._upload(_xlsx_bytes([
			_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC", dien_thoai="0842345678"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tao_moi", dong["errors"])
		self.assertEqual(
			dong["dien_thoai"], "0842345678", "không được cắt '84' của số nội địa",
		)

	def test_chuoi_84_sai_do_dai_quoc_te_khong_bi_cat_bua_bai(self):
		"""Vế THẬT SỰ canh điều kiện độ dài của phép quy đổi tiền tố quốc tế
		(`len(tho) == 11`): một chuỗi 12 số bắt đầu bằng "84" KHÔNG đúng hình
		mã nước (2 số) + số thuê bao (9 số) = 11 số — nếu cắt "84" một cách mù
		quáng theo `startswith` một mình (không xét độ dài), "849123456789"
		(12 số) sẽ bị biến thành "09123456789" (11 số) — TRÔNG hợp lệ nhưng là
		một số BỊA, không liên quan gì tới chuỗi gốc. Đúng phải TỪ CHỐI."""
		f = self._upload(_xlsx_bytes([
			_row("Trần Văn Bình", BINH, "Huyết học", "HUYETHOC", dien_thoai="849123456789"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, BINH)
		self.assertEqual(dong["trang_thai"], "tu_choi")
		self.assertNotEqual(dong["dien_thoai"], "09123456789", "không được bịa số")

	def test_tai_khoan_da_co_dien_thoai_rong_thi_duoc_dien(self):
		"""Bài 6a (QĐ-4, vế điền): tài khoản chạy TRƯỚC Task 10 chưa có số —
		nhập lại đúng file có số của người đó phải BỔ SUNG vào chỗ trống,
		không đòi phải sửa tay trên Desk như trước."""
		portal_api.portal_provision(CUST_A, HOA, first_name="Nguyễn Thị Hoa", vai_tro="Quản lý")
		self.assertFalse(frappe.db.get_value("User", HOA, "mobile_no"))

		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý", dien_thoai=SDT_CHUAN_HOA_TEST),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, HOA)
		self.assertEqual(dong["trang_thai"], "bo_qua")
		self.assertIn("bổ sung", dong["ghi_chu"].lower())

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertEqual(frappe.db.get_value("User", HOA, "mobile_no"), SDT_CHUAN_HOA_TEST)

	def test_tai_khoan_da_co_dien_thoai_khac_thi_giu_nguyen_va_canh_bao(self):
		"""Bài 6b (QĐ-4, vế chống mất dữ liệu): `mobile_no` đang có giá trị
		KHÁC là dữ liệu người khác đã nhập — im lặng ghi đè là mất dữ liệu."""
		portal_api.portal_provision(
			CUST_A, HOA, first_name="Nguyễn Thị Hoa", vai_tro="Quản lý", dien_thoai="0911111111",
		)
		self.assertEqual(frappe.db.get_value("User", HOA, "mobile_no"), "0911111111")

		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý", dien_thoai="0922222222"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, HOA)
		self.assertEqual(dong["trang_thai"], "canh_bao")
		self.assertIn("0911111111", " ".join(dong["errors"]))
		self.assertIn("0922222222", " ".join(dong["errors"]))

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertEqual(
			frappe.db.get_value("User", HOA, "mobile_no"), "0911111111",
			"lệch số thì giữ nguyên số cũ — không tự ý ghi đè",
		)


class TestDienThoaiTrung(_NhanSuTestBase):
	"""VÒNG SỬA 2 (04/09/2026) — `User.mobile_no` mang UNIQUE index THẬT trên
	CSDL (tự xác minh `SHOW INDEX FROM tabUser`, `Non_unique = 0`). TRƯỚC vòng
	sửa này, hai dòng trùng số (hoặc một dòng trùng số với `User` đã tồn tại)
	làm nổ `pymysql.err.IntegrityError` GIỮA vòng ghi ở `_ghi()` — và vì luật
	"một dòng bị từ chối là không ghi gì cả" của màn này, MỘT số điện thoại
	trùng chặn cứng việc cấp tài khoản cho CẢ bệnh viện. Ở bệnh viện, hai điều
	dưỡng cùng khoa khai chung số máy bàn của khoa là chuyện bình thường,
	không phải ca hiếm.

	Đúng tinh thần QĐ-1 (chủ đầu tư đã chốt): trùng số KHÔNG được chặn tạo tài
	khoản. Tài khoản vẫn tạo, chỉ KHÔNG gán số cho (những) dòng bị trùng."""

	def test_hai_dong_trong_tep_trung_so_dong_dau_giu_so_dong_sau_bo_trong(self):
		"""Hai điều dưỡng cùng khoa khai chung một số bàn — dòng ĐẦU TIÊN
		trong tệp giữ số, dòng SAU bỏ trống + cảnh báo. Lưới quan trọng nhất:
		CẢ HAI tài khoản vẫn được tạo (không phải `tu_choi`, không phải
		`canh_bao` — CANH_BAO trong file này mang nghĩa "không ghi")."""
		mot = f"trung_mot{DOMAIN}"
		hai = f"trung_hai{DOMAIN}"
		so_chung = "0938271001"
		f = self._upload(_xlsx_bytes([
			_row("Người Một", mot, "Huyết học", "HUYETHOC", dien_thoai=so_chung),
			_row("Người Hai", hai, "Huyết học", "HUYETHOC", dien_thoai=so_chung),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)

		dong_mot = self._dong(ket_qua, mot)
		dong_hai = self._dong(ket_qua, hai)
		self.assertEqual(dong_mot["trang_thai"], "tao_moi")
		self.assertEqual(dong_hai["trang_thai"], "tao_moi")
		self.assertEqual(dong_mot["dien_thoai"], so_chung, "dòng đầu trong tệp giữ số")
		self.assertEqual(dong_hai["dien_thoai"], "", "dòng sau bỏ trống, không giữ số trùng")
		self.assertIn("Người Một", dong_hai["ghi_chu"], "ghi chú phải nêu trùng với AI")
		self.assertIn(so_chung, dong_hai["ghi_chu"], "ghi chú phải nêu trùng SỐ NÀO")
		canh_bao = " ".join(ket_qua["canh_bao_toan_tep"])
		self.assertIn(so_chung, canh_bao, "canh_bao_toan_tep phải liệt kê được ca trùng")

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertTrue(frappe.db.exists("User", mot), "trùng số KHÔNG được chặn tạo tài khoản")
		self.assertTrue(frappe.db.exists("User", hai), "trùng số KHÔNG được chặn tạo tài khoản")
		self.assertEqual(frappe.db.get_value("User", mot, "mobile_no"), so_chung)
		self.assertFalse(frappe.db.get_value("User", hai, "mobile_no"))

	def test_trung_voi_user_da_ton_tai_tren_he_thong_khong_gan_so_nhung_van_tao(self):
		"""Số trong tệp đã thuộc một tài khoản KHÁC đang tồn tại (không phải
		người của chính dòng đó) — tài khoản của dòng này vẫn được tạo, chỉ
		không gán số, và ghi chú nêu rõ tài khoản nào đang giữ số đó."""
		chu_cu = f"chu_cu{DOMAIN}"
		portal_api.portal_provision(
			CUST_B, chu_cu, first_name="Chủ Cũ", vai_tro="Quản lý", dien_thoai="0938271002",
		)
		self.assertEqual(frappe.db.get_value("User", chu_cu, "mobile_no"), "0938271002")

		moi = f"nguoi_moi{DOMAIN}"
		f = self._upload(_xlsx_bytes([
			_row("Người Mới", moi, "Huyết học", "HUYETHOC", dien_thoai="0938271002"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, moi)
		self.assertEqual(
			dong["trang_thai"], "tao_moi", "trùng với User khác KHÔNG được chặn tạo tài khoản",
		)
		self.assertEqual(dong["dien_thoai"], "", "không gán số đang thuộc tài khoản khác")
		self.assertIn(chu_cu, dong["ghi_chu"])
		canh_bao = " ".join(ket_qua["canh_bao_toan_tep"])
		self.assertIn(chu_cu, canh_bao)

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertTrue(frappe.db.exists("User", moi), "trùng số KHÔNG được chặn tạo tài khoản")
		self.assertFalse(frappe.db.get_value("User", moi, "mobile_no"))
		self.assertEqual(
			frappe.db.get_value("User", chu_cu, "mobile_no"), "0938271002",
			"tài khoản cũ không bị đụng vào",
		)

	def test_tai_khoan_da_co_duoc_bo_sung_so_trung_thi_khong_gan_khong_vo(self):
		"""Nhánh BỔ SUNG số cho tài khoản ĐÃ CÓ (QĐ-4 vế điền, `trang_thai ==
		bo_qua`) cũng ghi vào CHÍNH `User.mobile_no` — cùng UNIQUE index, cùng
		rủi ro vỡ. Số trong tệp trùng với một User KHÁC đang tồn tại thì
		KHÔNG được bổ sung (tránh vỡ khi ghi), dòng vẫn giữ nguyên `bo_qua`
		như trước (đây là người ĐÃ CÓ tài khoản, không phải tạo mới)."""
		chu_cu = f"chu_cu_bs{DOMAIN}"
		portal_api.portal_provision(
			CUST_B, chu_cu, first_name="Chủ Cũ BS", vai_tro="Quản lý", dien_thoai="0938271005",
		)
		portal_api.portal_provision(CUST_A, HOA, first_name="Nguyễn Thị Hoa", vai_tro="Quản lý")
		self.assertFalse(frappe.db.get_value("User", HOA, "mobile_no"))

		f = self._upload(_xlsx_bytes([
			_row("Nguyễn Thị Hoa", HOA, vai_tro="Quản lý", dien_thoai="0938271005"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, HOA)
		self.assertEqual(dong["trang_thai"], "bo_qua")
		self.assertIn(chu_cu, dong["ghi_chu"])
		# Review độc lập (vòng sửa 2): ghi chú KHÔNG được nói tài khoản này
		# "vẫn được tạo" — HOA đã có tài khoản TỪ TRƯỚC (dòng 847), không có
		# gì được tạo ở đây cả. Nói "tạo" cho một tài khoản đã tồn tại là một
		# câu SAI SỰ THẬT đưa cho Miyano đọc.
		self.assertNotIn("được tạo", dong["ghi_chu"].lower())

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertFalse(
			frappe.db.get_value("User", HOA, "mobile_no"),
			"không bổ sung số đang thuộc tài khoản khác",
		)

	def test_ve_am_khong_trung_thi_khong_ai_bi_gan_co_va_moi_so_deu_duoc_gan(self):
		"""Vế ÂM: tệp không có số nào trùng nhau và không trùng ai đã có —
		không dòng nào bị cảnh báo vì lý do trùng, mọi số đều được gán."""
		a = f"khong_trung_a{DOMAIN}"
		b = f"khong_trung_b{DOMAIN}"
		f = self._upload(_xlsx_bytes([
			_row("Người A", a, "Huyết học", "HUYETHOC", dien_thoai="0938271003"),
			_row("Người B", b, "Huyết học", "HUYETHOC", dien_thoai="0938271004"),
		]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong_a = self._dong(ket_qua, a)
		dong_b = self._dong(ket_qua, b)
		self.assertEqual(dong_a["dien_thoai"], "0938271003")
		self.assertEqual(dong_b["dien_thoai"], "0938271004")
		self.assertNotIn("trùng", dong_a["ghi_chu"].lower())
		self.assertNotIn("trùng", dong_b["ghi_chu"].lower())
		# Tệp này không có Quản lý nên vẫn còn CẢNH BÁO khác (thiếu Quản lý) —
		# vế âm chỉ đòi hỏi không CẢNH BÁO VÌ LÝ DO TRÙNG SỐ, không đòi hỏi
		# `canh_bao_toan_tep` rỗng tuyệt đối.
		canh_bao = " ".join(ket_qua["canh_bao_toan_tep"]).lower()
		self.assertNotIn("trùng", canh_bao, "không có số nào trùng thì không được cảnh báo trùng")

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertEqual(frappe.db.get_value("User", a, "mobile_no"), "0938271003")
		self.assertEqual(frappe.db.get_value("User", b, "mobile_no"), "0938271004")


class TestGhiHongGiuaChung(_NhanSuTestBase):
	"""Vòng sửa 1 — hai bài canh CHÍNH lúc ghi vỡ giữa chừng.

	Trước vòng này không bài nào lái tới đó: bài "tất-cả-hoặc-không" duy nhất
	ném ở TIỀN KIỂM `so_tu_choi`, tức trước khi ghi một chữ, nên điểm lưu
	(savepoint) trong `_ghi` chưa từng được một test nào chạy qua — mà "ghi
	dở dang" lại đúng là kiểu hỏng brief nêu đích danh.

	Cách lái: chặn ở `update_password` — biên NGOÀI CÙNG mà mật khẩu thô đi
	qua. Cho hai người đầu ghi xong rồi ném ở người thứ ba: đúng hình một lần
	ghi 60/100 dòng rồi vỡ.
	"""

	def _chan_o_nguoi_thu_ba(self):
		"""Chặn ở `update_password` — biên ngoài cùng mà mật khẩu thô đi qua.

		HAI chi tiết của cái bẫy này KHÔNG phải chuyện phong cách, chúng quyết
		định bài test đo được cái gì (đã thấy tận mắt ở vòng đỏ đầu tiên):

		* `new=` chứ KHÔNG `side_effect=`. Với `side_effect`, lời gọi đi qua ba
		  khung của `unittest.mock`, mỗi khung giữ `args = (email, 'mật khẩu
		  thô')` — một tuple tên `args`, KHÔNG nằm trong danh sách chặn của
		  Frappe. Bài test khi đó tự bơm mật khẩu thô vào chính cái traceback
		  nó đang soi, và sẽ đỏ vĩnh viễn dù mã sản xuất đã sạch.
		* Tham số tên `pwd` và bản ghi nhận cất trên `self`. `pwd` khớp danh
		  sách chặn theo tên nên khung của hàm giả này tự che; `self` in ra
		  bằng repr của TestCase, không kết xuất thuộc tính. Một biến cục bộ
		  `da_dat = [...]` thì in nguyên văn — đúng lỗi của vòng đỏ đầu.
		"""
		from unittest.mock import patch

		self._mk_da_dat = []

		def gia_lap(user, pwd, *a, **kw):
			self._mk_da_dat.append(pwd)
			if len(self._mk_da_dat) == 3:
				raise frappe.ValidationError("ZZTEST hỏng giữa chừng khi đặt mật khẩu")
			return update_password(user, pwd, *a, **kw)

		return patch.object(nhan_su_api, "update_password", new=gia_lap)

	def test_ghi_hong_giua_chung_khong_de_lai_gi(self):
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		before = self._counts()

		with self._chan_o_nguoi_thu_ba():
			with self.assertRaises(frappe.ValidationError) as cm:
				nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertIn("ZZTEST hỏng giữa chừng", str(cm.exception))

		# Đã đi tới người thứ ba nghĩa là hai người đầu ĐÃ được ghi trước khi
		# vỡ — nếu không, bài này không canh cái nó tưởng đang canh.
		self.assertEqual(len(self._mk_da_dat), 3)
		self.assertEqual(before, self._counts(), "ghi vỡ giữa chừng phải quay lại sạch")

	def test_mat_khau_tho_khong_lo_ra_trong_traceback(self):
		"""QĐ-G19 cấm đích danh việc mật khẩu vào log.

		`frappe.get_traceback(with_context=True)` (đường mà `log_error()` và
		`log_error_snapshot()` dùng khi có lỗi 500) KẾT XUẤT BIẾN CỤC BỘ của
		mọi khung ngăn xếp bằng `repr()`. Bộ khử của Frappe chỉ che theo TÊN
		BIẾN (`re.search` với password|passwd|secret|token|key|pwd) và chỉ che
		KHOÁ của dict — giá trị trong dict in nguyên văn. `tabError Log` lại
		là MyISAM: dòng đó sống sót qua rollback.
		"""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))

		with self._chan_o_nguoi_thu_ba():
			try:
				nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
				self.fail("phải ném ở người thứ ba")
			except frappe.ValidationError:
				vet = frappe.get_traceback(with_context=True)

		self.assertTrue(self._mk_da_dat, "chưa sinh mật khẩu nào thì bài này vô nghĩa")
		self.assertIn("nhan_su.py", vet, "traceback phải có khung của chính hàm ghi")
		for mat_khau in self._mk_da_dat:
			self.assertNotIn(
				mat_khau, vet,
				"mật khẩu thô lọt vào traceback — đường này đi thẳng vào tabError Log",
			)


class TestTepTaiLen(_NhanSuTestBase):
	def test_tep_cong_khai_bi_tu_choi(self):
		"""Tệp mang họ tên + email nhân viên bệnh viện. Một tệp công khai được
		phục vụ từ `/files/` KHÔNG cần đăng nhập — chốt phải nằm ở server, không
		chỉ ở tuỳ chọn của uploader phía JS."""
		f = self._upload_cong_khai(_xlsx_bytes(TEP_HOP_LE))
		with self.assertRaises(frappe.ValidationError) as cm:
			nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		self.assertIn("riêng tư", str(cm.exception))

	def test_tep_bi_xoa_sau_khi_ghi_xong(self):
		"""Ghi xong thì tệp không còn việc gì để làm — để lại là để một danh
		sách nhân sự đầy đủ nằm vĩnh viễn trên đĩa.

		Khẳng định CẢ BYTE TRÊN ĐĨA, không chỉ dòng trong `tabFile`: mối lo ở
		đây là tệp nằm lại trên máy chủ, mà `delete_doc(..., force=True)` là
		đúng cái cờ có thể bỏ qua các bước dọn của `File.on_trash` — xoá được
		dòng mà bỏ lại tệp thì test vẫn xanh trong khi thứ cần dọn vẫn nằm đó."""
		import os

		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		duong_dan = frappe.get_doc("File", f.name).get_full_path()
		self.assertTrue(os.path.exists(duong_dan), "tệp phải có thật trước đã")

		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		self.assertFalse(frappe.db.exists("File", f.name))
		self.assertFalse(os.path.exists(duong_dan), "byte của tệp vẫn còn trên đĩa")

	def test_tep_van_con_khi_ghi_that_bai(self):
		"""Ngược lại: ghi hỏng thì GIỮ tệp — người nhập còn phải sửa và thử lại."""
		f = self._upload(_xlsx_bytes([_row("Trần Văn Bình", BINH)]))  # thiếu khoa
		with self.assertRaises(frappe.ValidationError):
			nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		self.assertTrue(frappe.db.exists("File", f.name))


class TestCacCaCon(_NhanSuTestBase):
	def test_user_permission_cua_khach_khac_cung_bi_canh_bao(self):
		"""Chốt (b) trước vòng này chỉ soi `Portal Member`. Một `User` đã có
		`User Permission` trỏ về bệnh viện khác mà CHƯA có `Portal Member` thì
		lọt qua — và Frappe OR các User Permission cùng doctype lại với nhau,
		nên tài khoản đó nhìn thấy dữ liệu của HAI bệnh viện."""
		frappe.get_doc({
			"doctype": "User", "email": DUNG, "first_name": "Phạm Văn Dũng",
			"user_type": "Website User", "send_welcome_email": 0,
		}).insert(ignore_permissions=True)
		frappe.get_doc({
			"doctype": "User Permission", "user": DUNG,
			"allow": "Customer", "for_value": CUST_A,
		}).insert(ignore_permissions=True)
		self.assertFalse(frappe.db.exists("Portal Member", {"user": DUNG}))

		f = self._upload(_xlsx_bytes([_row("Phạm Văn Dũng", DUNG, "Hồi sức", "HOISUC")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_B, f.file_url)
		dong = self._dong(ket_qua, DUNG)
		self.assertEqual(dong["trang_thai"], "canh_bao")
		self.assertIn(CUST_A, " ".join(dong["errors"]))

		nhan_su_api.nhan_su_import_commit(CUST_B, f.file_url)
		self.assertFalse(frappe.db.exists("Portal Member", {"user": DUNG}))
		self.assertFalse(frappe.db.exists(
			"User Permission", {"user": DUNG, "allow": "Customer", "for_value": CUST_B}
		))

	def test_tai_khoan_bi_vo_hieu_hoa_duoc_bao_ro(self):
		"""Gắn một tài khoản đã tắt vào bệnh viện rồi báo "Đã tạo" mà không
		nói gì là đẩy người ta đi tìm nguyên nhân "sao đăng nhập không được"."""
		u = frappe.get_doc({
			"doctype": "User", "email": DUNG, "first_name": "Phạm Văn Dũng",
			"user_type": "Website User", "send_welcome_email": 0,
		})
		u.insert(ignore_permissions=True)
		frappe.db.set_value("User", DUNG, "enabled", 0)

		f = self._upload(_xlsx_bytes([_row("Phạm Văn Dũng", DUNG, vai_tro="Quản lý")]))
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f.file_url)
		dong = self._dong(ket_qua, DUNG)
		self.assertEqual(dong["trang_thai"], "tao_moi")
		self.assertIn("vô hiệu hoá", dong["ghi_chu"].lower())

	def test_dong_bo_qua_noi_ro_man_nay_khong_sua_duoc_gi(self):
		"""Màn này KHÔNG có đường cập nhật: đổi khoa, bật lại tài khoản đã tắt,
		sửa vai trò gõ nhầm đều phải làm trên bản ghi Portal Member. Câu chữ
		phải nói đúng năng lực thật, đừng đọc như "không có gì phải làm"."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)

		f2 = self._upload(_xlsx_bytes(TEP_HOP_LE), filename="lan_2.xlsx")
		ket_qua = nhan_su_api.nhan_su_import_preview(CUST_A, f2.file_url)
		ghi_chu = self._dong(ket_qua, BINH)["ghi_chu"].lower()
		self.assertIn("không", ghi_chu)
		self.assertIn("khoa", ghi_chu)
		self.assertIn("portal member", ghi_chu)

	def test_khach_cong_khong_tai_duoc_mau(self):
		"""Chốt vai trò trên endpoint tệp mẫu — bản mẫu cũng là dữ liệu của
		Miyano, và ba endpoint phải cùng một chốt."""
		f = self._upload(_xlsx_bytes(TEP_HOP_LE))
		nhan_su_api.nhan_su_import_commit(CUST_A, f.file_url)
		frappe.set_user(HOA)
		with self.assertRaises(frappe.PermissionError):
			nhan_su_api.nhan_su_import_template()
		frappe.set_user("Administrator")


class TestManDeskHienDuocOChonBenhVien(FrappeTestCase):
	"""Ô "Bệnh viện" của màn Desk phải THẬT SỰ nằm trong DOM.

	Không có hạ tầng test JS trong repo này (`package.json` chỉ có vite), nên
	bất biến giao diện được canh bằng lưới Python đọc thẳng file `.js`.

	VÌ SAO CÓ LỚP NÀY — lượt chạy thử toàn tuyến 04/09/2026 phát hiện màn
	"Nhập nhân sự bệnh viện" CHẶN CỨNG 100% người dùng thật, và đã chặn như
	vậy từ commit dựng nó (`9d84345`): Frappe dựng `page_form` đúng một lần
	trong constructor của `Page` rồi prepend vào `main`, còn `render_body()`
	gọi `this.page.main.html(...)` — thay thẳng innerHTML của `main`, gỡ luôn
	`page_form` ra khỏi DOM. Ô nhập vẫn truy cập được từ JS
	(`page.fields_dict.customer`), nên MỌI phép kiểm ở tầng dữ liệu đều xanh;
	chỉ con mắt nhìn vào màn hình mới thấy nó biến mất.

	Đó là lý do bài này canh THỨ TỰ trong file, chứ không canh sự tồn tại của
	một chuỗi: cả `main.html(` lẫn `page_form` đều đã tồn tại suốt thời gian
	màn hình hỏng. Thứ hỏng là cái nào chạy trước.
	"""

	@classmethod
	def setUpClass(cls):
		super().setUpClass()
		from pathlib import Path

		duong_dan = (
			Path(frappe.get_app_path("miyano_portal"))
			/ "miyano_portal"
			/ "page"
			/ "nhap_nhan_su"
			/ "nhap_nhan_su.js"
		)
		cls.ma = duong_dan.read_text(encoding="utf-8")

	def test_page_form_duoc_gan_lai_sau_khi_ve_than_man(self):
		"""`page_form` phải được gắn lại vào `main` SAU lệnh `main.html(...)`.

		Đảo thứ tự hai lệnh này (hoặc bỏ lệnh gắn lại) làm ô "Bệnh viện" biến
		mất khỏi màn hình mà không endpoint nào đỏ — Miyano không cấp được tài
		khoản cho bất kỳ bệnh viện nào, còn Bước 1 trên màn vẫn thản nhiên bảo
		"chọn ở ô trên đầu màn hình".
		"""
		i_html = self.ma.find("this.page.main.html(")
		self.assertNotEqual(
			i_html, -1, "Không tìm thấy lệnh vẽ thân màn `this.page.main.html(`"
		)
		i_gan_lai = self.ma.find("this.page.page_form.prependTo(this.page.main)")
		self.assertNotEqual(
			i_gan_lai,
			-1,
			"Thiếu lệnh gắn `page_form` trở lại `main`. `main.html()` gỡ nó ra "
			"khỏi DOM, nên không có lệnh này thì ô 'Bệnh viện' không bao giờ "
			"hiện lên và màn Desk cấp tài khoản chặn cứng người dùng.",
		)
		self.assertLess(
			i_html,
			i_gan_lai,
			"`page_form` đang được gắn lại TRƯỚC `main.html(...)` — lệnh html() "
			"chạy sau sẽ gỡ nó ra lần nữa. Phải gắn lại SAU khi vẽ xong thân màn.",
		)
