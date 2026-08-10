"""Import/export bảng dòng của phiếu nhập và phiếu xuất."""

import io

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import Workbook, load_workbook

from miyano_portal.kho import dong_phieu
from miyano_portal.setup.seed_kho_demo import seed_kho_demo


def _xlsx(loai, rows):
	wb = Workbook()
	ws = wb.active
	ws.append([label for label, _ in dong_phieu.COLUMNS[loai]])
	for r in rows:
		ws.append(r)
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


class TestDocFileNhap(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		self.kho_bm = self.kho["kho_bm"]

	def test_ma_da_co_thi_trang_thai_khop_va_gan_san_vat_tu(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["MYN-GLOVE-M", "", "", "LO-1", None, 10, 1000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["trang_thai"], "khop")
		self.assertEqual(row["vat_tu"], self.kho["vt_bm"])

	def test_ma_la_thi_trang_thai_ma_moi(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["BM-LA-01", "Vật tư lạ", "Cái", "LO-2", None, 5, 2000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["trang_thai"], "ma_moi")
		self.assertEqual(row["vat_tu"], "")
		self.assertEqual(row["ten_vat_tu"], "Vật tư lạ")

	def test_so_luong_sai_thi_trang_thai_loi_neu_dung_so_dong(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["MYN-GLOVE-M", "", "", "LO-1", None, "abc", 1000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["trang_thai"], "loi")
		self.assertEqual(row["line"], 2)  # header ở dòng 1
		self.assertIn("Số lượng", " ".join(row["loi"]))

	def test_dong_loi_co_ma_khop_vat_tu_that_thi_van_khong_lo_vat_tu(self):
		# Khoá lỗ hổng round 2: bỏ chặn `not loi` (round 1) khiến _match_vat_tu
		# chạy cả trên dòng đã có lỗi định dạng khác. Nếu mã khớp một Customer
		# Warehouse Item thật, `vat_tu_name` bị gán rồi lọt ra ngoài dù trạng
		# thái cuối là "loi" — một consumer rẽ nhánh theo "vat_tu có giá trị
		# hay không" thay vì theo trang_thai sẽ âm thầm coi dòng lỗi là khớp.
		# Bất biến bắt buộc: dòng "loi" không bao giờ mang định danh vat_tu thật.
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["MYN-GLOVE-M", "", "", "LO-1", None, "abc", 1000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["trang_thai"], "loi")
		self.assertEqual(row["vat_tu"], "")

	def test_ma_moi_thieu_ten_hoac_dvt_thi_loi(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["BM-LA-02", "", "", "LO-3", None, 5, 2000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		self.assertEqual(kq["rows"][0]["trang_thai"], "loi")

	def test_ma_da_co_thi_bo_qua_ten_va_dvt_trong_file(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["MYN-GLOVE-M", "TÊN SAI", "ĐVT SAI", "LO-1", None, 10, 1000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["ten_vat_tu"], "Găng tay y tế size M")
		self.assertNotEqual(row["dvt"], "ĐVT SAI")

	def test_so_lo_trong_thi_nhan_lo_mac_dinh(self):
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["MYN-GLOVE-M", "", "", "", None, 10, 1000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		self.assertEqual(kq["rows"][0]["so_lo"], "KHONG-LO")

	def test_dong_vua_sai_so_luong_vua_ma_moi_thieu_ten_dvt_thi_gom_du_loi(self):
		# Khoá lỗ hổng round 1: trước sửa, khối kiểm mã mới (thiếu Tên/ĐVT) chỉ
		# chạy khi `not loi`, nên lỗi Số lượng đứng trước sẽ nuốt mất hai lý do
		# kia. Dòng này cố tình phạm CẢ BA lỗi cùng lúc: mã "BM-NEW-01" chưa có
		# trong kho (không phải "MYN-*" nên cũng không khớp Item Miyano), Số
		# lượng "abc" không phải số, và Tên/ĐVT đều bỏ trống.
		kq = dong_phieu.doc_file(
			_xlsx("nhap", [["BM-NEW-01", "", "", "", None, "abc", 2000, "", "", ""]]),
			self.kho_bm, "nhap",
		)
		row = kq["rows"][0]
		self.assertEqual(row["trang_thai"], "loi")
		loi_gop = " ".join(row["loi"])
		self.assertIn("Số lượng", loi_gop)
		self.assertIn("Tên vật tư", loi_gop)
		self.assertIn("ĐVT", loi_gop)
		self.assertEqual(len(row["loi"]), 3)


class TestDocFileXuat(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		self.kho_bm = self.kho["kho_bm"]

	def test_file_xuat_khong_co_cot_don_gia(self):
		labels = [label for label, _ in dong_phieu.COLUMNS["xuat"]]
		self.assertNotIn("Đơn giá", labels)
		self.assertNotIn("Hạn sử dụng", labels)

	def test_file_xuat_co_quy_cach_va_nhom(self):
		# Chốt hợp đồng: dòng xuất VẪN mang Quy cách/Nhóm dù không có giá/hạn —
		# modal tạo nhanh cho mã mới cần hai trường này. Khoá lại để không ai vô
		# tình bỏ chúng khỏi COLUMNS["xuat"] sau này.
		labels = [label for label, _ in dong_phieu.COLUMNS["xuat"]]
		self.assertIn("Quy cách", labels)
		self.assertIn("Nhóm", labels)

	def test_doc_file_xuat_khong_tra_don_gia(self):
		kq = dong_phieu.doc_file(
			_xlsx("xuat", [["MYN-GLOVE-M", "", "", "LO-1", 3, "", "", ""]]),
			self.kho_bm, "xuat",
		)
		self.assertNotIn("don_gia", kq["rows"][0])

	def test_loai_la_bi_chan(self):
		with self.assertRaises(frappe.ValidationError):
			dong_phieu.doc_file(b"", self.kho_bm, "linh tinh")


class TestExportDong(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		self.phieu = frappe.get_doc({
			"doctype": "Customer Stock Receipt",
			"kho": self.kho["kho_bm"],
			"ngay": frappe.utils.today(),
			"loai_nhap": "Nhập khác",
			"items": [{
				"vat_tu": self.kho["vt_bm"], "so_lo": "LO-EXP-01",
				"so_luong": 7, "don_gia": 1234,
			}],
		})
		self.phieu.insert(ignore_permissions=True)

	def test_export_ra_dung_bo_cot_va_du_lieu(self):
		content = dong_phieu.build_export_xlsx("Customer Stock Receipt", self.phieu.name)
		ws = load_workbook(io.BytesIO(content), data_only=True).active
		self.assertEqual([c.value for c in ws[1]], [label for label, _ in dong_phieu.COLUMNS["nhap"]])
		self.assertEqual(ws.cell(row=2, column=1).value, "MYN-GLOVE-M")
		self.assertEqual(ws.cell(row=2, column=6).value, 7)

	def test_export_roi_nap_lai_ra_dong_khop(self):
		content = dong_phieu.build_export_xlsx("Customer Stock Receipt", self.phieu.name)
		kq = dong_phieu.doc_file(content, self.kho["kho_bm"], "nhap")
		self.assertEqual(kq["rows"][0]["trang_thai"], "khop")
		self.assertEqual(kq["rows"][0]["so_luong"], 7)


from miyano_portal.api import kho as kho_api

BM_USER = "bvbm@demo.miyano"
PXN_USER = "pxnabc@demo.miyano"


class TestDongPhieuEndpoint(FrappeTestCase):
	def setUp(self):
		self.kho = seed_kho_demo()
		self.phieu_bm = frappe.get_doc({
			"doctype": "Customer Stock Receipt",
			"kho": self.kho["kho_bm"],
			"ngay": frappe.utils.today(),
			"loai_nhap": "Nhập khác",
			"items": [{
				"vat_tu": self.kho["vt_bm"], "so_lo": "LO-EP-01",
				"so_luong": 2, "don_gia": 100,
			}],
		})
		self.phieu_bm.insert(ignore_permissions=True)
		frappe.set_user(BM_USER)

	def tearDown(self):
		frappe.set_user("Administrator")

	def test_mau_tra_ve_file(self):
		kho_api.kho_dong_phieu_mau("nhap")
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent)

	def test_export_phieu_cua_kho_khac_bi_chan(self):
		frappe.set_user(PXN_USER)
		with self.assertRaises(frappe.PermissionError):
			kho_api.kho_dong_phieu_export("Customer Stock Receipt", self.phieu_bm.name)

	def test_export_doctype_ngoai_danh_sach_trang_bi_chan(self):
		with self.assertRaises(frappe.ValidationError):
			kho_api.kho_dong_phieu_export("Sales Invoice", self.phieu_bm.name)

	def test_luu_phieu_co_dong_thieu_vat_tu_bi_chan_o_server(self):
		with self.assertRaises(frappe.ValidationError) as ctx:
			kho_api.kho_phieu_nhap_save({
				"ngay": frappe.utils.today(),
				"loai_nhap": "Nhập khác",
				"items": [{"vat_tu": "", "so_lo": "LO-X", "so_luong": 1, "don_gia": 100}],
			})
		self.assertIn("chưa chọn vật tư", str(ctx.exception))
