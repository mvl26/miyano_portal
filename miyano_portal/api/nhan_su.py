"""Nhập nhân sự bệnh viện từ tệp Excel — tạo tài khoản, khoa phòng, phân quyền.

Task 15 (QĐ-G17…G23). Khuôn BA BƯỚC lấy nguyên của `api/kho.py`
(`kho_import_template` → `kho_import_preview` → `kho_import_commit`), chỉ khác
chỗ đứng: màn này nằm trong **Desk của Miyano**, không phải SPA của khách.
Không đẻ khuôn thứ ba.

Bốn điều đáng nhớ trước khi sửa file này:

1. **Khách hàng CHỌN TRÊN MÀN HÌNH, không nằm trong tệp** (QĐ-G18). Bộ cột cố
   ý KHÔNG có "tên bệnh viện": một cột gõ tay là đường để nhập nhầm người của
   viện này sang viện khác — đúng thứ cơ chế cách ly dữ liệu sinh ra để chặn.
2. **Bước xem trước là chốt chính, không phải trang trí.** Tạo tài khoản đăng
   nhập là việc khó lùi. `_phan_tich()` KHÔNG GHI GÌ, và nó phải đoán trước
   TẤT CẢ các chốt mà `PortalMember.validate()`/`CustomerDepartment.validate()`
   sẽ nổ lúc ghi — một lần "xem trước bảo được, commit nổ giữa chừng" là đúng
   kiểu hỏng mà task này sinh ra để dẹp.
3. **Tất-cả-hoặc-không cho dòng BỊ TỪ CHỐI** (cùng khuôn `import_ton_dau`),
   nhưng dòng *cảnh báo* (QĐ-G23) thì không chặn ai — nó chỉ báo và để Miyano
   quyết.
4. **Mật khẩu trả về ĐÚNG MỘT LẦN** trong kết quả của `commit` (QĐ-G19), không
   ghi vào tệp, không gửi email, không ghi vào Comment. Việc giữ nó ra khỏi
   `tabError Log` KHÔNG phải chuyện hiển nhiên và không phải chuyện đặt tên
   biến — đọc `_MatKhauTho` bên dưới trước khi đụng vào đường ghi.
"""

import io
import re
import secrets
import unicodedata

import frappe
from frappe.utils.password import update_password
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Dùng lại NGUYÊN phép kiểm sở hữu tệp của đường import kho (`api/kho.py`):
# tra File bằng `file_url`, so `owner` với người đang gọi, ép đuôi .xlsx, và
# trả mọi lỗi bằng tiếng Việt không lộ tên doctype. Một bản sao thứ hai của
# phép kiểm đó là một chỗ nữa để quên vá.
from miyano_portal.api.kho import _resolve_owned_spreadsheet
from miyano_portal.api.portal import chan_neu_khong_phai_nhan_vien_miyano, portal_provision
from miyano_portal.kho import import_ton_dau, similarity
from miyano_portal.miyano_portal.doctype.portal_member.portal_member import (
	NHAN_VIEN_KHOA,
	QUAN_LY,
)

# (nhãn cột hiển thị, tên field nội bộ) — MỘT bộ cột duy nhất cho cả ba việc:
# sinh tệp mẫu, đọc tệp xem trước, đọc tệp lúc ghi.
#
# Task 10: "Số điện thoại" thêm vào ĐÂY, không phải một tệp/luồng riêng — dòng
# thời gian "Ai đã làm gì" (Task 1-8) hiện tên NGƯỜI THAO TÁC kèm số điện
# thoại để khách bấm gọi thẳng; tài khoản không có số thì dòng thời gian chỉ
# hiện tên, một màn hình nêu câu hỏi "hỏi ai" rồi không cho gọi. Trước task
# này số phải điền tay từng tài khoản trên Desk sau khi nhập — bước thủ công
# chắc chắn có ngày bị quên.
COLUMNS = [
	("Họ tên", "ho_ten"),
	("Email", "email"),
	("Khoa", "ten_khoa"),
	("Mã khoa", "ma_khoa"),
	("Vai trò", "vai_tro"),
	("Số điện thoại", "dien_thoai"),
]

# Cột bắt buộc phải CÓ MẶT trong header. "Khoa"/"Mã khoa" không bắt buộc có
# mặt vì một tệp toàn Quản lý là hợp lệ (Quản lý nhìn toàn viện).
#
# "dien_thoai" CÓ MẶT ở đây (ai còn dùng tệp mẫu 5 cột cũ bị báo ngay ở bước
# đọc header, không lặng lẽ tạo cả viện không số) nhưng Ô TRỐNG của nó KHÔNG
# bị coi là lỗi hình thức như "ho_ten"/"email" — xem nhánh CANH_BAO trong
# `_phan_tich()`. "Bắt buộc có mặt cột" và "bắt buộc có giá trị mỗi dòng" là
# hai luật khác nhau, cố ý.
REQUIRED_HEADER = {"ho_ten", "email", "vai_tro", "dien_thoai"}

# QĐ-2 (Task 10): số SAI ĐỊNH DẠNG bị TỪ CHỐI — khác hẳn Ô TRỐNG (chỉ cảnh
# báo). Bỏ trống là một lựa chọn có ý thức của người điền; gõ sai là một lỗi
# im lặng sẽ in ra trước mặt bệnh viện và dẫn người ta gọi nhầm số. Số di
# động Việt Nam: bắt đầu bằng 0, 10-11 chữ số.
#
# VÒNG SỬA 2 (coordinator, 04/09/2026): regex này áp lên chuỗi ĐÃ CHUẨN HOÁ
# (`_chuan_hoa_dien_thoai()`), không phải chuỗi thô — trước vòng sửa này nó
# áp thẳng lên chuỗi thô, nên "+84912345678"/"091 234 5678"/"0912.345.678"/
# "091-234-5678" (bốn cách viết hợp lệ của CÙNG một số) đều bị TỪ CHỐI, và
# TỪ CHỐI kéo theo "một dòng bị từ chối là không ghi gì cả" — chặn cứng cả
# tệp, đúng cái bẫy QĐ-1 dựng ra để tránh, chỉ dịch sang QĐ-2.
_RE_DIEN_THOAI = re.compile(r"^0\d{9,10}$")

VAI_TRO_HOP_LE = (QUAN_LY, NHAN_VIEN_KHOA)

# Trạng thái của một dòng ở bước xem trước — đúng bốn thứ mà người ngồi duyệt
# tệp cần phân biệt.
TAO_MOI = "tao_moi"
BO_QUA = "bo_qua"
CANH_BAO = "canh_bao"
TU_CHOI = "tu_choi"


def _norm(value) -> str:
	if value is None:
		return ""
	return unicodedata.normalize("NFC", str(value)).strip()


def _chuan_hoa_dien_thoai(value) -> str:
	"""Chuẩn hoá ô Số điện thoại — CHUẨN HOÁ TRƯỚC, KIỂM SAU (VÒNG SỬA 2).

	Hai cạm bẫy khác nhau, xử ở đây theo đúng thứ tự vì cái thứ hai chỉ lộ ra
	SAU khi cái thứ nhất được vá đúng cách:

	1. **CẠM BẪY EXCEL (QĐ-3)**: một ô định dạng SỐ (không phải văn bản) làm
	   Excel NUỐT MẤT số 0 đứng đầu — `0912345678` gõ vào ô số lưu thành số
	   `912345678`, và `openpyxl` trả `int`/`float` đó nguyên văn. Ô kiểu số ở
	   đây luôn được hiểu là "0 bị nuốt", trả về NGAY sau khi phục hồi — số
	   này không mang dấu phân cách hay tiền tố quốc tế, không cần đi qua các
	   bước bên dưới (giữ NGUYÊN hành vi đã có, không đụng).
	2. **BỐN CÁCH VIẾT HỢP LỆ CÙNG MỘT SỐ (VÒNG SỬA 2)**: người gõ tay số
	   điện thoại rất hay chèn khoảng trắng/dấu chấm/gạch ngang/ngoặc đơn
	   (`091 234 5678`, `0912.345.678`, `091-234-5678`) hoặc viết tiền tố
	   quốc tế (`+84912345678`, `84912345678`) — lột hết các ký tự phân cách,
	   rồi quy tiền tố quốc tế về dạng nội địa (`0…`), TRƯỚC KHI kiểm định
	   dạng ở `_dien_thoai_hop_le()`. Không làm vậy thì bốn cách viết đúng đó
	   bị TỪ CHỐI oan, và TỪ CHỐI kéo theo "một dòng bị từ chối là không ghi
	   gì cả" — chặn cứng cả tệp, đúng cái bẫy QĐ-1 dựng ra để tránh.

	BẪY của bước 2: một số NỘI ĐỊA hợp lệ bắt đầu bằng `084…` (đầu số
	Vinaphone, đúng 10 chữ số) TUYỆT ĐỐI không được hiểu nhầm là tiền tố quốc
	gia — số đó LUÔN có `0` đứng trước `84` nên `startswith("84")` (ký tự đầu
	phải là `8`) không bao giờ khớp nó, và vì vậy không rơi vào nhánh quy đổi
	bên dưới. Chỉ quy đổi khi TIN CHẮC đó là tiền tố quốc tế: có dấu `+`
	(tín hiệu không mơ hồ), hoặc không có `+` nhưng tổng độ dài khớp ĐÚNG hình
	quốc tế (mã nước 2 số + 9 số thuê bao = 11 số, không có `0` đứng trước).
	"""
	if isinstance(value, bool):
		return _norm(value)
	if isinstance(value, (int, float)):
		return "0" + str(int(value))

	tho = _norm(value)
	if not tho:
		return tho

	# Lột khoảng trắng, dấu chấm, gạch ngang, ngoặc đơn — CHỈ những ký tự
	# phân cách, KHÔNG lột dấu "+": "+" là tín hiệu duy nhất chắc chắn báo
	# tiền tố quốc tế, cần giữ lại để phân biệt với số nội địa `084…`.
	tho = re.sub(r"[\s.\-()]", "", tho)

	if tho.startswith("+84"):
		tho = "0" + tho[3:]
	elif tho.startswith("84") and len(tho) == 11 and tho.isdigit():
		tho = "0" + tho[2:]
	return tho


def _dien_thoai_hop_le(so: str) -> bool:
	return bool(_RE_DIEN_THOAI.fullmatch(so))


def _kiem_trung_dien_thoai(
	dong: dict, dien_thoai: str, email_dong: str, so_da_dung: dict[str, dict],
	canh_bao_toan_tep: list[str], dang_tao_moi: bool,
) -> str:
	"""VÒNG SỬA 2 — `User.mobile_no` mang UNIQUE index THẬT trên CSDL (tự xác
	minh `SHOW INDEX FROM tabUser`, `Non_unique = 0`). Không có phép kiểm này,
	hai người trùng số nổ `pymysql.err.IntegrityError` GIỮA vòng ghi ở
	`_ghi()`, và vì luật "một dòng bị từ chối là không ghi gì cả" của màn
	này, MỘT số trùng chặn cứng việc cấp tài khoản cho CẢ bệnh viện. Ở bệnh
	viện, hai điều dưỡng cùng khoa khai chung số máy bàn của khoa là chuyện
	bình thường — không phải ca hiếm cần lường.

	Đúng tinh thần QĐ-1 (chủ đầu tư đã chốt, áp dụng tiếp cho ca này): trùng
	số KHÔNG được chặn tạo tài khoản. Trả về CHUỖI RỖNG nếu số này bị trùng
	(gọi nơi dùng PHẢI gán đè lại `dong["dien_thoai"]` bằng giá trị trả về —
	đây là nguồn sự thật DUY NHẤT mà `_ghi()` đọc lại lúc quyết có ghi số hay
	không), ngược lại trả nguyên `dien_thoai` và GHI NHẬN dòng này đã "giữ"
	số đó để các dòng sau trong tệp so.

	Gọi được từ HAI nhánh khác nhau của `_phan_tich()` (tạo mới VÀ bổ sung số
	cho tài khoản đã có) vì cả hai đều ghi vào CHÍNH một cột `User.mobile_no`
	lúc `_ghi()` chạy — cùng một cột, cùng một rủi ro vỡ. `dang_tao_moi` PHÂN
	BIỆT ngữ cảnh gọi (review độc lập, vòng sửa 2 — bản đầu dùng CHUNG một câu
	"tài khoản của dòng này vẫn được tạo" cho cả hai nhánh; câu đó ĐÚNG cho
	nhánh tạo mới nhưng SAI SỰ THẬT cho nhánh bổ sung — tài khoản ở đó ĐÃ CÓ
	từ trước, không có gì được "tạo" cả, chỉ là số không được bổ sung).
	"""
	# (1) trùng với một User ĐÃ CÓ trên hệ thống, không phải người của CHÍNH
	# dòng này (loại trừ email của dòng: tài khoản đã tồn tại giữ số của
	# MÌNH không phải một ca trùng).
	chu_cu = frappe.db.get_value(
		"User", {"mobile_no": dien_thoai, "name": ["!=", email_dong]}, "name"
	)
	if chu_cu:
		if dang_tao_moi:
			ket_cuc = "tài khoản của dòng này vẫn được tạo."
		else:
			ket_cuc = "tài khoản ĐÃ CÓ của dòng này không bị ảnh hưởng gì khác."
		_them_ghi_chu(dong, (
			f'Số điện thoại "{dien_thoai}" đã thuộc tài khoản khác đang tồn tại trên hệ '
			f'thống ("{chu_cu}") — KHÔNG gán số này để tránh trùng, {ket_cuc}'
		))
		canh_bao_toan_tep.append(
			f'Dòng {dong["line"]} ({dong["ho_ten"]}): số "{dien_thoai}" đã thuộc tài khoản '
			f'"{chu_cu}" đang tồn tại — không gán số này.'
		)
		dong["_trung_dien_thoai"] = True
		return ""
	# (2) trùng với một dòng KHÁC đã đi qua TRƯỚC nó trong CHÍNH tệp này —
	# dòng ĐẦU TIÊN trong tệp giữ số, các dòng sau bỏ trống (người điền tệp
	# liệt kê ai trước thường là chủ số thật, các dòng sau nhiều khả năng là
	# copy-paste nhầm số của khoa/phòng).
	trung = so_da_dung.get(dien_thoai)
	if trung:
		_them_ghi_chu(dong, (
			f'Số điện thoại "{dien_thoai}" trùng với dòng {trung["line"]} '
			f'({trung["ho_ten"]}) trong cùng tệp — dòng đó giữ số, dòng này bỏ trống để '
			"tránh trùng."
		))
		canh_bao_toan_tep.append(
			f'Dòng {dong["line"]} ({dong["ho_ten"]}) và dòng {trung["line"]} '
			f'({trung["ho_ten"]}) cùng khai số "{dien_thoai}" trong tệp — chỉ dòng '
			f'{trung["line"]} được gán số, (các) dòng sau bỏ trống.'
		)
		dong["_trung_dien_thoai"] = True
		return ""
	so_da_dung[dien_thoai] = dong
	return dien_thoai


# ---------------------------------------------------------------------------
# Bước 1 — tệp mẫu
# ---------------------------------------------------------------------------


def build_template_bytes() -> bytes:
	"""Tệp mẫu .xlsx: đúng 6 cột theo thứ tự COLUMNS, kèm hai dòng ví dụ (một
	cho mỗi vai trò — dòng Quản lý CỐ Ý để trống Khoa/Mã khoa để người điền
	thấy ngay rằng đó là hợp lệ).

	Cả hai dòng ví dụ ĐỀU có số điện thoại (Task 10, QĐ-1): cột này bắt buộc
	có mặt trong header, nên tệp mẫu phải làm gương — bỏ trống ví dụ sẽ dạy
	người điền rằng để trống là bình thường.

	Tải mẫu xuống rồi nạp lại ngay (không sửa gì) phải đi lọt bước xem trước
	KHÔNG một cảnh báo nào — cùng cam kết như tệp mẫu của kho, xem
	`import_ton_dau.build_template_bytes`.
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = "Nhân sự"
	ws.append([label for label, _ in COLUMNS])
	for cell in ws[1]:
		cell.font = Font(bold=True)
	ws.append(["Nguyễn Thị Hoa", "quanly@benhvien.example", "", "", QUAN_LY, "0912345678"])
	ws.append([
		"Trần Văn Bình", "huyethoc@benhvien.example", "Huyết học", "HUYETHOC", NHAN_VIEN_KHOA,
		"0987654321",
	])
	for i, width in enumerate([26, 32, 24, 14, 18, 18], start=1):
		ws.column_dimensions[get_column_letter(i)].width = width
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


# ---------------------------------------------------------------------------
# Bước 2 — đọc và phân tích, KHÔNG GHI GÌ
# ---------------------------------------------------------------------------


def _kiem_khach(customer) -> str:
	customer = _norm(customer)
	if not customer or not frappe.db.exists("Customer", customer):
		frappe.throw(
			"Không tìm thấy khách hàng để nhập nhân sự. Chọn lại bệnh viện trên màn hình.",
			frappe.ValidationError,
		)
	return customer


def _them_ghi_chu(dong: dict, text: str) -> None:
	"""Ghi chú CỘNG DỒN. Một dòng có thể mang nhiều lời nhắc cùng lúc (tài
	khoản đang tắt + khoa đang tắt); gán đè sẽ nuốt mất cái trước."""
	dong["ghi_chu"] = f'{dong["ghi_chu"]} · {text}'.strip(" ·") if dong["ghi_chu"] else text


def _kiem_tep_rieng_tu(file_url) -> None:
	"""Tệp nhân sự mang họ tên + email của nhân viên bệnh viện.

	`_resolve_owned_spreadsheet` chỉ kiểm `owner`, KHÔNG kiểm `is_private` — mà
	một tệp công khai được phục vụ thẳng từ `/files/` cho bất kỳ ai có đường
	dẫn, không cần đăng nhập. Tuỳ chọn `make_attachments_public: false` phía JS
	là lớp tiện lợi, không phải chốt: chốt phải đứng ở server, nơi client không
	đổi được.
	"""
	if not file_url:
		return  # `_resolve_owned_spreadsheet` lo phần thiếu tệp, với thông điệp của nó
	if frappe.db.get_value("File", {"file_url": file_url}, "is_private") == 0:
		frappe.throw(
			"Tệp nhân sự phải ở chế độ riêng tư (Private). Tệp này đang công khai — "
			"bất kỳ ai có đường dẫn đều tải được danh sách nhân sự. Xoá tệp đi và "
			"tải lại bằng nút trên màn hình này.",
			frappe.ValidationError,
		)


def _kiem_ma_khoa(ma: str) -> str | None:
	"""Soi trước ĐÚNG các luật của `CustomerDepartment._chuan_hoa_ma_khoa()`.

	Cố ý là một BẢN SOI, không phải bản sao thứ hai của luật: nguồn sự thật
	vẫn là controller (nó vẫn chạy lúc `insert`). Chỗ này chỉ để bước xem
	trước nói ra được lỗi TRƯỚC khi ai đó bấm Ghi — nếu controller đổi luật
	mà quên chỗ này, hậu quả là commit nổ và savepoint quay lại, không phải
	dữ liệu sai lọt xuống.
	"""
	if len(ma) > 20:
		return "Mã khoa không được quá 20 ký tự."
	if not ma.isalnum() or not ma.isascii():
		return "Mã khoa chỉ được dùng chữ cái không dấu và chữ số (ví dụ HUYETHOC)."
	if ma in {"CHUNG"}:
		return f'"{ma}" là mã dành riêng của hệ thống, không đặt cho khoa phòng được.'
	return None


def _tim_khoa_da_co(khoa_hien_co: list, ten: str, ma: str) -> tuple[dict | None, str | None]:
	"""Khớp một ô Khoa/Mã khoa trong tệp với khoa phòng ĐÃ CÓ của bệnh viện.

	Mã khoa có tiếng nói cuối cùng khi cả hai cùng có: nó là định danh, tên là
	chữ tự do. So tên không dấu, cùng phép so `_chan_trung_tuyet_doi()` của
	`Customer Department` dùng — nếu so lệch nhau, tệp sẽ "tạo khoa mới" ở xem
	trước rồi vỡ vì trùng tên lúc ghi.
	"""
	if ma:
		for row in khoa_hien_co:
			if (row.get("ma_khoa") or "").upper() == ma:
				return row, None
	if ten:
		for row in khoa_hien_co:
			if similarity.la_trung_tuyet_doi(ten, row.get("ten_khoa_phong")):
				ma_cu = (row.get("ma_khoa") or "").upper()
				if ma and ma_cu and ma_cu != ma:
					return None, (
						f'Khoa "{row["ten_khoa_phong"]}" của bệnh viện này đang mang mã '
						f'"{ma_cu}", không khớp Mã khoa "{ma}" trong tệp.'
					)
				return row, None
	return None, None


def _phan_tich(content: bytes, customer: str) -> dict:
	"""Đọc và kiểm toàn bộ tệp, KHÔNG GHI GÌ vào database.

	Trả verdict theo khuôn `previewing-imports-before-writing`: MỖI dòng của
	tệp đều có mặt trong `rows` kèm số dòng gốc (1-based, tính cả header) và
	một trạng thái nói rõ *sẽ tạo / bỏ qua vì đã có / chỉ cảnh báo / bị từ
	chối vì lý do gì* — không dòng nào biến mất lặng lẽ.
	"""
	ws = import_ton_dau.mo_workbook(content)
	header_row, col_index = import_ton_dau.read_header(ws, COLUMNS, REQUIRED_HEADER)

	khoa_hien_co = frappe.get_all(
		"Customer Department",
		filters={"customer": customer},
		fields=["name", "ten_khoa_phong", "ma_khoa", "active"],
	)
	quan_ly_hien_co = frappe.db.get_value(
		"Portal Member", {"customer": customer, "vai_tro": QUAN_LY, "active": 1}, "user"
	)
	co_ma_ngan = bool(frappe.db.get_value("Customer", customer, "custom_ma_ngan"))

	rows: list[dict] = []
	khoa_se_tao: list[dict] = []
	email_da_gap: dict[str, int] = {}
	quan_ly_trong_tep: dict | None = None
	loi_toan_tep: list[str] = []
	canh_bao_toan_tep: list[str] = []
	# VÒNG SỬA 2: dòng ĐẦU TIÊN trong tệp "giữ" một số điện thoại — theo dõi ở
	# đây (dùng chung cho cả nhánh TẠO MỚI và nhánh BỔ SUNG số cho tài khoản
	# đã có, xem `_kiem_trung_dien_thoai`) để dòng SAU trùng số biết mà bỏ
	# trống, không phải đợi CSDL ném `IntegrityError` giữa vòng ghi.
	so_da_dung: dict[str, dict] = {}

	for line, row_cells in enumerate(
		ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1
	):
		raw = {field: _cell(row_cells, col) for field, col in col_index.items()}
		if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw.values()):
			continue  # dòng trắng hoàn toàn — bỏ qua, không tính vào tổng

		ho_ten = _norm(raw.get("ho_ten"))
		email = _norm(raw.get("email")).lower()
		ten_khoa = _norm(raw.get("ten_khoa"))
		ma_khoa = _norm(raw.get("ma_khoa")).upper()
		vai_tro_raw = _norm(raw.get("vai_tro"))
		dien_thoai = _chuan_hoa_dien_thoai(raw.get("dien_thoai"))

		dong = {
			"line": line, "ho_ten": ho_ten, "email": email, "ten_khoa": ten_khoa,
			"ma_khoa": ma_khoa, "vai_tro": "", "khoa": None, "khoa_moi": False,
			"khoa_key": None, "trang_thai": TAO_MOI, "errors": [], "ghi_chu": "",
			"dien_thoai": dien_thoai,
		}
		rows.append(dong)
		errors = dong["errors"]

		# --- (a) chốt hình thức -------------------------------------------
		if not ho_ten:
			errors.append("Thiếu Họ tên")
		if not email:
			errors.append("Thiếu Email")
		elif not frappe.utils.validate_email_address(email):
			errors.append(f'Email không hợp lệ: "{email}"')
		elif email in email_da_gap:
			errors.append(f"Email này đã có ở dòng {email_da_gap[email]} trong tệp")
		elif email:
			email_da_gap[email] = line
		if dien_thoai and not _dien_thoai_hop_le(dien_thoai):
			# QĐ-2: SAI định dạng bị TỪ CHỐI — khác Ô TRỐNG (xem nhánh CANH_BAO
			# ở cuối vòng lặp). Gõ sai là lỗi im lặng sẽ in ra trước mặt bệnh
			# viện và dẫn người ta gọi nhầm số.
			errors.append(
				f'Số điện thoại không hợp lệ: "{dien_thoai}" — chỉ nhận số bắt đầu '
				"bằng 0, gồm 10-11 chữ số."
			)

		vai_tro = next(
			(v for v in VAI_TRO_HOP_LE if similarity.la_trung_tuyet_doi(vai_tro_raw, v)), None
		)
		if not vai_tro:
			errors.append(
				f'Vai trò không hợp lệ: "{vai_tro_raw}" — chỉ nhận "{QUAN_LY}" hoặc '
				f'"{NHAN_VIEN_KHOA}"'
			)
		dong["vai_tro"] = vai_tro or vai_tro_raw
		if errors:
			dong["trang_thai"] = TU_CHOI
			continue

		# --- (b) email này đã là ai chưa? ----------------------------------
		# Gõ nhầm email NỘI BỘ Miyano vào tệp của bệnh viện là một tai nạn có
		# bán kính rộng: `portal_provision` sẽ gắn `User Permission` trên
		# Customer cho tài khoản đó, và một System User bị giới hạn theo một
		# khách hàng sẽ mất tầm nhìn ở khắp ERPNext — hỏng theo kiểu rất khó
		# lần ra nguyên nhân. Chặn ở đây, không phải sau khi đã ghi.
		if frappe.db.get_value("User", email, "user_type") == "System User":
			errors.append(
				f'"{email}" là tài khoản nội bộ (System User) trên hệ thống Miyano, '
				"không cấp làm tài khoản cổng của bệnh viện được."
			)
			dong["trang_thai"] = TU_CHOI
			continue
		khach_dang_thuoc = frappe.db.get_value("Portal Member", {"user": email}, "customer")
		if not khach_dang_thuoc:
			# `Portal Member` là nguồn DANH TÍNH cổng, nhưng không phải thứ duy
			# nhất cấp quyền đọc: Frappe OR các `User Permission` cùng doctype
			# lại với nhau, nên một tài khoản đã có `User Permission` trỏ về
			# bệnh viện KHÁC mà chưa có `Portal Member` sẽ nhận quyền của HAI
			# bệnh viện nếu cấp tiếp ở đây. Soi cả hai trục, không chỉ trục
			# danh tính.
			khach_dang_thuoc = frappe.db.get_value(
				"User Permission",
				{"user": email, "allow": "Customer", "for_value": ["!=", customer]},
				"for_value",
			)
		if khach_dang_thuoc and khach_dang_thuoc != customer:
			# QĐ-G23 — có thể là người thật làm ở hai nơi, cũng có thể là gõ
			# nhầm. KHÔNG tạo lại, KHÔNG đổi mật khẩu của họ, không chặn các
			# dòng khác: báo và để Miyano quyết.
			dong["trang_thai"] = CANH_BAO
			errors.append(
				f'Email này đã thuộc khách hàng "{khach_dang_thuoc}" — không tạo lại và '
				"không đổi mật khẩu của họ. Kiểm tra lại: gõ nhầm, hay đúng là một người "
				"làm ở hai nơi?"
			)
			continue
		if khach_dang_thuoc == customer:
			# Nói ĐÚNG năng lực thật của màn này: nó chỉ TẠO MỚI. Đổi khoa cho
			# một nhân viên, bật lại tài khoản đã tắt, hay sửa vai trò gõ nhầm
			# đều KHÔNG làm được ở đây — câu "bỏ qua, không đụng tới" một mình
			# đọc như "không có gì phải làm", và người nhập sẽ tưởng tệp vừa
			# cập nhật xong khoa phòng cho cả viện.
			dong["trang_thai"] = BO_QUA
			# QĐ-4 (Task 10): tài khoản ĐÃ CÓ chỉ được ĐIỀN VÀO CHỖ TRỐNG, KHÔNG
			# BAO GIỜ ĐÈ. Hôm nay gần như mọi tài khoản đang chạy đều chưa có
			# số — nếu chỉ đặt số cho tài khoản MỚI thì toàn bộ người dùng hiện
			# hữu vẫn không có số và tính năng gần như vô dụng ngay ngày bật.
			mobile_hien_co = _norm(frappe.db.get_value("User", email, "mobile_no"))
			if dien_thoai and not mobile_hien_co:
				# VÒNG SỬA 2: nhánh BỔ SUNG này ghi thẳng vào `User.mobile_no`
				# (UNIQUE) ở `_ghi()` — soi trùng TRƯỚC, cùng hàm dùng cho nhánh
				# TẠO MỚI bên dưới, vì cả hai cùng đụng một cột.
				dien_thoai = _kiem_trung_dien_thoai(
					dong, dien_thoai, email, so_da_dung, canh_bao_toan_tep, dang_tao_moi=False,
				)
				dong["dien_thoai"] = dien_thoai
				if dien_thoai:
					_them_ghi_chu(dong, (
						f'Đã có tài khoản ở bệnh viện này — bỏ qua tạo mới, nhưng số '
						f'điện thoại "{dien_thoai}" trong tệp sẽ được BỔ SUNG vào tài '
						"khoản (đang chưa có số)."
					))
				# else: `_kiem_trung_dien_thoai` đã tự ghi chú lý do không gán.
			elif dien_thoai and mobile_hien_co and mobile_hien_co != dien_thoai:
				# Lệch số: có thể người thật đổi số, cũng có thể gõ nhầm. Đây LÀ
				# dữ liệu người khác đã nhập — im lặng ghi đè là mất dữ liệu, nên
				# nâng thành CANH_BAO nêu cả hai số để Miyano tự xử, không tự ý
				# chọn số nào.
				dong["trang_thai"] = CANH_BAO
				errors.append(
					f'Số điện thoại trong tệp ("{dien_thoai}") khác với số đang lưu '
					f'trên tài khoản ("{mobile_hien_co}") — không tự động ghi đè. '
					"Kiểm tra lại rồi tự cập nhật trên Portal Member/User nếu cần."
				)
			else:
				_them_ghi_chu(dong, (
					"Đã có tài khoản ở bệnh viện này — bỏ qua. Màn này KHÔNG sửa được "
					"khoa phòng, vai trò hay trạng thái của tài khoản đã có: sửa trực "
					"tiếp trên bản ghi Portal Member."
				))
			continue

		if frappe.db.get_value("User", email, "enabled") == 0:
			# Gắn vào bệnh viện thì được, nhưng người ta sẽ KHÔNG đăng nhập
			# được — báo ngay ở xem trước, đừng để họ đi tìm nguyên nhân.
			_them_ghi_chu(dong, (
				"Tài khoản này đang bị VÔ HIỆU HOÁ trên hệ thống — bật lại "
				"(User → Enabled) thì mới đăng nhập được."
			))

		# --- (c) chốt nghiệp vụ --------------------------------------------
		if vai_tro == QUAN_LY:
			if ten_khoa or ma_khoa:
				errors.append(
					"Quản lý nhìn xuyên mọi khoa nên không gắn vào khoa phòng nào. "
					"Bỏ trống cột Khoa và Mã khoa."
				)
			elif quan_ly_hien_co:
				errors.append(
					f"Bệnh viện này đã có quản lý là {quan_ly_hien_co}. Tắt thành viên đó "
					f'trước, hoặc đặt tài khoản này là "{NHAN_VIEN_KHOA}".'
				)
			elif quan_ly_trong_tep:
				errors.append(
					f"Tệp đã có một Quản lý ở dòng {quan_ly_trong_tep['line']} "
					f"({quan_ly_trong_tep['email']}) — mỗi bệnh viện chỉ một quản lý."
				)
			else:
				quan_ly_trong_tep = dong
		else:
			if not co_ma_ngan:
				errors.append(
					f'Khách hàng "{customer}" chưa có Mã ngắn — đặt Mã ngắn cho bệnh viện '
					"trước khi cấp tài khoản theo khoa phòng."
				)
			if not ten_khoa and not ma_khoa:
				errors.append(
					f"{NHAN_VIEN_KHOA} phải được gán một khoa phòng — điền cột Khoa "
					"(và Mã khoa nếu có)."
				)
			else:
				da_co, loi_khoa = _tim_khoa_da_co(khoa_hien_co, ten_khoa, ma_khoa)
				if loi_khoa:
					errors.append(loi_khoa)
				elif da_co:
					dong["khoa"] = da_co["name"]
					dong["ten_khoa"] = da_co["ten_khoa_phong"]
					if not da_co.get("active"):
						_them_ghi_chu(
							dong,
							f'Khoa "{da_co["ten_khoa_phong"]}" đang TẮT — bật lại nếu vẫn dùng.',
						)
				else:
					errors.extend(_ghi_nhan_khoa_moi(dong, khoa_se_tao, ten_khoa, ma_khoa))

		if errors:
			dong["trang_thai"] = TU_CHOI
		else:
			# VÒNG SỬA 2: soi trùng TRƯỚC KHI GHI — dòng này sắp thành TAO_MOI,
			# và `_ghi()` sẽ đặt thẳng `dien_thoai` vào `User.mobile_no` (UNIQUE)
			# lúc tạo tài khoản. Chỉ soi khi dòng CÓ số — dòng trống thì rơi
			# thẳng xuống nhánh "Không có số điện thoại" bên dưới như cũ.
			if dien_thoai:
				dien_thoai = _kiem_trung_dien_thoai(
					dong, dien_thoai, email, so_da_dung, canh_bao_toan_tep, dang_tao_moi=True,
				)
				dong["dien_thoai"] = dien_thoai
			if not dien_thoai:
				# VÒNG SỬA 1 (chủ đầu tư, 04/09/2026): brief gốc bảo dùng CANH_BAO
				# cho ô trống, và CANH_BAO trong file này SẴN mang nghĩa "gắn cờ VÀ
				# KHÔNG GHI" (`_ghi()`: `if dong["trang_thai"] != TAO_MOI: continue`)
				# — hai chỗ dùng CANH_BAO trước đó (email thuộc bệnh viện khác; số
				# cũ khác số mới) ĐÚNG là không nên ghi. Nhưng thiếu số điện thoại
				# KHÔNG phải lý do để hoãn cấp tài khoản: chủ đầu tư chọn tường
				# minh là VẪN TẠO, chỉ cảnh báo để Miyano thấy ai còn thiếu mà đi
				# xin bổ sung. Nên GIỮ `trang_thai = TAO_MOI` (không đổi), đưa cảnh
				# báo vào `ghi_chu` (không phải `errors` — `errors` gắn với
				# TU_CHOI/CANH_BAO ở nơi khác trong file, dùng nó ở đây sẽ đọc nhầm
				# là "có gì đó chặn dòng này"). Mục `canh_bao_toan_tep` (sau vòng
				# lặp) mới là nơi liệt kê TÊN những người thiếu số — không có nó,
				# "cảnh báo" chỉ là chữ nằm im trong từng dòng, người duyệt tệp
				# phải dò từng dòng mới thấy.
				#
				# VÒNG SỬA 2: nếu số vừa bị XOÁ vì TRÙNG (không phải vốn dĩ để
				# trống), `_kiem_trung_dien_thoai` đã tự ghi chú lý do cụ thể rồi
				# — không nối thêm câu "Không có số điện thoại" chung chung đè lên,
				# tránh một dòng mang hai lời giải thích khác nhau cho cùng một sự
				# thật. Cờ `_trung_dien_thoai` (đặt bên dưới) còn được dùng để loại
				# dòng này khỏi mục tổng "thiếu số điện thoại" cấp tệp — dòng này
				# KHÔNG "thiếu" theo nghĩa người điền quên, nó bị soi trùng.
				if dong.get("_trung_dien_thoai"):
					pass
				else:
					_them_ghi_chu(dong, (
						"Không có số điện thoại — khách sẽ thấy tên nhưng không gọi được."
					))

	if not co_ma_ngan and any(
		r["vai_tro"] == NHAN_VIEN_KHOA and r["trang_thai"] in (TAO_MOI, TU_CHOI) for r in rows
	):
		loi_toan_tep.append(
			f'Khách hàng "{customer}" chưa có Mã ngắn. Đặt Mã ngắn trên hồ sơ khách hàng '
			"rồi nhập lại — mã ngắn đi vào tên phiếu Đề nghị mua của khoa phòng."
		)
	se_co_quan_ly = quan_ly_trong_tep is not None and quan_ly_trong_tep["trang_thai"] == TAO_MOI
	if not quan_ly_hien_co and not se_co_quan_ly:
		canh_bao_toan_tep.append(
			"Tệp này không tạo Quản lý nào và bệnh viện cũng chưa có quản lý đang hoạt "
			"động — sẽ không ai duyệt được phiếu Đề nghị mua của các khoa."
		)
	# VÒNG SỬA 1: liệt kê TÊN người thiếu số ở CẤP TỆP — đây là thứ biến "cảnh
	# báo" ở từng dòng (`ghi_chu`) thành việc làm được: người duyệt tệp thấy
	# ngay ở đầu màn ai còn thiếu số mà đi xin bổ sung, không phải dò từng dòng
	# trong bảng có thể dài 200 dòng.
	# VÒNG SỬA 2: loại các dòng bị BỎ TRỐNG VÌ TRÙNG khỏi mục "thiếu số điện
	# thoại" — dòng đó không "thiếu" theo nghĩa người điền quên, nó đã có số
	# trong tệp nhưng bị soi trùng; lý do đã nằm riêng trong `canh_bao_toan_tep`
	# ở `_kiem_trung_dien_thoai`. Gộp chung hai lý do khác nhau vào một câu
	# chung chung sẽ đọc sai bản chất với người duyệt tệp.
	thieu_dien_thoai = [
		r for r in rows
		if r["trang_thai"] == TAO_MOI and not r["dien_thoai"] and not r.get("_trung_dien_thoai")
	]
	if thieu_dien_thoai:
		canh_bao_toan_tep.append(
			f"{len(thieu_dien_thoai)} người chưa có số điện thoại trong tệp — vẫn được tạo "
			"tài khoản, nhưng dòng thời gian đơn hàng sẽ chỉ hiện tên, không gọi được: "
			+ ", ".join(f'{r["ho_ten"]} ({r["email"]})' for r in thieu_dien_thoai) + "."
		)

	dem = {trang_thai: 0 for trang_thai in (TAO_MOI, BO_QUA, CANH_BAO, TU_CHOI)}
	for row in rows:
		dem[row["trang_thai"]] += 1
	# Khoa mới chỉ còn ý nghĩa nếu dòng dẫn tới nó không bị từ chối vì lý do khác.
	khoa_se_tao = [k for k in khoa_se_tao if any(
		r["trang_thai"] == TAO_MOI and r["khoa_moi"] and r["khoa_key"] == k["key"] for r in rows
	)]
	return {
		"customer": customer,
		"total": len(rows),
		"so_tao_moi": dem[TAO_MOI],
		"so_bo_qua": dem[BO_QUA],
		"so_canh_bao": dem[CANH_BAO],
		"so_tu_choi": dem[TU_CHOI],
		"rows": rows,
		"khoa_se_tao": khoa_se_tao,
		"loi_toan_tep": loi_toan_tep,
		"canh_bao_toan_tep": canh_bao_toan_tep,
	}


def _ghi_nhan_khoa_moi(dong: dict, khoa_se_tao: list[dict], ten: str, ma: str) -> list[str]:
	"""QĐ-G20: khoa chưa có thì tự tạo — nhưng phải NÓI RÕ ở bước xem trước.

	Gộp các dòng cùng trỏ về một khoa mới (theo mã, hoặc theo tên không dấu)
	để tệp có 30 người của khoa Huyết học không đẻ ra 30 khoa.
	"""
	errors: list[str] = []
	if not ten:
		return [
			f'Mã khoa "{ma}" chưa có trong danh mục khoa phòng của bệnh viện. Điền thêm '
			"cột Khoa (tên khoa) nếu muốn tạo mới."
		]
	if ma:
		loi_ma = _kiem_ma_khoa(ma)
		if loi_ma:
			return [loi_ma]
	for muc in khoa_se_tao:
		trung_ma = bool(ma) and muc["ma_khoa"] == ma
		trung_ten = similarity.la_trung_tuyet_doi(ten, muc["ten_khoa_phong"])
		if trung_ma or trung_ten:
			if trung_ma and not trung_ten:
				errors.append(
					f'Mã khoa "{ma}" đang dùng cho hai tên khoa khác nhau trong cùng tệp: '
					f'"{muc["ten_khoa_phong"]}" và "{ten}".'
				)
				return errors
			dong["khoa_moi"] = True
			dong["khoa_key"] = muc["key"]
			return errors
	key = f"{len(khoa_se_tao)}|{ma or similarity.khong_dau(ten)}"
	khoa_se_tao.append({"key": key, "ten_khoa_phong": ten, "ma_khoa": ma})
	dong["khoa_moi"] = True
	dong["khoa_key"] = key
	return errors


def _cell(row_cells, col: int):
	idx = col - 1
	return row_cells[idx].value if idx < len(row_cells) else None


# ---------------------------------------------------------------------------
# Bước 3 — ghi thật
# ---------------------------------------------------------------------------


class _MatKhauTho:
	"""Mật khẩu thô, KHÔNG BAO GIỜ tự in ra mình.

	Vì sao phải có một kiểu riêng thay vì chỉ đặt tên biến cho khéo — đo được,
	không phải suy đoán (vòng sửa 1, review bảo mật độc lập):

	* `frappe.get_traceback(with_context=True)` — đường mà `log_error()` và
	  `log_error_snapshot()` (frappe/app.py, bắn với MỌI mã HTTP ≥ 500) đi qua
	  — **kết xuất biến cục bộ của mọi khung ngăn xếp**, in bằng `repr()`
	  (traceback_with_variables/core.py:187).
	* Bộ khử của Frappe (`_get_traceback_sanitizer`) che theo **TÊN BIẾN**,
	  bằng `re.search` với `password|passwd|secret|token|key|pwd`, và với dict
	  thì chỉ che **KHOÁ** nằm trong danh sách đó. Bản đồ `{email: mật khẩu}`
	  có khoá là email nên **in ra nguyên văn**.
	* `tabError Log` là MyISAM: dòng đó **sống sót qua rollback** — giao dịch
	  cuộn ngược xoá sạch tài khoản, nhưng mật khẩu thì nằm lại vĩnh viễn.

	Đặt tên biến là `pwd`/`pwd_map` thì bộ khử theo tên cũng che được, nhưng
	cách đó mong manh: một lần đổi tên vô hại của người sau là mở lại đúng lỗ
	này, và không có test nào của họ nói cho họ biết. Ở đây **kiểu dữ liệu**
	gánh việc đó — `repr()` không bao giờ trả ra bí mật, bất kể biến tên gì,
	nằm trong dict/list/tuple nào, ở khung của ai.

	**Và vì thế các biến ở đây CỐ Ý mang tên tiếng Việt không khớp danh sách
	chặn** (`mat_khau`, `bang_mat_khau` — không chứa `password|passwd|secret|
	token|key|pwd`). Nghe ngược đời, nhưng đã đo: bản đầu của vòng sửa này đặt
	tên `pwd`/`pwd_map`, và khi thử **phá `__repr__`** cho nó in thẳng bí mật
	ra, `test_mat_khau_tho_khong_lo_ra_trong_traceback` **vẫn XANH** — bộ khử
	theo tên đã che hộ, tức bài test đang canh CÁI TÊN chứ không canh cái lớp
	này. Hai lớp chồng nhau nghe thì an toàn hơn, nhưng lớp ngoài che mất hồi
	quy của lớp trong, và một chốt không test được là một chốt sẽ mục lặng lẽ.
	Một lớp DUY NHẤT, mạnh hơn, và có test thật sự phân biệt được — đổi tên
	biến ở đây thành `pwd_*` là làm bài test đó mù.

	Bí mật chỉ hiện hình khi ai đó gọi `lo_ra()` — một chỗ duy nhất để đọc, và
	trong file này nó được gọi ở ĐÚNG câu lệnh cuối cùng của `_ghi()`.
	"""

	__slots__ = ("_bi_mat",)

	def __init__(self, bi_mat: str):
		self._bi_mat = bi_mat

	def lo_ra(self) -> str:
		return self._bi_mat

	def __repr__(self) -> str:
		return "********"

	__str__ = __repr__

	def __format__(self, spec) -> str:
		return "********"


def _sinh_mat_khau() -> _MatKhauTho:
	"""Mật khẩu bàn giao tay: 12 ký tự, bỏ các ký tự dễ đọc nhầm khi chép ra
	giấy (0/O, 1/l/I). `secrets` chứ không phải `random` — đây là bí mật đăng
	nhập, không phải một con số ngẫu nhiên cho vui.

	Dựng thẳng vào `_MatKhauTho(...)`: chuỗi thô không bao giờ là biến cục bộ
	có tên trong khung này."""
	bang = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
	return _MatKhauTho("".join(secrets.choice(bang) for _ in range(12)))


def _ghi(content: bytes, customer: str, file_url: str | None = None) -> dict:
	"""Đọc lại TỪ ĐẦU trên server (không tin dữ liệu client gửi), rồi ghi.

	Dòng BỊ TỪ CHỐI khiến KHÔNG GÌ được ghi — tất-cả-hoặc-không, cùng khuôn
	`import_ton_dau.commit_workbook`. Dòng CẢNH BÁO (QĐ-G23) thì không chặn ai:
	nó được nhắc lại nguyên vẹn trong kết quả để cờ đó không rơi mất giữa hai
	bước.
	"""
	ket_qua = _phan_tich(content, customer)
	if ket_qua["loi_toan_tep"]:
		# Fail-closed. Hôm nay nguồn DUY NHẤT của `loi_toan_tep` (thiếu Mã
		# ngắn) cũng từ chối mọi dòng liên quan, nên nhánh này chưa với tới
		# được — nó đứng đây để phép kiểm CẤP TỆP tiếp theo không lặng lẽ
		# trôi qua bước ghi chỉ vì nó không nói gì về một dòng cụ thể.
		frappe.throw(
			" ".join(ket_qua["loi_toan_tep"]) + " Chưa có dữ liệu nào được ghi.",
			frappe.ValidationError,
		)
	if ket_qua["so_tu_choi"]:
		dau_tien = next(r for r in ket_qua["rows"] if r["trang_thai"] == TU_CHOI)
		frappe.throw(
			f"Tệp có {ket_qua['so_tu_choi']} dòng bị từ chối trong tổng số "
			f"{ket_qua['total']} dòng (ví dụ dòng {dau_tien['line']}: "
			f"{'; '.join(dau_tien['errors'])}). Vui lòng sửa và tải lại — "
			"chưa có dữ liệu nào được ghi.",
			frappe.ValidationError,
		)

	# `bang_mat_khau` giữ `_MatKhauTho`, KHÔNG giữ chuỗi thô — xem docstring của
	# lớp đó (kể cả đoạn giải thích vì sao biến ở đây KHÔNG được đặt tên
	# `pwd_*`): mọi khung dưới đây đều có thể ném, và mỗi lần ném là một lần
	# `log_error_snapshot()` kết xuất biến cục bộ vào `tabError Log`.
	bang_mat_khau: dict[str, _MatKhauTho] = {}
	khoa_da_tao: list[dict] = []
	sp = "nhan_su_import_commit_sp"
	frappe.db.savepoint(sp)
	try:
		key_to_khoa: dict[str, str] = {}
		for muc in ket_qua["khoa_se_tao"]:
			kp = frappe.get_doc({
				"doctype": "Customer Department", "customer": customer,
				"ten_khoa_phong": muc["ten_khoa_phong"], "ma_khoa": muc["ma_khoa"] or None,
				"active": 1,
			})
			kp.insert(ignore_permissions=True)
			key_to_khoa[muc["key"]] = kp.name
			khoa_da_tao.append({
				"name": kp.name, "ten_khoa_phong": kp.ten_khoa_phong, "ma_khoa": kp.ma_khoa or "",
			})

		for dong in ket_qua["rows"]:
			if dong["trang_thai"] == BO_QUA:
				# QĐ-4 (Task 10): đọc lại TỪ ĐẦU ngay tại lúc ghi (không tin
				# verdict của `_phan_tich()` đã tính vài dòng lệnh trước — cùng
				# triết lý "đọc lại từ đầu trên server" của cả hàm này). Số
				# khác đã bị `_phan_tich()` nâng thành CANH_BAO ở trên rồi nên
				# KHÔNG rơi vào đây — tới được đây nghĩa là hoặc đang trống
				# (điền), hoặc đã khớp sẵn (set lại giá trị y hệt, vô hại).
				if dong["dien_thoai"] and not frappe.db.get_value(
					"User", dong["email"], "mobile_no"
				):
					frappe.db.set_value(
						"User", dong["email"], "mobile_no", dong["dien_thoai"],
						update_modified=False,
					)
				continue
			if dong["trang_thai"] != TAO_MOI:
				continue
			khoa = dong["khoa"] or key_to_khoa.get(dong["khoa_key"])
			# Tài khoản ĐÃ TỒN TẠI (User có sẵn nhưng chưa thuộc bệnh viện
			# nào) chỉ được GẮN vào bệnh viện này — không đặt lại mật khẩu của
			# một người đang dùng tài khoản đó cho việc khác.
			la_nguoi_moi = not frappe.db.exists("User", dong["email"])
			portal_provision(
				customer, dong["email"], send_invite=False,
				first_name=dong["ho_ten"], vai_tro=dong["vai_tro"], khoa_phong=khoa,
				dien_thoai=dong["dien_thoai"] or None,
			)
			dong["khoa"] = khoa
			if la_nguoi_moi:
				mat_khau = _sinh_mat_khau()
				# `.lo_ra()` ngay trong biểu thức lời gọi: chuỗi thô không trở
				# thành biến cục bộ của khung này. Ở khung của `update_password`
				# nó mang tên `pwd` — trùng danh sách chặn theo tên của Frappe.
				update_password(dong["email"], mat_khau.lo_ra())
				# "Bắt đổi ở lần đăng nhập đầu" (QĐ-G19): bản Frappe này chỉ
				# có chính sách theo SỐ NGÀY ở System Settings
				# (`force_user_to_reset_password`), không có cờ cho từng
				# người — đặt mốc đổi mật khẩu về quá khứ để chính sách đó
				# (khi bật) chộp đúng các tài khoản này ngay lần đăng nhập
				# đầu. Phải đặt SAU `update_password`.
				frappe.db.set_value(
					"User", dong["email"], "last_password_reset_date", "2000-01-01",
					update_modified=False,
				)
				bang_mat_khau[dong["email"]] = mat_khau
			else:
				_them_ghi_chu(dong, (
					"Tài khoản đã tồn tại từ trước — chỉ gắn vào bệnh viện này, "
					"không đặt lại mật khẩu."
				))
	except Exception:
		frappe.db.rollback(save_point=sp)
		raise

	ket_qua["khoa_da_tao"] = khoa_da_tao
	ket_qua.pop("khoa_se_tao", None)
	# Ghi xong thì tệp hết việc: để lại là để một danh sách nhân sự đầy đủ nằm
	# vĩnh viễn trên đĩa. Xoá TRƯỚC câu lệnh bung mật khẩu bên dưới — hàm xoá
	# có thể ném, và không được ném khi khung này đang cầm mật khẩu thô.
	if file_url:
		_xoa_tep_da_nhap(file_url)
	# CÂU LỆNH CUỐI CÙNG và không có gì ném được giữa đây với `return`: đây là
	# chỗ duy nhất trong file mật khẩu thô hiện hình.
	ket_qua["mat_khau"] = {email: p.lo_ra() for email, p in bang_mat_khau.items()}
	return ket_qua


def _xoa_tep_da_nhap(file_url: str) -> None:
	"""Xoá tệp nhân sự sau khi đã ghi xong. Nuốt mọi lỗi CÓ CHỦ Ý: tệp còn sót
	lại là chuyện vệ sinh, còn một lỗi ném lên từ đây sẽ (a) làm hỏng một lần
	nhập ĐÃ THÀNH CÔNG và (b) kéo theo một lần kết xuất biến cục bộ của khung
	gọi. Bắt lỗi NGAY TRONG hàm này để traceback (nếu có ai log) chỉ chứa các
	khung từ đây trở xuống."""
	try:
		name = frappe.db.get_value("File", {"file_url": file_url}, "name")
		if name:
			frappe.delete_doc("File", name, ignore_permissions=True, force=True)
	except Exception:
		pass


# ---------------------------------------------------------------------------
# Ba endpoint của màn Desk
# ---------------------------------------------------------------------------


@frappe.whitelist()
def nhan_su_import_template() -> None:
	"""Tải bảng Excel để nhân viên bệnh viện điền."""
	chan_neu_khong_phai_nhan_vien_miyano()
	frappe.local.response.filename = "mau_nhap_nhan_su_benh_vien.xlsx"
	frappe.local.response.filecontent = build_template_bytes()
	frappe.local.response.type = "download"
	frappe.local.response.content_type = (
		"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	)


@frappe.whitelist()
def nhan_su_import_preview(customer, file_url) -> dict:
	"""Đọc và phân tích tệp, KHÔNG GHI GÌ. Xem `_phan_tich`."""
	chan_neu_khong_phai_nhan_vien_miyano()
	customer = _kiem_khach(customer)
	_kiem_tep_rieng_tu(file_url)
	return _phan_tich(_resolve_owned_spreadsheet(file_url), customer)


@frappe.whitelist()
def nhan_su_import_commit(customer, file_url) -> dict:
	"""Đọc lại VÀ kiểm lại từ đầu ở server rồi mới ghi — không tin bất kỳ dòng
	dữ liệu nào mà client (đã gọi xem trước trước đó) có thể gửi kèm.

	Khoá `mat_khau` trong kết quả là thứ DUY NHẤT mang mật khẩu vừa đặt, và
	nó chỉ đi đúng một chuyến về màn hình đang mở (QĐ-G19)."""
	chan_neu_khong_phai_nhan_vien_miyano()
	customer = _kiem_khach(customer)
	_kiem_tep_rieng_tu(file_url)
	return _ghi(_resolve_owned_spreadsheet(file_url), customer, file_url)
